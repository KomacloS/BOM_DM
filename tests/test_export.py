# root: tests/test_export.py
from fastapi.testclient import TestClient
from sqlmodel import SQLModel, create_engine
import sqlalchemy
import pytest
import os
import io
from datetime import datetime
from openpyxl import load_workbook
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import app.main as main


@pytest.fixture(name="client")
def client_fixture():
    test_engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=sqlalchemy.pool.StaticPool,
    )
    main.engine = test_engine
    SQLModel.metadata.create_all(test_engine)
    with TestClient(main.app) as c:
        yield c


@pytest.fixture
def admin_header(client):
    token = client.post(
        "/auth/token", data={"username": "admin", "password": "123456789"}
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
def user_header(client, admin_header):
    client.post(
        "/auth/register",
        json={"username": "u1", "password": "pw", "role": "viewer"},
        headers=admin_header,
    )
    token = client.post(
        "/auth/token", data={"username": "u1", "password": "pw"}
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def seed_data(client, admin_header):
    client.post(
        "/bom/items",
        json={"part_number": "P1", "description": "A", "quantity": 2},
        headers=admin_header,
    )
    client.post(
        "/testresults",
        json={"serial_number": "SN1", "result": True},
        headers=admin_header,
    )


def test_admin_can_export(client, admin_header, tmp_path):
    seed_data(client, admin_header)
    resp = client.get("/export/bom.csv", headers=admin_header)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    first_line = resp.text.splitlines()[0]
    assert "part_number" in first_line

    resp2 = client.get("/export/testresults.xlsx", headers=admin_header)
    assert resp2.status_code == 200
    wb = load_workbook(filename=io.BytesIO(resp2.content))
    assert "results" in wb.sheetnames
    ws = wb["results"]
    assert ws.max_row >= 2

    # manual backup
    main.nightly_backup(dest=tmp_path)
    stamp = datetime.utcnow().strftime("%Y%m%d")
    assert (tmp_path / f"bom_{stamp}.csv").exists()
    assert (tmp_path / f"testresults_{stamp}.xlsx").exists()


def test_non_admin_forbidden(client, user_header):
    r1 = client.get("/export/bom.csv", headers=user_header)
    r2 = client.get("/export/testresults.xlsx", headers=user_header)
    assert r1.status_code == 403
    assert r2.status_code == 403
