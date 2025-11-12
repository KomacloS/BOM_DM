"""BOM editor for classifying parts as active or passive.

Features implemented:
- Clickable pill switch (no combobox) cycling empty â†’ passive â†’ active â†’ passive â†’ â€¦
- Two view modes: By PN (grouped) and By Reference (flat)
- Auto-passive inference for R/L/C when value is empty (UI-only)
- Column visibility per-mode, persisted in QSettings
- Filter works across visible columns; natural sort for reference columns
"""

from __future__ import annotations

from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple
from pathlib import Path
import json
import os

from PyQt6.QtGui import (
    QAction,
    QStandardItemModel,
    QStandardItem,
    QIcon,
    QKeySequence,
    QUndoStack,
    QUndoCommand,
)  # QAction/QStandardItem* are in QtGui
from PyQt6.QtCore import (
    Qt,
    QSortFilterProxyModel,
    QModelIndex,
    pyqtSignal,
    QSettings,
    QRect,
    QSize,
    QTimer,
    QUrl,
    QRectF,
    QPersistentModelIndex,
)
from PyQt6.QtWidgets import (
    QWidget,
    QTableView,
    QVBoxLayout,
    QLineEdit,
    QPushButton,
    QToolBar,
    QMenu,
    QToolButton,
    QStyle,
    QApplication,
    QStyledItemDelegate,
    QHeaderView,
    QAbstractItemView,
    QStyleOptionViewItem,
    QLabel,
    QMessageBox,
    QSpinBox,
    QComboBox,
    QFileDialog,
    QInputDialog,
    QStyleOptionButton,
    QSplitter,
    QDialog,
    QDialogButtonBox,
    QHBoxLayout,
    QTableWidget,
    QTableWidgetItem,
)
from PyQt6.QtGui import QPainter, QBrush, QColor, QDesktopServices, QGuiApplication, QTextDocument, QTextOption
import logging
from .. import services
from ..services.datasheets import get_local_open_path
from ..config import (
    DATA_ROOT,
    LOG_DIR,
    PDF_VIEWER,
    PDF_VIEWER_PATH,
    PDF_OPEN_DEBUG,
    get_complex_editor_settings,
    get_viva_export_settings,
    save_viva_export_settings,
)
from ..integration import ce_bridge_client
from ..integration.ce_bridge_client import (
    CEAuthError,
    CENetworkError,
    CEExportBusyError,
    CEExportError,
    CEExportStrictError,
    CEPNResolutionError,
)
import subprocess, shutil, time, os
from datetime import datetime
from functools import partial
from ..logic.autofill_rules import infer_from_pn_and_desc
from ..logic.prefix_macros import load_prefix_macros
from ..models import Assembly, TestMode, TestProfile
from . import state as app_state
from .widgets.complex_panel import ComplexPanel
from .widgets.schematic_panel import SchematicPanel


PartIdRole = Qt.ItemDataRole.UserRole + 1
ModeRole = Qt.ItemDataRole.UserRole + 2  # stores 'active'/'passive'/None in AP column for delegate
DatasheetRole = Qt.ItemDataRole.UserRole + 3
"""
Additional per-cell data roles
- BOMItemIdRole: primary key of BOMItem in by_ref view for per-row actions.
- LinkUrlRole: stores the actual product link while display text stays empty.
"""
BOMItemIdRole = Qt.ItemDataRole.UserRole + 4
LinkUrlRole = Qt.ItemDataRole.UserRole + 5


# Helpers ---------------------------------------------------------------
def natural_key(s: str) -> List[object]:
    """Natural sort key: split numbers out of strings (case-insensitive)."""
    import re

    token = re.compile(r"(\d+)")
    return [int(t) if t.isdigit() else t.lower() for t in token.split(s or "")]


class BOMFilterProxy(QSortFilterProxyModel):
    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._filter = ""
        self._ref_col: Optional[int] = None  # column index for natural sorting of references
        self._skip_cols: set[int] = set()

    def setFilterString(self, text: str) -> None:
        self._filter = text.lower()
        self.invalidateFilter()

    def setReferenceColumn(self, col: Optional[int]) -> None:
        self._ref_col = col
        self.invalidate()

    def setSkipColumns(self, cols: set[int]) -> None:
        self._skip_cols = set(cols)
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row: int, source_parent):  # pragma: no cover - Qt glue
        if not self._filter:
            return True
        model = self.sourceModel()
        cols = model.columnCount()
        for col in range(cols):
            if col in self._skip_cols:
                continue
            idx = model.index(source_row, col, source_parent)
            data = model.data(idx)
            if data and self._filter in str(data).lower():
                return True
        return False

    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:  # pragma: no cover - UI glue
        if self._ref_col is not None and left.column() == self._ref_col == right.column():
            l = (self.sourceModel().data(left) or "")
            r = (self.sourceModel().data(right) or "")
            return natural_key(l) < natural_key(r)
        return super().lessThan(left, right)


class NoWheelComboBox(QComboBox):
    """QComboBox that ignores wheel events to prevent accidental changes."""

    def wheelEvent(self, event):  # pragma: no cover - Qt glue
        event.ignore()


class NoWheelSpinBox(QSpinBox):
    """QSpinBox that ignores wheel events to prevent accidental changes."""

    def wheelEvent(self, event):  # pragma: no cover - Qt glue
        event.ignore()


class UndoableStandardItemModel(QStandardItemModel):
    """QStandardItemModel that emits old/new values on changes."""

    changed = pyqtSignal(QModelIndex, object, object)

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if isinstance(index, QPersistentModelIndex):
            index = QModelIndex(index)
        if role == Qt.ItemDataRole.EditRole:
            old = super().data(index, role)
            res = super().setData(index, value, role)
            if res and old != value:
                self.changed.emit(index, old, value)
            return res
        return super().setData(index, value, role)


class SetCellCommand(QUndoCommand):
    """Undo command encapsulating a single cell edit."""

    def __init__(self, pane: "BOMEditorPane", index: QModelIndex, old: object, new: object) -> None:
        super().__init__("Set Cell")
        self.pane = pane
        self.index = QPersistentModelIndex(index)
        self.old = old
        self.new = new
        # When first pushed the model already contains ``new``.  Track this so
        # that the initial ``redo`` executed by :class:`QUndoStack` is a no-op
        # and does not trigger change handlers again.
        self._first = True

    def undo(self) -> None:  # pragma: no cover - Qt glue
        model = self.pane.model
        self.pane._updating = True
        model.setData(QModelIndex(self.index), self.old)
        self.pane._updating = False

    def redo(self) -> None:  # pragma: no cover - Qt glue
        if self._first:
            self._first = False
            return
        model = self.pane.model
        self.pane._updating = True
        model.setData(QModelIndex(self.index), self.new)
        self.pane._updating = False


class CycleToggleDelegate(QStyledItemDelegate):
    """Delegate that draws a small pill and toggles value on click.

    Values: None â†’ 'passive' â†’ 'active' â†’ 'passive' â†’ â€¦
    """

    valueChanged = pyqtSignal(int, object)  # part_id, new_value (str or None)

    def paint(self, painter: QPainter, option, index):  # pragma: no cover - UI glue
        opt = QStyleOptionViewItem(option)
        opt.state &= ~QStyle.StateFlag.State_HasFocus

        # Fill the background based on state so it matches QSS selection/hover
        if opt.state & QStyle.StateFlag.State_Selected:
            painter.fillRect(opt.rect, QColor("#DCEBFF"))
        elif opt.state & QStyle.StateFlag.State_MouseOver:
            painter.fillRect(opt.rect, QColor("#F3F4F6"))
        else:
            painter.fillRect(opt.rect, opt.palette.base().color())

        value = index.data() or None
        rect: QRect = opt.rect.adjusted(6, 4, -6, -4)
        pal = opt.palette

        if value == "active":
            bg = pal.highlight().color()
            fg = pal.highlightedText().color()
            text = "Active"
        elif value == "passive":
            bg = pal.alternateBase().color()
            fg = pal.text().color()
            text = "Passive"
        else:
            bg = pal.window().color().lighter(110)
            fg = pal.mid().color()
            text = "—"

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setBrush(QBrush(bg))
        painter.setPen(QColor(bg).darker(120))
        painter.drawRoundedRect(rect, 10, 10)
        painter.setPen(fg)
        # Center text
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, text)
        painter.restore()

    def editorEvent(self, event, model, option, index):  # pragma: no cover - UI glue
        if event.type() == event.Type.MouseButtonRelease and event.button() == Qt.MouseButton.LeftButton:
            cur = index.data() or None
            new = self._next_state(cur)
            part_id = index.data(PartIdRole)
            # write to model first for immediate UI feedback
            model.setData(index, new)
            self.valueChanged.emit(part_id, new)
            return True
        return False

    @staticmethod
    def _next_state(cur):
        if cur in (None, ""):
            return "passive"
        if cur == "passive":
            return "active"
        if cur == "active":
            return "passive"
        return "passive"

    def createEditor(self, parent, option, index):  # pragma: no cover - no inline editor
        # No inline editor; clicks are handled in editorEvent
        return None


class TestMethodDelegate(QStyledItemDelegate):
    """Delegate providing a combobox for selecting test method."""

    methodChanged = pyqtSignal(int, str, int)  # part_id, method, column

    def createEditor(self, parent, option, index):  # pragma: no cover - Qt glue
        combo = NoWheelComboBox(parent)
        combo.addItems(["", "Macro", "Complex", "Quick test (QT)", "Python code"])
        combo.activated.connect(lambda _=0, w=combo: self._commit_close(w))
        return combo

    def setEditorData(self, editor: QComboBox, index):  # pragma: no cover - Qt glue
        value = index.data(int(Qt.ItemDataRole.EditRole))
        if value in (None, ""):
            value = index.data() or ""
        editor.setCurrentText(value)

    def setModelData(self, editor: QComboBox, model, index):  # pragma: no cover - Qt glue
        text = editor.currentText()
        model.setData(index, text, int(Qt.ItemDataRole.EditRole))
        part_id = index.data(PartIdRole)
        self.methodChanged.emit(part_id, text, index.column())

    # Do not force editor on single click; rely on double-click/Edit key
    def editorEvent(self, event, model, option, index):  # pragma: no cover - Qt glue
        return False

    def _commit_close(self, editor):  # pragma: no cover - Qt glue
        self.commitData.emit(editor)
        self.closeEditor.emit(editor)




class TestDetailDelegate(QStyledItemDelegate):
    """Delegate for Test Detail column.

    - When Test Method is 'Macro', shows a combobox with function options.
    - Otherwise renders a button-like area and emits detailClicked on click.
    """

    detailClicked = pyqtSignal(int, int)  # part_id, column
    detailChanged = pyqtSignal(int, object, int)  # part_id, new_detail (str|None), column

    def __init__(self, parent=None, options: list[str] | None = None, test_method_col: Optional[int] = None) -> None:
        super().__init__(parent)
        self._options = options or []
        self._default_method_col = test_method_col
        self._method_columns: dict[int, int] = {}
        if test_method_col is not None:
            self._method_columns[test_method_col] = test_method_col

    def set_test_method_col(self, col: int) -> None:
        self._default_method_col = col

    def set_method_column_map(self, mapping: dict[int, int]) -> None:
        self._method_columns = dict(mapping)

    def _method_for_index(self, index) -> str:
        try:
            model = index.model()  # proxy
            method_col = self._method_columns.get(index.column(), self._default_method_col)
            if method_col is None:
                return ""
            tm_idx = model.index(index.row(), method_col)
            value = model.data(tm_idx, int(Qt.ItemDataRole.EditRole))
            if value in (None, ""):
                value = model.data(tm_idx) or ""
            return value
        except Exception:
            return ""

    def paint(self, painter: QPainter, option, index):  # pragma: no cover - UI glue
        # Render as a normal text cell; editor appears on double-click
        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        style = opt.widget.style() if opt.widget else QApplication.style()
        style.drawControl(QStyle.ControlElement.CE_ItemViewItem, opt, painter, opt.widget)

    def editorEvent(self, event, model, option, index):  # pragma: no cover - Qt glue
        # Allow single-click action for non-Macro rows (Quick Test picker, Complex, etc.)
        if event.type() == event.Type.MouseButtonRelease and event.button() == Qt.MouseButton.LeftButton:
            method = (self._method_for_index(index) or "")
            if method != "Macro":
                part_id = index.data(PartIdRole)
                if part_id is not None:
                    self.detailClicked.emit(part_id, index.column())
                    return True
        return False

    def createEditor(self, parent, option, index):  # pragma: no cover - Qt glue
        method = (self._method_for_index(index) or "")
        if method != "Macro":
            return None
        combo = NoWheelComboBox(parent)
        items = [""] + list(self._options)
        combo.addItems(items)
        combo.activated.connect(lambda _=0, w=combo: self._commit_close(w))
        return combo

    def setEditorData(self, editor: QComboBox, index):  # pragma: no cover - Qt glue
        value = index.data(int(Qt.ItemDataRole.EditRole))
        if value in (None, ""):
            value = index.data() or ""
        editor.setCurrentText(value)

    def setModelData(self, editor: QComboBox, model, index):  # pragma: no cover - Qt glue
        text = (editor.currentText() or "").strip()
        model.setData(index, text, int(Qt.ItemDataRole.EditRole))
        part_id = index.data(PartIdRole)
        self.detailChanged.emit(part_id, text or None, index.column())

    def _commit_close(self, editor):  # pragma: no cover - Qt glue
        self.commitData.emit(editor)
        self.closeEditor.emit(editor)


class VivaStrictFailDialog(QDialog):
    """Modal dialog presenting PNs that blocked a VIVA export."""

    def __init__(
        self,
        parent: QWidget,
        error: CEExportStrictError,
        pn_map: Dict[str, int],
        open_part_cb,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Some PNs aren't linked to complexes")
        self._pn_map = pn_map
        self._open_part_cb = open_part_cb
        self._entries: list[tuple[str, str]] = []

        layout = QVBoxLayout(self)
        label = QLabel("Some PNs aren't linked to complexes", self)
        label.setWordWrap(True)
        layout.addWidget(label)

        self.table = QTableWidget(0, 3, self)
        self.table.setHorizontalHeaderLabels(["PN", "Status", ""])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)

        for status, pn in self._iter_entries(error):
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(pn))
            self.table.setItem(row, 1, QTableWidgetItem(status))
            self._entries.append((pn, status))
            part_id = pn_map.get(pn.lower())
            if part_id is not None:
                button = QPushButton("Open in CE", self)
                button.clicked.connect(partial(self._open_clicked, pn.lower()))
                self.table.setCellWidget(row, 2, button)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close, self)
        copy_button = buttons.addButton("Copy list", QDialogButtonBox.ButtonRole.ActionRole)
        copy_button.clicked.connect(self._copy_entries)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.resize(520, 320)

    def _iter_entries(self, error: CEExportStrictError) -> Iterable[tuple[str, str]]:
        for pn in error.unlinked:
            yield "Unlinked", str(pn)
        for pn in error.missing:
            yield "Missing", str(pn)

    def _open_clicked(self, pn_key: str) -> None:
        part_id = self._pn_map.get(pn_key)
        if part_id is None:
            return
        self._open_part_cb(part_id)

    def _copy_entries(self) -> None:
        text = "\n".join(f"{status}: {pn}" for pn, status in self._entries)
        QApplication.clipboard().setText(text)


class BOMEditorPane(QWidget):
    """Widget presenting BOM items with active/passive classification."""

    def __init__(self, assembly_id: int, parent=None) -> None:
        super().__init__(parent)
        self._assembly_id = assembly_id
        self._settings = QSettings("BOM_DB", f"BOMEditorPane/{assembly_id}")
        self._dirty_parts: Dict[int, str] = {}
        self._parts_state: Dict[int, Optional[str]] = {}
        self._rows_raw: list = []  # canonical read-model rows
        self._rows_by_part: Dict[int, object] = {}
        self._part_datasheets: Dict[int, Optional[str]] = {}
        self._datasheet_failed: set[int] = set()
        self._part_manual_links: Dict[int, str] = {}
        # Track parts with an in-progress Auto Datasheet operation
        self._datasheet_loading: set[int] = set()
        self._test_assignments: Dict[int, dict] = {}
        self._part_packages: Dict[int, Optional[str]] = {}
        self._dirty_packages: Dict[int, Optional[str]] = {}
        self._part_values: Dict[int, Optional[str]] = {}
        self._dirty_values: Dict[int, Optional[str]] = {}
        self._tolerances: dict[int, tuple[Optional[str], Optional[str]]] = {}
        self._dirty_tolerances: dict[int, tuple[Optional[str], Optional[str]]] = {}
        self._dirty_tests: set[int] = set()
        # Track manual edits to the Link column before Save
        self._dirty_links: Dict[int, Optional[str]] = {}
        # New staged edit trackers
        self._dirty_desc: Dict[int, Optional[str]] = {}
        self._dirty_mfg: Dict[int, Optional[str]] = {}
        self._locked_parts: set[int] = set()
        self._part_product_links: Dict[int, Optional[str]] = {}
        # Staged datasheet add/remove when Apply is off
        self._dirty_datasheets: Dict[int, Optional[str]] = {}
        # View mode: 'by_pn' or 'by_ref'
        self._part_numbers: Dict[int, str] = {}
        self._complex_links: Dict[int, object] = {}
        self._export_settings = QSettings("BOM_DB", "ExportForVIVA")
        self._complex_settings = get_complex_editor_settings()
        bridge_cfg = self._complex_settings.get('bridge', {}) if isinstance(self._complex_settings, dict) else {}
        bridge_enabled = bool(bridge_cfg.get('enabled')) if isinstance(bridge_cfg, dict) else False
        ui_enabled = bool(self._complex_settings.get('ui_enabled')) if isinstance(self._complex_settings, dict) else False
        self._complex_ui_enabled = ui_enabled and bridge_enabled
        self._assembly_mode: TestMode = TestMode.unpowered
        self._complex_splitter: Optional[QSplitter] = None
        self._complex_panel: Optional[ComplexPanel] = None
        self._main_splitter: Optional[QSplitter] = None
        self._schematic_panel: Optional[SchematicPanel] = None
        self._view_mode = self._settings.value("view_mode", "by_pn")
        self._col_indices = {  # will be updated on model rebuild
            "pn": 0,
            "ref": 1,
            "desc": 2,
            "mfg": 3,
            "ap": 4,
            "ds": 5,
            "link": 6,
            "test_method": 7,
            "test_detail": 8,
            "package": 9,
            "value": 10,
            "tol_p": 11,
            "tol_n": 12,
        }

        layout = QVBoxLayout(self)
        self.setWindowTitle(f"BOM Editor — Assembly {assembly_id}")
        self.toolbar = QToolBar()
        layout.addWidget(self.toolbar)

        # Filter
        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText("Filter…")

        # View mode toggle
        self.view_by_pn_act = QAction("By PN", self)
        self.view_by_ref_act = QAction("By Reference", self)
        self.view_by_pn_act.setCheckable(True)
        self.view_by_ref_act.setCheckable(True)
        from PyQt6.QtGui import QActionGroup
        self.view_group = QActionGroup(self)
        self.view_group.setExclusive(True)
        self.view_group.addAction(self.view_by_pn_act)
        self.view_group.addAction(self.view_by_ref_act)
        if self._view_mode == "by_ref":
            self.view_by_ref_act.setChecked(True)
        else:
            self._view_mode = "by_pn"
            self.view_by_pn_act.setChecked(True)

        # Columns menu
        self.columns_menu = QMenu("Columns", self)

        # Visible lines control
        self.lines_label = QLabel("Lines:")
        self.lines_spin = NoWheelSpinBox()
        self.lines_spin.setRange(1, 10)
        self.lines_spin.setValue(self._settings.value("lines", 1, type=int))

        # Apply immediately toggle
        self.apply_act = QAction("Apply Immediately", self)
        self.apply_act.setCheckable(True)

        # Save button
        self.save_act = QAction("Save", self)
        self.save_act.setEnabled(False)

        # Autofill button
        self.autofill_act = QAction("Autofill", self)

        # Auto datasheet button
        self.auto_ds_act = QAction("Auto Datasheet…", self)
        self.auto_ds_act.setEnabled(False)
        self.auto_ds_act.triggered.connect(self._auto_datasheet)

        # Export for VIVA
        self.export_viva_act = QAction("Export for VIVA", self)
        self.export_viva_act.triggered.connect(self._on_export_viva)

        # Table
        self.table = QTableView()

        # Enable per-cell hover
        self.table.setMouseTracking(True)
        self.table.viewport().setMouseTracking(True)

        # Style: white background, light-blue selection, light-gray hover.
        # Ensure hover doesn’t look like native “blue selection” on Windows.
        self.table.setStyleSheet(
            """
QTableView {
    background: #FFFFFF;
    alternate-background-color: #FAFAFA;
}

QTableView::item {
    selection-color: black;  /* readable text on light backgrounds */
}

/* Hover on non-selected cells */
QTableView::item:hover:!selected {
    background: #F3F4F6;  /* light gray */
}

/* Selected cells, whether view is active or not */
QTableView::item:selected,
QTableView::item:selected:active,
QTableView::item:selected:!active {
    background: #DCEBFF;  /* light blue */
    color: black;
}

/* When a selected cell is also hovered, keep it in the same palette (slightly stronger if you prefer) */
QTableView::item:selected:hover {
    background: #CFE3FF;  /* optional slightly darker blue on hover+selected */
}
"""
        )

        self.model = UndoableStandardItemModel(0, 5, self)
        self.model.itemChanged.connect(self._on_item_changed)
        self.model.changed.connect(self._on_model_changed)
        self._updating = False
        self.undo_stack = QUndoStack(self)
        self.undo_stack.setUndoLimit(10)
        # headers set in _rebuild_model()
        self.proxy = BOMFilterProxy(self)
        self.proxy.setSourceModel(self.model)
        self.table.setModel(self.proxy)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

        # Delegates
        self.delegate = CycleToggleDelegate(self.table)
        self.delegate.valueChanged.connect(self._on_value_changed)
        self.wrap_delegate = WrapTextDelegate(self.table)
        self.wrap_delegate.set_line_count(self.lines_spin.value())
        self.method_delegate = TestMethodDelegate(self.table)
        self.method_delegate.methodChanged.connect(self._on_method_changed)
        # Function options loaded for test detail delegate (with sensible fallback)
        self._function_options = self._load_function_options()
        # Map for case-insensitive validation and canonicalization
        self._function_options_canon = {opt.upper(): opt for opt in (self._function_options or [])}
        self._prefix_macros = load_prefix_macros()
        self.detail_delegate = TestDetailDelegate(self.table, options=self._function_options)
        self.detail_delegate.detailClicked.connect(self._on_detail_clicked)
        self.detail_delegate.detailChanged.connect(self._on_detail_changed)
        self.lines_spin.valueChanged.connect(self._on_lines_changed)
        # Column assignment done in _rebuild_model

        self.filter_edit.textChanged.connect(self.proxy.setFilterString)
        self.apply_act.toggled.connect(self._on_apply_toggled)
        self.save_act.triggered.connect(self._save_changes)
        self.autofill_act.triggered.connect(self._autofill_fields)
        self.view_by_pn_act.toggled.connect(lambda checked: self._on_view_mode_changed("by_pn", checked))
        self.view_by_ref_act.toggled.connect(lambda checked: self._on_view_mode_changed("by_ref", checked))

        # Copy/Paste support
        self.copy_act = QAction("Copy", self)
        self.copy_act.setShortcut(QKeySequence.StandardKey.Copy)
        self.copy_act.triggered.connect(self._copy_selection)
        self.table.addAction(self.copy_act)
        self.paste_act = QAction("Paste", self)
        self.paste_act.setShortcut(QKeySequence.StandardKey.Paste)
        self.paste_act.triggered.connect(self._paste_selection)
        self.table.addAction(self.paste_act)
        # Delete selected rows
        self.delete_act = QAction("Delete Selected", self)
        self.delete_act.setShortcut(QKeySequence.StandardKey.Delete)
        # Delete key: clear selected cells unless entire rows are selected
        self.delete_act.triggered.connect(self._delete_or_clear_selection)
        self.table.addAction(self.delete_act)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.ActionsContextMenu)
        # Link context actions
        self.open_link_act = QAction("Open Link", self)
        self.copy_link_act = QAction("Copy Link", self)
        self.open_link_act.triggered.connect(self._open_selected_link)
        self.copy_link_act.triggered.connect(self._copy_selected_link)
        self.table.addAction(self.open_link_act)
        self.table.addAction(self.copy_link_act)

        # Remove datasheet action (context menu)
        self.remove_ds_act = QAction("Remove Datasheet", self)
        self.remove_ds_act.triggered.connect(self._remove_selected_datasheets)
        self.table.addAction(self.remove_ds_act)

        if self._complex_ui_enabled:
            self._complex_splitter = QSplitter(Qt.Orientation.Horizontal, self)
            self._complex_splitter.addWidget(self.table)
            self._complex_panel = ComplexPanel(self)
            self._complex_panel.linkUpdated.connect(self._on_complex_link_updated)
            self._complex_panel.hide()
            self._complex_splitter.addWidget(self._complex_panel)
            self._complex_splitter.setStretchFactor(0, 3)
            self._complex_splitter.setStretchFactor(1, 2)
            try:
                self._complex_splitter.setCollapsible(1, True)
            except Exception:
                pass
            table_container: QWidget = self._complex_splitter
        else:
            table_container = self.table

        self._schematic_panel = SchematicPanel(self._assembly_id, self._open_pdf_path, self)
        self._schematic_panel.hide()
        self._main_splitter = QSplitter(Qt.Orientation.Vertical, self)
        self._main_splitter.addWidget(table_container)
        self._main_splitter.addWidget(self._schematic_panel)
        self._main_splitter.setStretchFactor(0, 3)
        self._main_splitter.setStretchFactor(1, 2)
        try:
            self._main_splitter.setCollapsible(1, True)
        except Exception:
            pass
        layout.addWidget(self._main_splitter)

        # Undo/Redo support via QUndoStack
        self.undo_act = QAction("Undo", self, shortcut=QKeySequence.StandardKey.Undo)
        self.redo_act = QAction("Redo", self, shortcut=QKeySequence.StandardKey.Redo)
        self.undo_act.triggered.connect(self.undo_stack.undo)
        self.redo_act.triggered.connect(self.undo_stack.redo)
        self.undo_act.setEnabled(False)
        self.redo_act.setEnabled(False)
        self.undo_stack.canUndoChanged.connect(self.undo_act.setEnabled)
        self.undo_stack.canRedoChanged.connect(self.redo_act.setEnabled)
        self.table.addAction(self.undo_act)
        self.table.addAction(self.redo_act)

        # Toolbar menus
        self.btn_edit = QToolButton(self)
        self.btn_edit.setText("Edit")
        self.btn_edit.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.menu_edit = QMenu(self.btn_edit)
        self.menu_edit.addAction(self.undo_act)
        self.menu_edit.addAction(self.redo_act)
        self.menu_edit.addAction(self.copy_act)
        self.menu_edit.addAction(self.paste_act)
        self.menu_edit.addAction(self.delete_act)
        self.btn_edit.setMenu(self.menu_edit)
        self.toolbar.addWidget(self.btn_edit)

        self.btn_view = QToolButton(self)
        self.btn_view.setText("View")
        self.btn_view.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.menu_view = QMenu(self.btn_view)
        self.menu_view.addAction(self.view_by_pn_act)
        self.menu_view.addAction(self.view_by_ref_act)
        self.menu_view.addMenu(self.columns_menu)
        self.view_schematics_act = QAction("Schematics Panel", self)
        self.view_schematics_act.setCheckable(True)
        self.view_schematics_act.toggled.connect(self._toggle_schematic_panel)
        self.menu_view.addAction(self.view_schematics_act)
        self.btn_view.setMenu(self.menu_view)
        self.toolbar.addWidget(self.btn_view)

        self.mode_badge = QLabel("", self)
        self.mode_badge.setObjectName("modeBadge")
        self.mode_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.mode_badge.setMinimumWidth(200)
        self.toolbar.addWidget(self.mode_badge)

        self.btn_data = QToolButton(self)
        self.btn_data.setText("Data")
        self.btn_data.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.menu_data = QMenu(self.btn_data)
        self.menu_data.addAction(self.apply_act)
        self.menu_data.addAction(self.save_act)
        self.publish_test_defaults_act = QAction("Publish Test Defaults to DB…", self)
        self.publish_test_defaults_act.triggered.connect(self._publish_test_defaults)
        self.menu_data.addAction(self.publish_test_defaults_act)
        self.btn_data.setMenu(self.menu_data)
        self.toolbar.addWidget(self.btn_data)

        self.btn_tools = QToolButton(self)
        self.btn_tools.setText("Tools")
        self.btn_tools.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.menu_tools = QMenu(self.btn_tools)
        self.menu_tools.addAction(self.autofill_act)
        self.menu_tools.addAction(self.auto_ds_act)
        self.menu_tools.addAction(self.export_viva_act)
        # Export to Excel
        self.export_xlsx_act = QAction("Export to Excel", self)
        self.export_xlsx_act.triggered.connect(self._export_excel)
        self.menu_tools.addAction(self.export_xlsx_act)
        self.btn_tools.setMenu(self.menu_tools)
        self.toolbar.addWidget(self.btn_tools)

        # Filter and lines widgets
        self.toolbar.addWidget(self.filter_edit)
        self.toolbar.addWidget(self.lines_label)
        self.toolbar.addWidget(self.lines_spin)

        # Enable wrapping and resize rows when columns change
        self.table.setWordWrap(True)
        self.table.horizontalHeader().sectionResized.connect(lambda *_: self.table.resizeRowsToContents())

        self._load_data()
        self._rebuild_model()
        self._on_lines_changed(self.lines_spin.value())
        if self.table.selectionModel():
            self.table.selectionModel().selectionChanged.connect(self._on_table_selection_changed)
        self._on_table_selection_changed()

    # ------------------------------------------------------------------
    def _setup_columns_menu(self) -> None:
        # Clear previous actions
        self.columns_menu.clear()
        self._column_actions: list[QAction] = []
        if self._view_mode == "by_pn":
            headers = [
                "PN",
                "References",
                "Description",
                "Manufacturer",
                "Active/Passive",
                "Datasheet",
                "Link",
                "Test Method",
                "Test Detail",
                "Package",
                "Value",
                "Tol Positive",
                "Tol Negative",
            ]
        else:
            headers = [
                "Reference",
                "PN",
                "Description",
                "Manufacturer",
                "Active/Passive",
                "Datasheet",
                "Link",
                "Test Method",
                "Test Detail",
                "Package",
                "Value",
                "Tol Positive",
                "Tol Negative",
            ]
        if "test_method_powered" in self._col_indices:
            insert_at = self._col_indices["test_method_powered"]
            headers.insert(insert_at, "Test Method (Powered)")
            headers.insert(insert_at + 1, "Test Detail (Powered)")
        self.model.setHorizontalHeaderLabels(headers)
        # Persist settings per mode
        mode_key = "cols_by_pn" if self._view_mode == "by_pn" else "cols_by_ref"
        for idx, name in enumerate(headers):
            act = QAction(name, self)
            act.setCheckable(True)
            visible = self._settings.value(f"{mode_key}/col{idx}_visible", True, type=bool)
            act.setChecked(visible)
            act.toggled.connect(lambda checked, col=idx: self.table.setColumnHidden(col, not checked))
            self.columns_menu.addAction(act)
            self._column_actions.append(act)
            self.table.setColumnHidden(idx, not visible)
            width = self._settings.value(f"{mode_key}/col{idx}_width", type=int)
            if width:
                self.table.setColumnWidth(idx, int(width))
        # Assign delegates
        ap_col = self._col_indices["ap"]
        self.table.setItemDelegateForColumn(ap_col, self.delegate)
        ref_col = self._col_indices["ref"]
        self.table.setItemDelegateForColumn(ref_col, self.wrap_delegate)
        tm_col = self._col_indices["test_method"]
        self.table.setItemDelegateForColumn(tm_col, self.method_delegate)
        powered_tm_col = self._col_indices.get("test_method_powered")
        if powered_tm_col is not None:
            self.table.setItemDelegateForColumn(powered_tm_col, self.method_delegate)
        td_col = self._col_indices["test_detail"]
        powered_td_col = self._col_indices.get("test_detail_powered")
        # Keep delegate up to date on column positions and options
        relevant_method_col = self._col_indices.get(
            self._relevant_test_column_keys()[0],
            tm_col,
        )
        method_map: Dict[int, int] = {}
        if tm_col is not None:
            method_map[td_col] = tm_col
        if powered_tm_col is not None and powered_td_col is not None:
            method_map[powered_td_col] = powered_tm_col
        self.detail_delegate.set_method_column_map(method_map)
        self.detail_delegate.set_test_method_col(relevant_method_col)
        self.table.setItemDelegateForColumn(td_col, self.detail_delegate)
        if powered_td_col is not None:
            self.table.setItemDelegateForColumn(powered_td_col, self.detail_delegate)
        self._update_test_header_labels()
        # Make icon-only columns compact while keeping icons prominent
        try:
            from PyQt6.QtWidgets import QHeaderView as _QHV
            hdr = self.table.horizontalHeader()
            ds_col = self._col_indices.get("ds")
            link_col = self._col_indices.get("link")
            icon_col_width = 36
            if ds_col is not None:
                hdr.setSectionResizeMode(ds_col, _QHV.ResizeMode.Fixed)
                self.table.setColumnWidth(ds_col, icon_col_width)
            if link_col is not None:
                hdr.setSectionResizeMode(link_col, _QHV.ResizeMode.Fixed)
                self.table.setColumnWidth(link_col, icon_col_width)
        except Exception:
            pass
        # Ensure combo is visible for 'Macro' rows
        self._sync_detail_editors()
        # Set reference column for natural sorting and skip AP from filter
        self.proxy.setReferenceColumn(ref_col)
        self.proxy.setSkipColumns({ap_col})

    def _relevant_test_column_keys(self) -> tuple[str, str]:
        powered_cols_present = (
            "test_method_powered" in self._col_indices
            and "test_detail_powered" in self._col_indices
        )
        if self._assembly_mode == TestMode.powered and powered_cols_present:
            return "test_method_powered", "test_detail_powered"
        return "test_method", "test_detail"

    def _update_test_header_labels(self) -> None:
        label_map = {
            "test_method": "Test Method",
            "test_detail": "Test Detail",
        }
        if "test_method_powered" in self._col_indices:
            label_map["test_method_powered"] = "Test Method (Powered)"
        if "test_detail_powered" in self._col_indices:
            label_map["test_detail_powered"] = "Test Detail (Powered)"
        relevant_method, relevant_detail = self._relevant_test_column_keys()
        relevant_cols = {
            self._col_indices.get(relevant_method),
            self._col_indices.get(relevant_detail),
        }
        for key, base_label in label_map.items():
            col = self._col_indices.get(key)
            if col is None:
                continue
            item = self.model.horizontalHeaderItem(col)
            if item is None:
                continue
            is_relevant = col in relevant_cols
            label = base_label + (" — Current Mode" if is_relevant else "")
            item.setText(label)
            if is_relevant:
                item.setData(
                    QBrush(QColor("#BFDBFE")),
                    int(Qt.ItemDataRole.BackgroundRole),
                )
                item.setData(
                    QBrush(QColor("#1E3A8A")),
                    int(Qt.ItemDataRole.ForegroundRole),
                )
                item.setToolTip(
                    "These values are used for Autofill, VIVA export, and Excel export."
                )
            else:
                item.setData(QBrush(), int(Qt.ItemDataRole.BackgroundRole))
                item.setData(QBrush(), int(Qt.ItemDataRole.ForegroundRole))
                if key.endswith("powered"):
                    item.setToolTip(
                        "Reference values for powered tests (read-only display)."
                    )
                else:
                    item.setToolTip("")

    def _update_mode_badge(self) -> None:
        badge = getattr(self, "mode_badge", None)
        if badge is None:
            return
        if self._assembly_mode == TestMode.powered:
            text = "Powered board — using powered test columns"
            bg = "#F97316"
            fg = "#FFFFFF"
        else:
            text = "Unpowered board — using unpowered test columns"
            bg = "#0EA5E9"
            fg = "#FFFFFF"
        badge.setText(text)
        badge.setStyleSheet(
            "QLabel#modeBadge {"
            " border-radius: 14px;"
            " padding: 4px 12px;"
            " font-weight: 600;"
            f" color: {fg};"
            f" background-color: {bg};"
            "}"
        )
        badge.setToolTip(
            "This reflects the assembly test mode currently applied to the BOM."
        )

    def _load_data(self) -> None:
        # Keep canonical raw rows
        with app_state.get_session() as session:
            self._rows_raw = services.get_joined_bom_for_assembly(session, self._assembly_id)
            asm = session.get(Assembly, self._assembly_id)
        if asm is not None:
            mode_val = getattr(asm, "test_mode", TestMode.unpowered)
            if isinstance(mode_val, TestMode):
                self._assembly_mode = mode_val
            else:
                try:
                    self._assembly_mode = TestMode(str(mode_val))
                except ValueError:
                    self._assembly_mode = TestMode.unpowered
        else:
            self._assembly_mode = TestMode.unpowered
        schem_visible = bool(self._settings.value("schematics_visible", False, type=bool))
        if self._schematic_panel:
            self.view_schematics_act.blockSignals(True)
            self.view_schematics_act.setChecked(schem_visible)
            self.view_schematics_act.blockSignals(False)
            self._set_schematic_panel_visible(schem_visible)
        self._update_mode_badge()
        # Seed parts state from DB
        self._parts_state.clear()
        self._part_datasheets.clear()
        self._part_packages.clear()
        self._part_numbers.clear()
        self._part_values.clear()
        self._tolerances.clear()
        self._dirty_datasheets.clear()
        self._complex_links.clear()
        self._rows_by_part = {}
        for r in self._rows_raw:
            # Use DB-provided value; may be None in future schema
            self._parts_state[r.part_id] = getattr(r, "active_passive", None)
            self._part_datasheets[r.part_id] = getattr(r, "datasheet_url", None)
            self._part_packages[r.part_id] = getattr(r, "package", None)
            self._part_numbers[r.part_id] = getattr(r, "part_number", "")
            self._part_values[r.part_id] = getattr(r, "value", None)
            self._tolerances[r.part_id] = (getattr(r, "tol_p", None), getattr(r, "tol_n", None))
            # Seed product link if available so the Link column shows after reload
            self._part_product_links[r.part_id] = getattr(r, "product_url", None)
            self._rows_by_part.setdefault(r.part_id, r)
        self._reload_complex_links({r.part_id for r in self._rows_raw})
        # Overlay any saved test assignments from settings for visible parts
        self._load_test_assignments_from_settings()

    def _auto_infer(self, value: Optional[str], reference: str) -> Optional[str]:
        # Do not override explicit value
        if value in ("active", "passive"):
            return value
        if isinstance(reference, str) and reference[:1].upper() in ("R", "L", "C"):
            return "passive"
        return None

    def _blank_assignment(self) -> dict:
        return {
            "method": "",
            "detail": None,
            "qt_path": None,
            "method_powered": "",
            "detail_powered": None,
            "qt_path_powered": None,
        }

    def _ensure_assignment_keys(self, assignment: dict) -> dict:
        defaults = self._blank_assignment()
        for key, default in defaults.items():
            assignment.setdefault(key, default)
        return assignment

    def _get_assignment(self, part_id: int) -> dict:
        assignment = self._test_assignments.get(part_id)
        if assignment is None:
            assignment = self._blank_assignment()
            self._test_assignments[part_id] = assignment
        else:
            self._ensure_assignment_keys(assignment)
        return assignment

    def _assignment_keys_for_mode(self, mode: TestMode) -> tuple[str, str, str]:
        if mode == TestMode.powered:
            return "method_powered", "detail_powered", "qt_path_powered"
        return "method", "detail", "qt_path"

    def _column_keys_for_mode(self, mode: TestMode) -> tuple[str, str]:
        if mode == TestMode.powered:
            return "test_method_powered", "test_detail_powered"
        return "test_method", "test_detail"

    def _assignment_view(self, assignment: dict, mode: TestMode) -> dict:
        method_key, detail_key, qt_key = self._assignment_keys_for_mode(mode)
        return {
            "method": assignment.get(method_key) or "",
            "detail": assignment.get(detail_key),
            "qt_path": assignment.get(qt_key),
        }

    def _mode_for_column(self, column: int) -> TestMode:
        powered_cols = {
            idx
            for key in ("test_method_powered", "test_detail_powered")
            if (idx := self._col_indices.get(key)) is not None
        }
        if column in powered_cols:
            return TestMode.powered
        return TestMode.unpowered

    def _collect_publish_selection(
        self,
        part_id: int,
        mode: TestMode,
        assignment: dict,
    ) -> tuple[dict | None, str | None]:
        view = self._assignment_view(assignment, mode)
        method = (view.get("method") or "").strip()
        if not method:
            return None, None
        detail = (view.get("detail") or "").strip() if view.get("detail") else None
        qt_path = (view.get("qt_path") or "").strip() or None
        selection: dict[str, object] = {
            "method": method,
            "detail": detail,
            "qt_path": qt_path,
        }
        if method == "Complex":
            selection["persist"] = False
            return selection, None
        resolved = self._rows_by_part.get(part_id)
        if detail is None and resolved is not None:
            detail_attr = "test_detail_powered" if mode == TestMode.powered else "test_detail"
            resolved_detail = getattr(resolved, detail_attr, None)
            if resolved_detail:
                selection["detail"] = str(resolved_detail)
                detail = selection["detail"]
        if method == "Quick test (QT)":
            if not qt_path:
                return None, "Quick test requires selecting an XML file"
            if not detail:
                selection["detail"] = Path(qt_path).name
            selection["persist"] = True
            return selection, None
        if method == "Macro":
            if not detail:
                return None, "Macro selections require a detail value"
            selection["persist"] = True
            return selection, None
        if method == "Python code":
            if not detail:
                return None, "Python code selections require a detail value"
            selection["persist"] = True
            return selection, None
        return None, f"Unsupported method '{method}'"

    def _make_method_item(
        self,
        resolved: Optional[str],
        assignment: dict,
        mode: TestMode = TestMode.unpowered,
    ) -> QStandardItem:
        item = QStandardItem(resolved or "")
        view = self._assignment_view(self._ensure_assignment_keys(assignment), mode)
        assigned = (view.get("method") or "").strip()
        item.setData(assigned, int(Qt.ItemDataRole.EditRole))
        item.setData(assigned, int(Qt.ItemDataRole.ToolTipRole))
        return item

    def _make_detail_item(
        self,
        resolved: Optional[str],
        assignment: dict,
        mode: TestMode = TestMode.unpowered,
    ) -> QStandardItem:
        item = QStandardItem(resolved or "")
        view = self._assignment_view(self._ensure_assignment_keys(assignment), mode)
        method = (view.get("method") or "").strip()
        if method == "Macro":
            edit_value = view.get("detail") or ""
        elif method == "Quick test (QT)":
            edit_value = view.get("qt_path") or ""
        else:
            edit_value = view.get("detail") or ""
        item.setData(edit_value, int(Qt.ItemDataRole.EditRole))
        tooltip = self._detail_text_for(method, view.get("detail"), view.get("qt_path"))
        item.setData(tooltip, int(Qt.ItemDataRole.ToolTipRole))
        return item

    def _rebuild_model(self) -> None:
        # Build model based on view mode
        self.model.setRowCount(0)
        powered = self._assembly_mode == TestMode.powered
        base_columns = 13
        extra = 2 if powered else 0
        self.model.setColumnCount(base_columns + extra)
        if self._view_mode == "by_pn":
            # Column map
            self._col_indices = {
                "pn": 0,
                "ref": 1,
                "desc": 2,
                "mfg": 3,
                "ap": 4,
                "ds": 5,
                "link": 6,
                "test_method": 7,
                "test_detail": 8,
            }
            next_col = 9
            if powered:
                self._col_indices["test_method_powered"] = next_col
                next_col += 1
                self._col_indices["test_detail_powered"] = next_col
                next_col += 1
            self._col_indices.update(
                {
                    "package": next_col,
                    "value": next_col + 1,
                    "tol_p": next_col + 2,
                    "tol_n": next_col + 3,
                }
            )
            self._build_by_pn()
        else:
            self._col_indices = {
                "ref": 0,
                "pn": 1,
                "desc": 2,
                "mfg": 3,
                "ap": 4,
                "ds": 5,
                "link": 6,
                "test_method": 7,
                "test_detail": 8,
            }
            next_col = 9
            if powered:
                self._col_indices["test_method_powered"] = next_col
                next_col += 1
                self._col_indices["test_detail_powered"] = next_col
                next_col += 1
            self._col_indices.update(
                {
                    "package": next_col,
                    "value": next_col + 1,
                    "tol_p": next_col + 2,
                    "tol_n": next_col + 3,
                }
            )
            self._build_by_ref()
        # Setup columns menu and sorting based on new headers
        self._setup_columns_menu()
        self.table.setSortingEnabled(True)
        self.table.sortByColumn(self._col_indices["ref"], Qt.SortOrder.AscendingOrder)
        # Defer autosize until columns are applied
        QTimer.singleShot(0, self._autosize_window_to_columns)
        QTimer.singleShot(0, self._install_datasheet_widgets)
        QTimer.singleShot(0, self._sync_detail_editors)
        if self._complex_panel:
            QTimer.singleShot(0, self._update_complex_panel_context)
        self._update_mode_badge()

    def _build_by_pn(self) -> None:
        from collections import defaultdict

        groups: Dict[int, List[object]] = defaultdict(list)
        for r in self._rows_raw:
            groups[r.part_id].append(r)

        for part_id, rows in groups.items():
            # Aggregate references
            refs_sorted = sorted((x.reference for x in rows), key=natural_key)
            refs_str = ",".join(refs_sorted)
            # Determine value: use explicit if present, else auto-infer, then overlay staged
            explicit = next((x.active_passive for x in rows if getattr(x, "active_passive", None) in ("active", "passive")), None)
            mode_val = explicit or self._auto_infer(None, rows[0].reference)
            if part_id in self._dirty_parts:
                mode_val = self._dirty_parts[part_id]
            # Persist or stage auto-inferred, if any
            self._handle_auto_infer_persistence(part_id, mode_val, explicit)

            ta = self._get_assignment(part_id)
            powered_cols = "test_method_powered" in self._col_indices
            show_powered = powered_cols and (
                mode_val == "active"
                or any(getattr(x, "active_passive", None) == "active" for x in rows)
            )
            resolved_powered_method = rows[0].test_method_powered if show_powered else ""
            resolved_powered_detail = rows[0].test_detail_powered if show_powered else ""
            method_item = self._make_method_item(rows[0].test_method, ta, TestMode.unpowered)
            detail_item = self._make_detail_item(rows[0].test_detail, ta, TestMode.unpowered)
            row_items = [
                QStandardItem(rows[0].part_number),
                QStandardItem(refs_str),
                QStandardItem(rows[0].description or ""),
                QStandardItem(rows[0].manufacturer or ""),
                QStandardItem(mode_val or ""),
                QStandardItem(""),
                QStandardItem(""),
                method_item,
                detail_item,
            ]
            if powered_cols:
                row_items.append(
                    self._make_method_item(
                        resolved_powered_method,
                        ta,
                        TestMode.powered,
                    )
                )
                row_items.append(
                    self._make_detail_item(
                        resolved_powered_detail,
                        ta,
                        TestMode.powered,
                    )
                )
            row_items.extend(
                [
                    QStandardItem(self._part_packages.get(part_id) or ""),
                    QStandardItem(self._part_values.get(part_id) or ""),
                    QStandardItem(self._tolerances.get(part_id, (None, None))[0] or ""),
                    QStandardItem(self._tolerances.get(part_id, (None, None))[1] or ""),
                ]
            )
            for i, it in enumerate(row_items):
                flags = it.flags() | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
                editable_keys = list(self._relevant_test_column_keys())
                if "test_method_powered" in self._col_indices:
                    editable_keys.extend(["test_method_powered", "test_detail_powered"])
                editable_keys.extend(["desc", "mfg", "package", "value", "tol_p", "tol_n"])
                editable = {
                    self._col_indices.get(key)
                    for key in editable_keys
                    if self._col_indices.get(key) is not None
                }
                if i not in editable or part_id in self._locked_parts:
                    flags &= ~Qt.ItemFlag.ItemIsEditable
                it.setFlags(flags)
                it.setData(part_id, PartIdRole)
                if i == self._col_indices["ap"]:
                    it.setData(mode_val, ModeRole)
                if i == self._col_indices["ds"]:
                    it.setData(self._part_datasheets.get(part_id), DatasheetRole)
                if i == self._col_indices["link"]:
                    it.setData(self._part_product_links.get(part_id) or "", LinkUrlRole)
            self.model.appendRow(row_items)

    def _build_by_ref(self) -> None:
        for r in self._rows_raw:
            explicit = getattr(r, "active_passive", None)
            mode_val = explicit or self._auto_infer(None, r.reference)
            if r.part_id in self._dirty_parts:
                mode_val = self._dirty_parts[r.part_id]
            self._handle_auto_infer_persistence(r.part_id, mode_val, explicit)
            ta = self._get_assignment(r.part_id)
            powered_cols = "test_method_powered" in self._col_indices
            is_active = mode_val == "active" or getattr(r, "active_passive", None) == "active"
            resolved_powered_method = r.test_method_powered if powered_cols and is_active else ""
            resolved_powered_detail = r.test_detail_powered if powered_cols and is_active else ""
            method_item = self._make_method_item(r.test_method, ta, TestMode.unpowered)
            detail_item = self._make_detail_item(r.test_detail, ta, TestMode.unpowered)
            items = [
                QStandardItem(r.reference),
                QStandardItem(r.part_number),
                QStandardItem(r.description or ""),
                QStandardItem(r.manufacturer or ""),
                QStandardItem(mode_val or ""),
                QStandardItem(""),
                QStandardItem(""),
                method_item,
                detail_item,
            ]
            if powered_cols:
                items.append(
                    self._make_method_item(
                        resolved_powered_method,
                        ta,
                        TestMode.powered,
                    )
                )
                items.append(
                    self._make_detail_item(
                        resolved_powered_detail,
                        ta,
                        TestMode.powered,
                    )
                )
            items.extend(
                [
                    QStandardItem(self._part_packages.get(r.part_id) or ""),
                    QStandardItem(self._part_values.get(r.part_id) or ""),
                    QStandardItem(self._tolerances.get(r.part_id, (None, None))[0] or ""),
                    QStandardItem(self._tolerances.get(r.part_id, (None, None))[1] or ""),
                ]
            )
            for i, it in enumerate(items):
                flags = it.flags() | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
                editable_keys = list(self._relevant_test_column_keys())
                if "test_method_powered" in self._col_indices:
                    editable_keys.extend(["test_method_powered", "test_detail_powered"])
                editable_keys.extend(["desc", "mfg", "package", "value", "tol_p", "tol_n"])
                editable = {
                    self._col_indices.get(key)
                    for key in editable_keys
                    if self._col_indices.get(key) is not None
                }
                if i not in editable or r.part_id in self._locked_parts:
                    flags &= ~Qt.ItemFlag.ItemIsEditable
                it.setFlags(flags)
                it.setData(r.part_id, PartIdRole)
                it.setData(getattr(r, "bom_item_id", None), BOMItemIdRole)
                if i == self._col_indices["ap"]:
                    it.setData(mode_val, ModeRole)
                if i == self._col_indices["ds"]:
                    it.setData(self._part_datasheets.get(r.part_id), DatasheetRole)
                if i == self._col_indices["link"]:
                    it.setData(self._part_product_links.get(r.part_id) or "", LinkUrlRole)
            self.model.appendRow(items)
    def _refresh_rows_for_part(self, part_id: int, mode: TestMode | None = None) -> None:
        assignment = self._get_assignment(part_id)
        modes = [mode] if mode is not None else [TestMode.unpowered, TestMode.powered]
        for current_mode in modes:
            method_key, detail_key = self._column_keys_for_mode(current_mode)
            method_col = self._col_indices.get(method_key)
            detail_col = self._col_indices.get(detail_key)
            if method_col is None and detail_col is None:
                continue
            view = self._assignment_view(assignment, current_mode)
            method_value = (view.get("method") or "").strip()
            detail_raw = view.get("detail") or ""
            qt_path = view.get("qt_path") or None
            detail_edit_value = detail_raw
            if method_value == "Quick test (QT)":
                detail_edit_value = qt_path or ""
            detail_label = self._detail_text_for(method_value, detail_raw, qt_path)
            for row in range(self.model.rowCount()):
                owns_part = False
                for c in range(self.model.columnCount()):
                    if self.model.data(self.model.index(row, c), PartIdRole) == part_id:
                        owns_part = True
                        break
                if not owns_part:
                    continue
                if method_col is not None:
                    idx = self.model.index(row, method_col)
                    self.model.setData(idx, method_value, int(Qt.ItemDataRole.EditRole))
                    self.model.setData(idx, method_value, int(Qt.ItemDataRole.ToolTipRole))
                if detail_col is not None:
                    idx = self.model.index(row, detail_col)
                    self.model.setData(idx, detail_edit_value, int(Qt.ItemDataRole.EditRole))
                    self.model.setData(idx, detail_label, int(Qt.ItemDataRole.ToolTipRole))
                break

    def _on_method_changed(self, part_id: int, new_method: str, column: int) -> None:
        assignment = self._get_assignment(part_id)
        mode = self._mode_for_column(column)
        method_key, detail_key, qt_key = self._assignment_keys_for_mode(mode)
        assignment[method_key] = new_method
        if new_method != "Macro":
            assignment[detail_key] = None
        if new_method != "Quick test (QT)":
            assignment[qt_key] = None
        self._refresh_rows_for_part(part_id, mode)
        if new_method == "Complex":
            self._ensure_complex_panel_visible_for_part(part_id)
        else:
            self._update_complex_panel_context()
        # Ensure editors reflect Macro selection state
        self._sync_detail_editors()
        # Persist/stage according to Apply toggle
        if self.apply_act.isChecked():
            self._persist_test_assignment(part_id)
        else:
            self._dirty_tests.add(part_id)
            self.save_act.setEnabled(True)

    def _on_detail_clicked(self, part_id: int, column: int) -> None:
        assignment = self._get_assignment(part_id)
        mode = self._mode_for_column(column)
        method_key, detail_key, qt_key = self._assignment_keys_for_mode(mode)
        method = assignment.get(method_key, "")
        if method == "Macro":
            self._show_stub_dialog(
                "This would open the Macro selector (closed list) and save the chosen Macro for this PN. (Not implemented yet)."
            )
        elif method == "Complex":
            if self._complex_panel:
                self._ensure_complex_panel_visible_for_part(part_id)
            else:
                self._show_stub_dialog(
                    "Complex Editor integration is disabled. Enable it in settings to link complexes."
                )
        elif method == "Quick test (QT)":
            path, _ = QFileDialog.getOpenFileName(self, "Select Quick Test XML", "", "Quick Test XML (*.xml)")
            if path:
                assignment[qt_key] = path
                assignment[detail_key] = assignment.get(detail_key) or None
                self._refresh_rows_for_part(part_id, mode)
        elif method == "Python code":
            self._show_stub_dialog(
                "This would open a project chooser (folder with code, description, library links) and link it to this PN. (Not implemented yet)."
            )


    def _on_detail_changed(self, part_id: int, new_detail: Optional[str], column: int) -> None:
        # Record selection and refresh label, keep persistent editor
        assignment = self._get_assignment(part_id)
        mode = self._mode_for_column(column)
        _, detail_key, _ = self._assignment_keys_for_mode(mode)
        assignment[detail_key] = new_detail or None
        self._refresh_rows_for_part(part_id, mode)
        # Keep editor visible where applicable
        self._sync_detail_editors()
        # Persist/stage according to Apply toggle
        if self.apply_act.isChecked():
            self._persist_test_assignment(part_id)
        else:
            self._dirty_tests.add(part_id)
            self.save_act.setEnabled(True)

    def _sync_detail_editors(self) -> None:
        # Do not auto-open editors; keep cells passive until user double-clicks
        model = self.table.model()  # proxy
        if model is None:
            return
        method_key, detail_key = self._relevant_test_column_keys()
        tm_col = self._col_indices.get(method_key)
        td_col = self._col_indices.get(detail_key)
        if tm_col is None or td_col is None:
            return
        rows = model.rowCount()
        for r in range(rows):
            tm_idx = model.index(r, tm_col)
            method = model.data(tm_idx) or ""
            td_idx = model.index(r, td_col)
            # Always keep closed; opening is controlled by QAbstractItemView triggers
            self.table.closePersistentEditor(td_idx)

    # ------------------------------------------------------------------
    # Test assignment persistence using QSettings
    def _persist_test_assignment(self, part_id: int) -> None:
        assignment = self._get_assignment(part_id)
        self._settings.setValue(f"test/method/{part_id}", assignment.get("method") or "")
        self._settings.setValue(f"test/detail/{part_id}", assignment.get("detail") or "")
        self._settings.setValue(f"test/qt_path/{part_id}", assignment.get("qt_path") or "")
        self._settings.setValue(
            f"test/method_powered/{part_id}", assignment.get("method_powered") or ""
        )
        self._settings.setValue(
            f"test/detail_powered/{part_id}", assignment.get("detail_powered") or ""
        )
        self._settings.setValue(
            f"test/qt_path_powered/{part_id}", assignment.get("qt_path_powered") or ""
        )

    def _load_test_assignments_from_settings(self) -> None:
        # Populate self._test_assignments with any saved values for parts in current view
        part_ids = {r.part_id for r in self._rows_raw}
        for pid in part_ids:
            method = self._settings.value(f"test/method/{pid}", "")
            detail = self._settings.value(f"test/detail/{pid}", "")
            qt_path = self._settings.value(f"test/qt_path/{pid}", "")
            method_p = self._settings.value(f"test/method_powered/{pid}", "")
            detail_p = self._settings.value(f"test/detail_powered/{pid}", "")
            qt_path_p = self._settings.value(f"test/qt_path_powered/{pid}", "")
            if any(
                [
                    (method or "").strip(),
                    (detail or "").strip(),
                    (qt_path or "").strip(),
                    (method_p or "").strip(),
                    (detail_p or "").strip(),
                    (qt_path_p or "").strip(),
                ]
            ):
                ta = self._get_assignment(pid)
                ta["method"] = (method or "").strip()
                ta["detail"] = (detail or "").strip() or None
                ta["qt_path"] = (qt_path or "").strip() or None
                ta["method_powered"] = (method_p or "").strip()
                ta["detail_powered"] = (detail_p or "").strip() or None
                ta["qt_path_powered"] = (qt_path_p or "").strip() or None

    def _publish_test_defaults(self) -> None:
        part_scope = {r.part_id for r in self._rows_raw}
        if not part_scope:
            QMessageBox.information(self, "Publish Test Defaults", "No parts are available to publish.")
            return

        plans: list[dict[str, object]] = []
        errors: list[tuple[int, str]] = []
        missing_complex: set[int] = set()

        for part_id in sorted(part_scope):
            assignment = self._test_assignments.get(part_id)
            if not assignment:
                continue
            self._ensure_assignment_keys(assignment)
            unpowered_sel, err = self._collect_publish_selection(part_id, TestMode.unpowered, assignment)
            powered_sel, err_powered = self._collect_publish_selection(part_id, TestMode.powered, assignment)
            if err:
                errors.append((part_id, err))
            if err_powered:
                errors.append((part_id, err_powered))
            if not unpowered_sel and not powered_sel:
                continue

            row = self._rows_by_part.get(part_id)
            pn = self._part_numbers.get(part_id) or getattr(row, "part_number", str(part_id))
            part_type = getattr(row, "active_passive", None)
            profile = TestProfile.PASSIVE
            if isinstance(part_type, str) and part_type.lower() == "active":
                profile = TestProfile.ACTIVE

            plan = {
                "part_id": part_id,
                "pn": pn,
                "profile": profile,
                "selections": {
                    TestMode.unpowered: unpowered_sel,
                    TestMode.powered: powered_sel,
                },
            }

            if unpowered_sel and unpowered_sel.get("method") == "Complex" and part_id not in self._complex_links:
                missing_complex.add(part_id)
            if powered_sel and powered_sel.get("method") == "Complex" and part_id not in self._complex_links:
                missing_complex.add(part_id)

            plans.append(plan)

        if errors:
            lines = []
            for part_id, message in errors:
                pn = self._part_numbers.get(part_id) or str(part_id)
                lines.append(f"PN {pn}: {message}")
            QMessageBox.warning(
                self,
                "Publish Test Defaults",
                "Unable to publish due to the following issues:\n\n" + "\n".join(lines),
            )
            return

        if missing_complex:
            pns = [self._part_numbers.get(pid) or str(pid) for pid in sorted(missing_complex)]
            QMessageBox.warning(
                self,
                "Publish Test Defaults",
                "Link the following parts to Complex Editor before publishing:\n\n" + "\n".join(pns),
            )
            self._ensure_complex_panel_visible_for_part(next(iter(missing_complex)))
            return

        actionable = [
            plan
            for plan in plans
            if any(
                sel
                for sel in plan["selections"].values()
                if sel and sel.get("method")
            )
        ]

        if not actionable:
            QMessageBox.information(
                self,
                "Publish Test Defaults",
                "No staged test defaults to publish.",
            )
            return

        summary_lines: list[str] = []
        for plan in actionable:
            ops: list[str] = []
            for mode, label in ((TestMode.unpowered, "unpowered"), (TestMode.powered, "powered")):
                sel = plan["selections"].get(mode)
                if not sel or not sel.get("method"):
                    continue
                method = sel["method"]
                detail = sel.get("detail")
                if method == "Complex":
                    ops.append("Complex (linked)")
                else:
                    if detail:
                        ops.append(f"{label} {method} ({detail})")
                    else:
                        ops.append(f"{label} {method}")
            if ops:
                summary_lines.append(f"PN {plan['pn']}: " + ", ".join(ops))

        confirm_message = f"This will publish test defaults for {len(actionable)} part(s)."
        if summary_lines:
            confirm_message += "\n\n" + "\n".join(summary_lines)

        confirm = QMessageBox.question(
            self,
            "Publish Test Defaults",
            confirm_message,
            QMessageBox.StandardButton.Ok,
            QMessageBox.StandardButton.Cancel,
        )
        if confirm != QMessageBox.StandardButton.Ok:
            return

        successes: list[str] = []
        failures: list[tuple[str, str]] = []

        with app_state.get_session() as session:
            for plan in actionable:
                try:
                    for mode, sel in plan["selections"].items():
                        if not sel or not sel.get("method"):
                            continue
                        if not sel.get("persist", True):
                            continue
                        services.upsert_part_test_map(
                            session,
                            plan["part_id"],
                            mode,
                            plan["profile"],
                            sel["method"],
                            sel.get("detail"),
                            sel.get("qt_path"),
                        )
                    session.commit()
                    successes.append(str(plan["pn"]))
                except Exception as exc:
                    session.rollback()
                    failures.append((str(plan["pn"]), str(exc)))

        if failures:
            lines = [f"PN {pn}: {error}" for pn, error in failures]
            if successes:
                lines.insert(0, f"Published {len(successes)} part(s).")
            QMessageBox.warning(
                self,
                "Publish Test Defaults",
                "Some defaults could not be published:\n\n" + "\n".join(lines),
            )
            return

        QMessageBox.information(
            self,
            "Publish Test Defaults",
            f"Published test defaults for {len(successes)} part(s).",
        )

    def _reload_complex_links(self, part_ids: Iterable[int]) -> None:
        ids: Set[int] = {int(pid) for pid in part_ids if isinstance(pid, int)}
        if not ids:
            return
        try:
            from sqlmodel import select
            from ..domain.complex_linker import ComplexLink

            with app_state.get_session() as session:
                rows = session.exec(
                    select(ComplexLink).where(ComplexLink.part_id.in_(ids))
                ).all()
        except Exception as exc:  # pragma: no cover - defensive
            logging.debug("Failed to refresh complex links: %s", exc)
            return
        for pid in ids:
            self._complex_links.pop(pid, None)
        for row in rows:
            try:
                self._complex_links[row.part_id] = row
            except Exception:  # pragma: no cover - defensive
                logging.debug("Skipping malformed complex link row: %s", row)

    def _show_stub_dialog(self, message: str) -> None:
        from .dialogs.tm_stub_dialog import TestMethodStubDialog

        dlg = TestMethodStubDialog(message, self)
        dlg.exec()

    # ------------------------------------------------------------------
    def _detail_text_for(
        self,
        method: str,
        detail: Optional[str],
        qt_path: Optional[str],
    ) -> str:
        method = method or ""
        if method == "Macro":
            sel = detail or None
            return sel or "Choose Macro..."
        if method == "Complex":
            return "Link Complex..."
        if method == "Quick test (QT)":
            if qt_path:
                return Path(qt_path).name
            return "Select QT XML..."
        if method == "Python code":
            return detail or "Open Python Project..."
        return ""

    def _handle_auto_infer_persistence(self, part_id: int, inferred: Optional[str], explicit: Optional[str]) -> None:
        # Persist/stage only when there is no explicit value in DB and we inferred a mode
        if explicit in ("active", "passive"):
            # keep parts_state from DB
            self._parts_state[part_id] = explicit
            return
        self._parts_state.setdefault(part_id, inferred)
        if inferred in ("active", "passive"):
            if self.apply_act.isChecked():
                # Apply immediately
                try:
                    with app_state.get_session() as session:
                        services.update_part_active_passive(session, part_id, inferred)
                except Exception:
                    # Do not block UI; leave as inferred only
                    return
                else:
                    self._parts_state[part_id] = inferred
            else:
                # Stage for save
                self._dirty_parts[part_id] = inferred
                self.save_act.setEnabled(True)

    # ------------------------------------------------------------------
    def _on_value_changed(self, part_id: int, value: Optional[str]) -> None:
        # Determine old value from tracked state
        old_value = self._parts_state.get(part_id, None)
        # Update all rows with this part_id in the current model
        ap_col = self._col_indices["ap"]
        for row in range(self.model.rowCount()):
            pid = self.model.data(self.model.index(row, 0), PartIdRole)
            # Some models put PN at col 0, others Reference. PartIdRole is on all cells
            if pid != part_id:
                # try any col
                matched = False
                for c in range(self.model.columnCount()):
                    if self.model.data(self.model.index(row, c), PartIdRole) == part_id:
                        matched = True
                        break
                if not matched:
                    continue
            self.model.setData(self.model.index(row, ap_col), value or "")

        if self.apply_act.isChecked():
            # Persist immediately
            try:
                if value not in ("active", "passive"):
                    # Treat empty click as passive transition handled by delegate; here guard
                    value = "passive"
                with app_state.get_session() as session:
                    services.update_part_active_passive(session, part_id, value)
            except Exception as exc:  # pragma: no cover - depends on DB
                QMessageBox.warning(self, "Update failed", str(exc))
                # revert UI
                for row in range(self.model.rowCount()):
                    for c in range(self.model.columnCount()):
                        if self.model.data(self.model.index(row, c), PartIdRole) == part_id:
                            self.model.setData(self.model.index(row, ap_col), old_value or "")
                return
            self._parts_state[part_id] = value
        else:
            # Stage for save
            if value not in ("active", "passive"):
                value = "passive"
            self._dirty_parts[part_id] = value
            self.save_act.setEnabled(True)

    def _fanout_part_field(self, part_id: int, field: str, value: Optional[str]) -> None:
        col = self._col_indices.get(field)
        if col is None:
            return
        self._updating = True
        for row in range(self.model.rowCount()):
            if self.model.data(self.model.index(row, col), PartIdRole) == part_id:
                self.model.setData(self.model.index(row, col), value or "")
        self._updating = False

    def _persist_or_stage_package(self, part_id: int, value: Optional[str]) -> None:
        if self.apply_act.isChecked():
            try:
                with app_state.get_session() as session:
                    services.update_part_package(session, part_id, value or "")
            except Exception as exc:
                QMessageBox.warning(self, "Update failed", str(exc))
                old_val = self._part_packages.get(part_id, None)
                self._fanout_part_field(part_id, "package", old_val)
                return
            self._part_packages[part_id] = value or None
            self._dirty_packages.pop(part_id, None)
        else:
            self._dirty_packages[part_id] = value or None
            self.save_act.setEnabled(True)

    def _persist_or_stage_value(self, part_id: int, value: Optional[str]) -> None:
        if self.apply_act.isChecked():
            try:
                with app_state.get_session() as session:
                    services.update_part_value(session, part_id, value or "")
            except Exception as exc:
                QMessageBox.warning(self, "Update failed", str(exc))
                old_val = self._part_values.get(part_id, None)
                self._fanout_part_field(part_id, "value", old_val)
                return
            self._part_values[part_id] = value or None
            self._dirty_values.pop(part_id, None)
        else:
            self._dirty_values[part_id] = value or None
            self.save_act.setEnabled(True)

    def _persist_or_stage_description(self, part_id: int, value: Optional[str]) -> None:
        if self.apply_act.isChecked():
            try:
                with app_state.get_session() as session:
                    services.update_part_description(session, part_id, value or "")
            except Exception as exc:
                QMessageBox.warning(self, "Update failed", str(exc))
                # Revert in UI from DB snapshot
                self._fanout_part_field(part_id, "desc", None)
                return
            self._dirty_desc.pop(part_id, None)
        else:
            self._dirty_desc[part_id] = value or None
            self.save_act.setEnabled(True)

    def _persist_or_stage_manufacturer(self, part_id: int, value: Optional[str]) -> None:
        if self.apply_act.isChecked():
            try:
                with app_state.get_session() as session:
                    services.update_manufacturer_for_part_in_assembly(
                        session, self._assembly_id, part_id, value or ""
                    )
            except Exception as exc:
                QMessageBox.warning(self, "Update failed", str(exc))
                # No reliable previous value to restore; leave UI as-is
                return
            self._dirty_mfg.pop(part_id, None)
        else:
            self._dirty_mfg[part_id] = value or None
            self.save_act.setEnabled(True)

    def _persist_or_stage_tolerances(
        self, part_id: int, tol_p: Optional[str], tol_n: Optional[str]
    ) -> None:
        if self.apply_act.isChecked():
            try:
                with app_state.get_session() as session:
                    services.update_part_tolerances(session, part_id, tol_p, tol_n)
            except Exception as exc:
                QMessageBox.warning(self, "Update failed", str(exc))
                old = self._tolerances.get(part_id, (None, None))
                self._fanout_part_field(part_id, "tol_p", old[0])
                self._fanout_part_field(part_id, "tol_n", old[1])
                return
            self._tolerances[part_id] = (tol_p, tol_n)
            self._dirty_tolerances.pop(part_id, None)
        else:
            self._tolerances[part_id] = (tol_p, tol_n)
            self._dirty_tolerances[part_id] = (tol_p, tol_n)
            self.save_act.setEnabled(True)

    def _on_item_changed(self, item: QStandardItem) -> None:
        if self._updating:
            return
        part_id = item.data(PartIdRole)
        if part_id is None:
            return
        col = item.column()
        text = (item.text() or "").strip() or None
        if col == self._col_indices.get("package"):
            self._part_packages[part_id] = text
            self._fanout_part_field(part_id, "package", text)
            self._persist_or_stage_package(part_id, text)
        elif col == self._col_indices.get("value"):
            self._part_values[part_id] = text
            self._fanout_part_field(part_id, "value", text)
            self._persist_or_stage_value(part_id, text)
        elif col == self._col_indices.get("desc"):
            self._fanout_part_field(part_id, "desc", text)
            self._persist_or_stage_description(part_id, text)
        elif col == self._col_indices.get("mfg"):
            self._fanout_part_field(part_id, "mfg", text)
            self._persist_or_stage_manufacturer(part_id, text)
        elif col == self._col_indices.get("tol_p"):
            cur = self._tolerances.get(part_id, (None, None))
            new_pair = (text, cur[1])
            self._tolerances[part_id] = new_pair
            self._fanout_part_field(part_id, "tol_p", text)
            self._persist_or_stage_tolerances(part_id, new_pair[0], new_pair[1])
        elif col == self._col_indices.get("tol_n"):
            cur = self._tolerances.get(part_id, (None, None))
            new_pair = (cur[0], text)
            self._tolerances[part_id] = new_pair
            self._fanout_part_field(part_id, "tol_n", text)
            self._persist_or_stage_tolerances(part_id, new_pair[0], new_pair[1])

    def _extract_prefix(self, ref: str) -> str:
        i = 0
        while i < len(ref) and ref[i].isalpha():
            i += 1
        return ref[:i].upper()

    def _macro_for_reference(self, reference: str) -> str | None:
        if not reference:
            return None
        pfx = self._extract_prefix(reference)
        if not pfx:
            return None
        for key, macro in self._prefix_macros:
            if pfx.startswith(key):
                canon = self._function_options_canon.get(macro.upper())
                return canon
        return None

    def _autofill_fields(self) -> None:
        proxy = self.table.model()
        if proxy is None:
            return
        pn_col = self._col_indices.get("pn")
        desc_col = self._col_indices.get("desc")
        rows = proxy.rowCount()
        for r in range(rows):
            pn = str(proxy.data(proxy.index(r, pn_col)) or "")
            desc = str(proxy.data(proxy.index(r, desc_col)) or "")
            part_id = proxy.data(proxy.index(r, pn_col), PartIdRole)
            if part_id is None:
                part_id = proxy.data(proxy.index(r, desc_col), PartIdRole)
            if part_id is None:
                continue
            ar = infer_from_pn_and_desc(pn, desc)
            if ar.package and not self._part_packages.get(part_id):
                self._part_packages[part_id] = ar.package
                self._fanout_part_field(part_id, "package", ar.package)
                self._persist_or_stage_package(part_id, ar.package)
            if ar.value and not self._part_values.get(part_id):
                self._part_values[part_id] = ar.value
                self._fanout_part_field(part_id, "value", ar.value)
                self._persist_or_stage_value(part_id, ar.value)
            if ar.tol_pos and ar.tol_neg:
                cur_p, cur_n = self._tolerances.get(part_id, (None, None))
                if not cur_p and not cur_n:
                    pair = (ar.tol_pos, ar.tol_neg)
                    self._tolerances[part_id] = pair
                    self._fanout_part_field(part_id, "tol_p", pair[0])
                    self._fanout_part_field(part_id, "tol_n", pair[1])
                    self._persist_or_stage_tolerances(part_id, pair[0], pair[1])
            method_key, _detail_key = self._relevant_test_column_keys()
            tm_col = self._col_indices.get(method_key)
            ref_col = self._col_indices.get("ref")
            if tm_col is not None and ref_col is not None:
                tm_text = str(proxy.data(proxy.index(r, tm_col)) or "").strip()
                if not tm_text:
                    ref_text = str(proxy.data(proxy.index(r, ref_col)) or "")
                    if self._view_mode == "by_pn":
                        first_ref = ref_text.split(",")[0].strip() if ref_text else ""
                    else:
                        first_ref = ref_text
                    macro = self._macro_for_reference(first_ref)
                    if macro:
                        ta = self._test_assignments.setdefault(part_id, {"method": "", "qt_path": None})
                        ta["method"] = "Macro"
                        ta["detail"] = macro
                        ta["qt_path"] = None
                        self._refresh_rows_for_part(part_id)
                        if self.apply_act.isChecked():
                            self._persist_test_assignment(part_id)
                        else:
                            self._dirty_tests.add(part_id)
                            self.save_act.setEnabled(True)

    def _auto_datasheet(self) -> None:
        proxy = self.table.model()
        sel = self.table.selectionModel()
        if proxy is None or sel is None or not sel.selectedIndexes():
            QMessageBox.information(self, "Auto Datasheet", "Select one or more rows first.")
            return
        rows = sorted({i.row() for i in sel.selectedIndexes()})
        pn_col = self._col_indices.get("pn")
        desc_col = self._col_indices.get("desc")
        mfg_col = self._col_indices.get("mfg")
        work = []
        for r in rows:
            pn = str(proxy.data(proxy.index(r, pn_col)) or "")
            desc = str(proxy.data(proxy.index(r, desc_col)) or "")
            mfg = str(proxy.data(proxy.index(r, mfg_col)) or "")
            pid = proxy.data(proxy.index(r, pn_col), PartIdRole) or proxy.data(proxy.index(r, desc_col), PartIdRole)
            if pid:
                from .auto_datasheet_dialog import WorkItem, AutoDatasheetDialog
                work.append(WorkItem(part_id=pid, pn=pn, mfg=mfg, desc=desc))
        if not work:
            QMessageBox.warning(self, "Auto Datasheet", "No resolvable parts in selection.")
            return
        logging.info("BOMEditor: starting Auto Datasheet for %d parts", len(work))
        # Mark selected parts as in-progress and update UI
        self._datasheet_loading |= {w.part_id for w in work}
        QTimer.singleShot(0, self._install_datasheet_widgets)
        dlg = AutoDatasheetDialog(self, work, on_locked_parts_changed=self._set_parts_locked)
        # Update UI immediately for each part as it's attached
        dlg.attached.connect(self._on_datasheet_attached)
        # Mark parts as searched-but-not-found in this BOM session
        dlg.failed.connect(self._on_datasheet_failed)
        # Receive manual page suggestions for hard downloads
        dlg.manualLink.connect(self._on_manual_link)
        dlg.exec()
        # Refresh datasheet paths for affected parts
        from ..models import Part
        with app_state.get_session() as session:
            for w in work:
                p = session.get(Part, w.part_id)
                if p:
                    self._part_datasheets[w.part_id] = p.datasheet_url
                    self._part_product_links[w.part_id] = getattr(p, "product_url", None)
                    # Update description live if inferred/filled during auto-datasheet
                    self._fanout_part_field(w.part_id, "desc", getattr(p, "description", None))
        # Clear loading state and refresh icons
        self._datasheet_loading -= {w.part_id for w in work}
        QTimer.singleShot(0, self._install_datasheet_widgets)

    def _set_parts_locked(self, parts: set[int], lock: bool):
        if lock:
            self._locked_parts |= set(parts)
        else:
            self._locked_parts -= set(parts)
        self._rebuild_model()

    def _on_table_selection_changed(self, *_args) -> None:
        self._update_auto_ds_act()
        self._update_complex_panel_context()
        self._update_schematic_panel_context()

    def _update_auto_ds_act(self) -> None:
        sel = self.table.selectionModel()
        self.auto_ds_act.setEnabled(bool(sel and sel.selectedIndexes()))

    def _selected_part_id(self) -> Optional[int]:
        sel = self.table.selectionModel()
        if not sel:
            return None
        for idx in sel.selectedIndexes():
            part_id = idx.data(PartIdRole)
            if isinstance(part_id, int):
                return part_id
        return None

    def _part_number_for_part(self, part_id: int) -> str:
        pn = self._part_numbers.get(part_id) or ""
        if pn:
            return pn
        for row in self._rows_raw:
            if getattr(row, 'part_id', None) == part_id:
                pn = getattr(row, 'part_number', '') or ''
                if pn:
                    self._part_numbers[part_id] = pn
                return pn
        return ''

    def _ensure_complex_panel_visible_for_part(self, part_id: int) -> None:
        if not self._complex_panel:
            return
        pn = self._part_number_for_part(part_id)
        self._complex_panel.set_context(part_id, pn)
        self._complex_panel.setVisible(True)
        if self._complex_splitter:
            sizes = self._complex_splitter.sizes()
            if len(sizes) == 2 and sizes[1] == 0:
                total = sum(sizes) or self.width() or 1
                self._complex_splitter.setSizes([max(total - 360, 320), 360])
        self._complex_panel.search_edit.setFocus()

    def _update_complex_panel_context(self) -> None:
        if not self._complex_panel:
            return
        part_id = self._selected_part_id()
        if part_id is None:
            self._complex_panel.set_context(None, None)
            self._complex_panel.hide()
            return
        assignment = self._get_assignment(part_id)
        method_unpowered = assignment.get("method", "")
        method_powered = assignment.get("method_powered", "")
        if method_unpowered != 'Complex' and method_powered != 'Complex':
            self._complex_panel.set_context(None, None)
            self._complex_panel.hide()
            return
        self._ensure_complex_panel_visible_for_part(part_id)

    def _selected_reference_text(self) -> str:
        ref_col = self._col_indices.get("ref")
        sel = self.table.selectionModel()
        if ref_col is None or not sel:
            return ""
        indexes = sel.selectedIndexes()
        if not indexes:
            return ""
        source_index = self.proxy.mapToSource(indexes[0])
        idx = self.model.index(source_index.row(), ref_col)
        ref = self.model.data(idx)
        return str(ref or "")

    def _update_schematic_panel_context(self) -> None:
        if not self._schematic_panel or not self.view_schematics_act.isChecked():
            return
        part_id = self._selected_part_id()
        if part_id is None:
            self._schematic_panel.clear_component_context()
            return
        pn = self._part_number_for_part(part_id)
        reference = self._selected_reference_text()
        self._schematic_panel.set_component_context(part_id, pn, reference)

    def _toggle_schematic_panel(self, checked: bool) -> None:
        self._settings.setValue("schematics_visible", bool(checked))
        self._set_schematic_panel_visible(checked)

    def _set_schematic_panel_visible(self, visible: bool) -> None:
        if not self._schematic_panel or not self._main_splitter:
            return
        self._schematic_panel.setVisible(visible)
        if visible:
            sizes = self._main_splitter.sizes()
            if len(sizes) == 2 and sizes[1] == 0:
                total = sum(sizes) or max(self.height(), 1)
                self._main_splitter.setSizes([int(total * 0.65), int(total * 0.35)])
            self._schematic_panel.refresh()
            self._update_schematic_panel_context()
        else:
            try:
                total = sum(self._main_splitter.sizes()) or 1
                self._main_splitter.setSizes([total, 0])
            except Exception:
                pass
            self._schematic_panel.clear_component_context()

    def _on_complex_link_updated(self, part_id: int) -> None:
        self._reload_complex_links([part_id])
        self._refresh_rows_for_part(part_id)
        self._ensure_complex_panel_visible_for_part(part_id)
        self._install_datasheet_widgets()

    def _on_apply_toggled(self, checked: bool) -> None:
        if checked and (
            self._dirty_parts
            or self._dirty_packages
            or self._dirty_values
            or self._dirty_tolerances
            or self._dirty_datasheets
        ):
            self._save_changes()

    def _on_lines_changed(self, lines: int) -> None:
        """Adjust row heights when the visible line count changes."""
        self.wrap_delegate.set_line_count(lines)
        self.table.resizeRowsToContents()
        self._settings.setValue("lines", lines)

    def _save_changes(self) -> None:
        if (
            not self._dirty_parts
            and not self._dirty_packages
            and not self._dirty_values
            and not self._dirty_tests
            and not self._dirty_tolerances
            and not getattr(self, "_dirty_desc", {})
            and not getattr(self, "_dirty_mfg", {})
            and not getattr(self, "_dirty_links", {})
            and not getattr(self, "_dirty_datasheets", {})
        ):
            return
        failures = []
        for part_id, value in list(self._dirty_parts.items()):
            try:
                with app_state.get_session() as session:
                    services.update_part_active_passive(session, part_id, value)
            except Exception as exc:
                failures.append(str(exc))
                # revert
                old_val = self._parts_state.get(part_id, None)
                for row in range(self.model.rowCount()):
                    for c in range(self.model.columnCount()):
                        if self.model.data(self.model.index(row, c), PartIdRole) == part_id:
                            self.model.setData(self.model.index(row, self._col_indices["ap"]), old_val or "")
            else:
                self._parts_state[part_id] = value
                del self._dirty_parts[part_id]
        for part_id, pkg in list(self._dirty_packages.items()):
            try:
                with app_state.get_session() as session:
                    services.update_part_package(session, part_id, pkg or "")
            except Exception as exc:
                failures.append(str(exc))
                old_val = self._part_packages.get(part_id, None)
                self._fanout_part_field(part_id, "package", old_val)
            else:
                self._part_packages[part_id] = pkg or None
                del self._dirty_packages[part_id]
        for part_id, val in list(self._dirty_values.items()):
            try:
                with app_state.get_session() as session:
                    services.update_part_value(session, part_id, val or "")
            except Exception as exc:
                failures.append(str(exc))
                old_val = self._part_values.get(part_id, None)
                self._fanout_part_field(part_id, "value", old_val)
            else:
                self._part_values[part_id] = val or None
                del self._dirty_values[part_id]
        for part_id, (tp, tn) in list(self._dirty_tolerances.items()):
            try:
                with app_state.get_session() as session:
                    services.update_part_tolerances(session, part_id, tp, tn)
            except Exception as exc:
                failures.append(str(exc))
                old = self._tolerances.get(part_id, (None, None))
                self._fanout_part_field(part_id, "tol_p", old[0])
                self._fanout_part_field(part_id, "tol_n", old[1])
            else:
                self._tolerances[part_id] = (tp, tn)
                del self._dirty_tolerances[part_id]
        # Persist edited descriptions
        for part_id, desc in list(getattr(self, "_dirty_desc", {}).items()):
            try:
                with app_state.get_session() as session:
                    services.update_part_description(session, part_id, desc or "")
            except Exception as exc:
                failures.append(str(exc))
            else:
                del self._dirty_desc[part_id]
        # Persist edited manufacturers (apply to all BOM items for this part in this assembly)
        for part_id, mfg in list(getattr(self, "_dirty_mfg", {}).items()):
            try:
                with app_state.get_session() as session:
                    services.update_manufacturer_for_part_in_assembly(session, self._assembly_id, part_id, mfg or "")
            except Exception as exc:
                failures.append(str(exc))
            else:
                del self._dirty_mfg[part_id]
        # Persist test assignments via settings
        for part_id in list(self._dirty_tests):
            try:
                self._persist_test_assignment(part_id)
            except Exception as exc:
                failures.append(str(exc))
            else:
                self._dirty_tests.discard(part_id)
        # Persist product links
        for part_id, link in list(self._dirty_links.items()):
            try:
                with app_state.get_session() as session:
                    services.update_part_product_url(session, part_id, link)
            except Exception as exc:
                failures.append(str(exc))
            else:
                self._part_product_links[part_id] = link
                del self._dirty_links[part_id]
        # Persist datasheet add/removes
        for part_id, ds in list(self._dirty_datasheets.items()):
            try:
                with app_state.get_session() as session:
                    if ds:
                        services.update_part_datasheet_url(session, part_id, ds)
                    else:
                        services.remove_part_datasheet(session, part_id, delete_file=True)
            except Exception as exc:
                failures.append(str(exc))
            else:
                self._part_datasheets[part_id] = ds or None
                del self._dirty_datasheets[part_id]
        if failures:
            QMessageBox.warning(self, "Save failed", "; ".join(failures))
        else:
            QMessageBox.information(self, "Saved", "Changes saved.")
        self.save_act.setEnabled(
            bool(self._dirty_parts)
            or bool(self._dirty_packages)
            or bool(self._dirty_values)
            or bool(self._dirty_tolerances)
            or bool(self._dirty_tests)
            or bool(getattr(self, "_dirty_desc", {}))
            or bool(getattr(self, "_dirty_mfg", {}))
            or bool(getattr(self, "_dirty_links", {}))
            or bool(getattr(self, "_dirty_datasheets", {}))
        )

    def _on_view_mode_changed(self, mode: str, checked: bool) -> None:
        if not checked:
            return
        if mode not in ("by_pn", "by_ref"):
            return
        self._view_mode = mode
        self._settings.setValue("view_mode", self._view_mode)
        # Rebuild model and columns menu to reflect new structure
        self._rebuild_model()

    def _autosize_window_to_columns(self) -> None:
        try:
            self.table.resizeColumnsToContents()
            total_w = self.table.verticalHeader().width() + (self.table.frameWidth() * 2)
            for c in range(self.model.columnCount()):
                if self.table.isColumnHidden(c):
                    continue
                w = self.table.columnWidth(c)
                total_w += w
            # Add a bit for scrollbar and padding
            total_w += 24
            # Keep a reasonable max to avoid huge windows
            total_w = min(max(total_w, 900), 1800)
            self.resize(total_w, max(self.height(), 600))
        except Exception:
            # Best-effort sizing; ignore issues in headless envs
            pass

    # ------------------------------------------------------------------
    def _collect_table_rows(self) -> tuple[list[dict], Set[int]]:
        """Return visible table rows and the set of part ids they reference."""

        model = self.table.model()
        if model is None:
            return [], set()
        method_key, detail_key = self._relevant_test_column_keys()
        tm_col = self._col_indices.get(method_key)
        td_col = self._col_indices.get(detail_key)
        ref_col = self._col_indices.get("ref")
        pn_col = self._col_indices.get("pn")
        rows: list[dict] = []
        part_scope: Set[int] = set()
        for r in range(model.rowCount()):
            pn_idx = model.index(r, pn_col) if pn_col is not None else None
            ref_idx = model.index(r, ref_col) if ref_col is not None else None
            tm_idx = model.index(r, tm_col) if tm_col is not None else None
            td_idx = model.index(r, td_col) if td_col is not None else None
            pn = model.data(pn_idx) if pn_idx is not None else ""
            refs = model.data(ref_idx) if ref_idx is not None else ""
            tm = model.data(tm_idx) if tm_idx is not None else ""
            td = model.data(td_idx) if td_idx is not None else ""
            part_id = None
            if pn_idx is not None:
                part_id = model.data(pn_idx, PartIdRole)
            if part_id is None and ref_idx is not None:
                part_id = model.data(ref_idx, PartIdRole)
            if isinstance(part_id, int):
                part_scope.add(part_id)
            if self._view_mode == "by_pn":
                ref_list = [x.strip() for x in str(refs or "").split(",") if x.strip()]
                for ref in ref_list:
                    rows.append(
                        {
                            "reference": ref,
                            "part_number": pn or "",
                            "test_method": tm or "",
                            "test_detail": td or "",
                        }
                    )
            else:
                rows.append(
                    {
                        "reference": str(refs or "").strip(),
                        "part_number": pn or "",
                        "test_method": tm or "",
                        "test_detail": td or "",
                    }
                )
        return rows, part_scope

    def _on_export_viva(self) -> None:  # pragma: no cover - UI glue
        table_rows, part_scope = self._collect_table_rows()
        self._reload_complex_links(part_scope)
        with app_state.get_session() as session:
            try:
                rows = services.build_viva_groups(table_rows, session, self._assembly_id)
            except ValueError as exc:
                QMessageBox.warning(self, "Cannot export", str(exc))
                return

            asm, proj = self._load_assembly_context()
            strict_mode = True
            viva_settings = config.get_viva_export_settings()
            force_prompt = bool(
                QApplication.keyboardModifiers() & Qt.KeyboardModifier.ShiftModifier
            )
            base_root = self._determine_export_base_root(
                proj,
                force_prompt=force_prompt,
                viva_settings=viva_settings,
            )
            if base_root is None:
                return
            try:
                base_root.mkdir(parents=True, exist_ok=True)
            except OSError as exc:
                detail = os.strerror(exc.errno) if getattr(exc, "errno", None) else str(exc)
                message = (
                    f"Could not prepare the export folder:\n{base_root}\n"
                    f"{detail}"
                )
                QMessageBox.warning(self, "Export failed", message)
                return
            except Exception as exc:
                QMessageBox.warning(self, "Export failed", f"Cannot create export root: {exc}")
                return

            while True:
                try:
                    result = services.perform_viva_export(
                        session,
                        self._assembly_id,
                        base_dir=base_root,
                        bom_rows=rows,
                        strict=strict_mode,
                    )
                except services.VIVAExportValidationError as exc:
                    self._show_unlinked_components(exc.missing)
                    return
                except CEPNResolutionError as exc:
                    if not exc.unresolved:
                        QMessageBox.warning(
                            self,
                            "VIVA export",
                            "Complex Editor could not resolve some part numbers.",
                        )
                        return
                    if not self._prompt_unresolved_pns(exc.unresolved):
                        return
                    strict_mode = False
                    continue
                except CEExportBusyError as exc:
                    retry = QMessageBox.question(
                        self,
                        "Complex Editor busy",
                        f"{exc}\nRetry the export?",
                        QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel,
                        QMessageBox.StandardButton.Retry,
                    )
                    if retry == QMessageBox.StandardButton.Retry:
                        continue
                    logging.info(
                        "VIVA export cancelled due to CE busy: assembly=%s",
                        self._assembly_id,
                    )
                    return
                except CEAuthError as exc:
                    QMessageBox.warning(self, "Complex Editor", str(exc))
                    return
                except CEExportError as exc:
                    self._show_ce_export_error(exc)
                    return
                except CENetworkError as exc:
                    self._show_export_error(str(exc))
                    return
                else:
                    ce_summary: Optional[Dict[str, object]] = None
                    ce_folder = result.paths.folder / "CE"
                    try:
                        ce_summary = services.export_bom_to_ce_bridge(
                            session,
                            self._assembly_id,
                            out_dir=ce_folder,
                        )
                    except Exception as exc:  # pragma: no cover - defensive
                        logging.exception(
                            "Complex Editor export crashed: assembly=%s", self._assembly_id
                        )
                        ce_summary = {
                            "status": "FAILED_BACKEND",
                            "trace_id": "",
                            "export_path": str(ce_folder),
                            "exported_count": 0,
                            "missing_count": 0,
                            "report_path": "",
                            "detail": f"Unexpected error running Complex Editor export: {exc}",
                        }
                    self._finalize_viva_success(
                        result,
                        base_root,
                        ce_summary=ce_summary,
                    )
                    return

    def _export_excel(self) -> None:  # pragma: no cover - UI glue
        # Save visible table to an .xlsx, preserving 'Link' hyperlinks; skip 'Datasheet' column
        from openpyxl import Workbook
        from openpyxl.styles import Font

        base = self._default_export_basename()
        default_name = f"{base} - BOM.xlsx" if base else f"BOM_{self._assembly_id}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Export BOM to Excel", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        proxy = self.table.model()
        if proxy is None:
            return
        ds_col = self._col_indices.get("ds")
        link_col = self._col_indices.get("link")
        # Build list of visible columns (skip 'Datasheet')
        export_cols: list[int] = []
        headers: list[str] = []
        for c in range(self.model.columnCount()):
            if self.table.isColumnHidden(c):
                continue
            if ds_col is not None and c == ds_col:
                continue
            export_cols.append(c)
            hi = self.model.horizontalHeaderItem(c)
            headers.append(hi.text() if hi else str(c))
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(headers)
        # Export visible rows in proxy order
        rows = proxy.rowCount()
        for r in range(rows):
            # First append values to keep shapes; then add hyperlinks for link column
            values = []
            for c in export_cols:
                idx = proxy.index(r, c)
                if link_col is not None and c == link_col:
                    link_val = proxy.data(idx, LinkUrlRole)
                    values.append(str(link_val or ""))
                else:
                    values.append(str(proxy.data(idx) or ""))
            ws.append(values)
            # Set hyperlink if link column included
            if link_col is not None and link_col in export_cols:
                idx = export_cols.index(link_col)
                url = values[idx]
                if isinstance(url, str) and url.lower().startswith(("http://", "https://")):
                    cell = ws.cell(row=ws.max_row, column=idx + 1)
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
        try:
            wb.save(path)
        except Exception as exc:
            QMessageBox.warning(self, "Export failed", str(exc))
            return
        QMessageBox.information(self, "Export", f"Saved Excel to {path}")

    def _default_export_basename(self) -> str:
        """Return 'Customer - Project' for current assembly, sanitized for filenames."""
        import re
        try:
            from ..models import Assembly, Project, Customer
            with app_state.get_session() as session:
                asm = session.get(Assembly, self._assembly_id)
                if not asm:
                    return ""
                proj = session.get(Project, asm.project_id) if getattr(asm, "project_id", None) else None
                cust = session.get(Customer, proj.customer_id) if proj and getattr(proj, "customer_id", None) else None
                cust_name = (cust.name if cust else "").strip()
                proj_title = (proj.title if proj else "").strip() or (getattr(proj, "name", "") or "").strip()
                base = " - ".join([p for p in (cust_name, proj_title) if p])
                # sanitize for filenames
                base = re.sub(r"[\\/:*?\"<>|]+", "_", base)[:120]
                return base
        except Exception:
            return ""

    def _load_assembly_context(self):
        try:
            from ..models import Assembly, Project

            with app_state.get_session() as session:
                asm = session.get(Assembly, self._assembly_id)
                proj = None
                if asm and getattr(asm, "project_id", None):
                    proj = session.get(Project, asm.project_id)
            return asm, proj
        except Exception as exc:  # pragma: no cover - defensive
            logging.debug("Failed to load assembly context: %s", exc)
            return None, None

    def _project_default_root(self, project) -> Optional[Path]:
        code = getattr(project, "code", None) if project else None
        if not code:
            return None
        sanitized = services.sanitize_token(code, str(code))
        return (DATA_ROOT / "Projects" / sanitized).resolve()

    def _determine_export_base_root(
        self,
        project,
        *,
        force_prompt: bool = False,
        viva_settings: Optional[Dict[str, object]] = None,
    ) -> Optional[Path]:
        stored_path = None
        if viva_settings:
            last_path = viva_settings.get("last_export_path")
            if isinstance(last_path, str) and last_path.strip():
                try:
                    stored_path = Path(last_path).expanduser().resolve()
                except Exception:
                    stored_path = Path(last_path).expanduser()
        if stored_path is None:
            stored_raw = self._export_settings.value("last_root", "")
            if isinstance(stored_raw, str) and stored_raw.strip():
                try:
                    stored_path = Path(stored_raw).expanduser().resolve()
                except Exception:
                    stored_path = Path(stored_raw).expanduser()
        if stored_path is not None and not force_prompt:
            return stored_path
        default_root = self._project_default_root(project)
        if default_root is None and viva_settings:
            cfg_dir = viva_settings.get("viva_export_base_dir")
            if isinstance(cfg_dir, str) and cfg_dir.strip():
                try:
                    default_root = Path(cfg_dir).expanduser().resolve()
                except Exception:
                    default_root = Path(cfg_dir).expanduser()
        if default_root is not None and not force_prompt:
            return default_root
        start = default_root or stored_path or Path(viva_settings.get("viva_export_base_dir") if viva_settings else DATA_ROOT)
        if not isinstance(start, Path):
            start = DATA_ROOT
        directory = QFileDialog.getExistingDirectory(
            self,
            "Select VIVA export root",
            str(start),
        )
        if not directory:
            return None
        return Path(directory).expanduser().resolve()

    def _show_export_success(
        self,
        paths: services.VIVAExportPaths,
        *,
        warnings: Sequence[str] | None = None,
        manifest: Optional[Dict[str, object]] = None,
    ) -> None:
        message = f"VIVA export ready • 1 txt, 1 mdb → {paths.folder}"
        details_lines = [
            f"BOM text: {paths.bom_txt}",
            f"Complex MDB: {paths.mdb_path}",
        ]
        trace_id = manifest.get("trace_id") if isinstance(manifest, dict) else None
        if isinstance(trace_id, str) and trace_id.strip():
            details_lines.append(f"Trace ID: {trace_id.strip()}")
        if warnings:
            warning_text = "\n".join(warnings)
            details_lines.append(f"Warnings:\n{warning_text}")
        details = "\n".join(details_lines)
        box = QMessageBox(self)
        box.setWindowTitle("Export for VIVA")
        box.setIcon(QMessageBox.Icon.Information)
        box.setText(message)
        box.setInformativeText(details)
        open_button = box.addButton("Open Folder", QMessageBox.ButtonRole.ActionRole)
        box.addButton(QMessageBox.StandardButton.Close)
        box.exec()
        if box.clickedButton() == open_button:
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(paths.folder)))

    def _show_ce_export_summary(self, summary: Dict[str, object]) -> None:
        status = str(summary.get("status") or "").strip()
        if not status:
            return
        exported = int(summary.get("exported_count") or 0)
        missing = int(summary.get("missing_count") or 0)
        export_path = str(summary.get("export_path") or "").strip()
        report_path = str(summary.get("report_path") or "").strip()
        detail = str(summary.get("detail") or "").strip()
        trace_id = str(summary.get("trace_id") or "").strip()

        box = QMessageBox(self)
        box.setWindowTitle("Complex Editor export")
        success_states = {"SUCCESS", "PARTIAL_SUCCESS"}
        if status in success_states:
            box.setIcon(QMessageBox.Icon.Information)
        else:
            box.setIcon(QMessageBox.Icon.Warning)

        if status == "SUCCESS":
            message = (
                f"Exported {exported} complexes to {export_path}."
                if export_path
                else f"Exported {exported} complexes."
            )
        elif status == "PARTIAL_SUCCESS":
            message = f"Exported {exported}; {missing} skipped. See report."
        elif status in {"FAILED_INPUT", "FAILED_BACKEND"}:
            message = detail or "Complex Editor export failed."
        elif status in {"RETRY_LATER", "RETRY_WITH_BACKOFF"}:
            message = detail or "Complex Editor export is temporarily unavailable."
        else:
            message = detail or f"Complex Editor export finished with status {status}."

        details_lines: List[str] = []
        if status not in {"SUCCESS", "PARTIAL_SUCCESS"} and detail and message != detail:
            details_lines.append(detail)
        if export_path:
            details_lines.append(f"Export path: {export_path}")
        if report_path:
            details_lines.append(f"Report: {report_path}")
        if trace_id:
            details_lines.append(f"Trace ID: {trace_id}")
        box.setText("Complex Editor export")
        box.setInformativeText("\n".join([message] + details_lines if details_lines else [message]))

        open_folder_button = None
        open_report_button = None
        if export_path:
            open_folder_button = box.addButton(
                "Open CE Folder", QMessageBox.ButtonRole.ActionRole
            )
        if report_path:
            open_report_button = box.addButton(
                "Open Report", QMessageBox.ButtonRole.ActionRole
            )
        box.addButton(QMessageBox.StandardButton.Close)
        box.exec()
        clicked = box.clickedButton()
        if open_folder_button and clicked == open_folder_button:
            try:
                target = Path(export_path)
                folder = target if target.is_dir() else target.parent
                if folder:
                    QDesktopServices.openUrl(QUrl.fromLocalFile(str(folder)))
            except Exception:
                logging.debug("Failed to open CE folder for %s", export_path)
        elif open_report_button and clicked == open_report_button:
            try:
                QDesktopServices.openUrl(QUrl.fromLocalFile(report_path))
            except Exception:
                logging.debug("Failed to open CE report: %s", report_path)

    def _show_export_error(self, message: str) -> None:
        box = QMessageBox(self)
        box.setWindowTitle("Export failed")
        box.setIcon(QMessageBox.Icon.Warning)
        box.setText("Export failed")
        box.setInformativeText(message)
        logs_button = box.addButton("Open Logs", QMessageBox.ButtonRole.ActionRole)
        box.addButton(QMessageBox.StandardButton.Close)
        box.exec()
        if box.clickedButton() == logs_button:
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(LOG_DIR)))

    def _finalize_viva_success(
        self,
        result: services.VIVAExportOutcome,
        base_root: Path,
        *,
        ce_summary: Optional[Dict[str, object]] = None,
    ) -> None:
        self._export_settings.setValue("last_root", str(base_root))
        try:
            save_viva_export_settings(last_export_path=str(base_root))
        except Exception as exc:  # pragma: no cover - best effort
            logging.debug("Failed to persist VIVA export path: %s", exc)
        manifest = result.manifest
        trace_id = manifest.get("trace_id") if isinstance(manifest, dict) else None
        logging.info(
            "VIVA export success: assembly=%s comp_ids=%d folder=%s trace_id=%s",
            self._assembly_id,
            len(result.comp_ids),
            result.paths.folder,
            trace_id,
        )
        self._show_export_success(
            result.paths,
            warnings=result.warnings,
            manifest=manifest,
        )
        if ce_summary:
            logging.info(
                "CE export summary: assembly=%s status=%s exported=%s trace_id=%s path=%s",
                self._assembly_id,
                ce_summary.get("status"),
                ce_summary.get("exported_count"),
                ce_summary.get("trace_id"),
                ce_summary.get("export_path"),
            )
            self._show_ce_export_summary(ce_summary)

    def _prompt_unresolved_pns(self, unresolved: Sequence[str]) -> bool:
        lines = [str(pn) for pn in unresolved]
        if len(lines) > 12:
            display = lines[:12] + ["…"]
        else:
            display = lines
        text = "\n".join(display)
        box = QMessageBox(self)
        box.setWindowTitle("Unresolved complexes")
        box.setIcon(QMessageBox.Icon.Warning)
        box.setText("Complex Editor could not resolve some part numbers.")
        box.setInformativeText(text)
        export_button = box.addButton(
            "Export only resolved",
            QMessageBox.ButtonRole.AcceptRole,
        )
        box.addButton(QMessageBox.StandardButton.Cancel)
        box.setDefaultButton(QMessageBox.StandardButton.Cancel)
        box.exec()
        return box.clickedButton() == export_button

    def _show_unlinked_components(
        self, missing: Sequence[services.VIVAMissingComplex]
    ) -> None:
        if not missing:
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Unlinked components")
        layout = QVBoxLayout(dialog)
        label = QLabel(
            "These BOM lines require a Complex assignment before exporting.", dialog
        )
        label.setWordWrap(True)
        layout.addWidget(label)

        table = QTableWidget(dialog)
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Line", "Part Number", "Description", "Reference"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setRowCount(len(missing))
        for row, item in enumerate(missing):
            table.setItem(row, 0, QTableWidgetItem(str(item.line_number)))
            table.setItem(row, 1, QTableWidgetItem(item.part_number or ""))
            table.setItem(row, 2, QTableWidgetItem(item.description or ""))
            table.setItem(row, 3, QTableWidgetItem(item.reference))
        layout.addWidget(table)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close, dialog)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.resize(600, 320)
        dialog.exec()

    def _show_ce_export_error(self, error: CEExportError) -> None:
        details: List[str] = []
        if error.reason:
            details.append(f"Reason: {error.reason}")
        if error.trace_id:
            details.append(f"Trace ID: {error.trace_id}")
        try:
            base_url = ce_bridge_client.get_active_base_url()
        except Exception:
            base_url = None
        if base_url:
            details.append(f"Bridge URL: {base_url}")
        message = str(error)
        box = QMessageBox(self)
        box.setWindowTitle("Export failed")
        box.setIcon(QMessageBox.Icon.Warning)
        box.setText("Export failed")
        box.setInformativeText("\n".join([message] + details if details else [message]))
        diagnostics_button = None
        diagnostics = getattr(error, "diagnostics", None)
        if isinstance(diagnostics, services.VIVAExportDiagnostics):
            diagnostics_button = box.addButton(
                "Copy diagnostics",
                QMessageBox.ButtonRole.ActionRole,
            )
        logs_button = box.addButton("Open Logs", QMessageBox.ButtonRole.ActionRole)
        box.addButton(QMessageBox.StandardButton.Close)
        box.exec()
        clicked = box.clickedButton()
        if clicked == logs_button:
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(LOG_DIR)))
        elif diagnostics_button and clicked == diagnostics_button:
            self._copy_viva_diagnostics(diagnostics, error)
        

    def _show_strict_fail_dialog(
        self,
        error: CEExportStrictError,
        pn_map: Dict[str, int],
    ) -> None:
        dialog = VivaStrictFailDialog(self, error, pn_map, self._open_complex_for_part)
        dialog.exec()

    def _copy_viva_diagnostics(
        self,
        diagnostics: services.VIVAExportDiagnostics | None,
        error: CEExportError,
    ) -> None:
        if diagnostics is None:
            return
        payload: Dict[str, object] = {}
        manifest_path = diagnostics.manifest_path
        if manifest_path and manifest_path.exists():
            try:
                payload["manifest"] = json.loads(manifest_path.read_text(encoding="utf-8"))
            except Exception as exc:
                payload["manifest_error"] = str(exc)
        ce_path = diagnostics.ce_response_path
        if ce_path and ce_path.exists():
            try:
                payload["ce_response"] = json.loads(ce_path.read_text(encoding="utf-8"))
            except Exception as exc:
                payload["ce_response_error"] = str(exc)
        elif error.payload:
            payload["ce_response"] = error.payload
        text = json.dumps(payload, indent=2, sort_keys=True)
        QApplication.clipboard().setText(text)
        QMessageBox.information(
            self,
            "Diagnostics copied",
            "Diagnostics JSON copied to clipboard.",
        )

    def _open_complex_for_part(self, part_id: int) -> bool:
        link = self._complex_links.get(part_id)
        ce_id = getattr(link, "ce_complex_id", None) if link else None
        if not ce_id:
            QMessageBox.information(
                self,
                "Complex Editor",
                "No linked complex is available for this part.",
            )
            return False
        try:
            ce_bridge_client.open_complex(str(ce_id))
        except CEAuthError as exc:
            QMessageBox.warning(self, "Complex Editor", str(exc))
            return False
        except CENetworkError as exc:
            QMessageBox.warning(self, "Complex Editor", str(exc))
            return False
        return True

    def closeEvent(self, event) -> None:  # pragma: no cover - UI glue
        # Persist column settings per mode
        mode_key = "cols_by_pn" if self._view_mode == "by_pn" else "cols_by_ref"
        for idx, act in enumerate(self._column_actions):
            self._settings.setValue(f"{mode_key}/col{idx}_visible", act.isChecked())
            self._settings.setValue(f"{mode_key}/col{idx}_width", self.table.columnWidth(idx))
        super().closeEvent(event)

    def _copy_selection(self) -> None:
        model = self.table.model()  # proxy
        if model is None:
            return
        idxs = self.table.selectionModel().selectedIndexes() if self.table.selectionModel() else []
        if not idxs:
            cur = self.table.currentIndex()
            if cur.isValid():
                idxs = [cur]
        if not idxs:
            return
        # Build grid by row/col
        rows = sorted({i.row() for i in idxs})
        cols = sorted({i.column() for i in idxs})
        lines = []
        for r in rows:
            vals = []
            for c in cols:
                idx = next((i for i in idxs if i.row() == r and i.column() == c), None)
                vals.append(str(model.data(idx) or "") if idx is not None else "")
            lines.append("\t".join(vals))
        text = "\n".join(lines)
        QGuiApplication.clipboard().setText(text)

    def _paste_selection(self) -> None:
        model = self.table.model()
        if model is None:
            return
        text = QGuiApplication.clipboard().text()
        if text == "":
            return
        idxs = self.table.selectionModel().selectedIndexes() if self.table.selectionModel() else []
        if not idxs:
            cur = self.table.currentIndex()
            if cur.isValid():
                idxs = [cur]
        if not idxs:
            return
        col = idxs[0].column()
        if any(i.column() != col for i in idxs):
            return
        value = text.splitlines()[0].split("\t")[0]
        # Validate against allowed options for combo columns
        method_columns = {
            idx for idx in (
                self._col_indices.get("test_method"),
                self._col_indices.get("test_method_powered"),
            )
            if idx is not None
        }
        detail_columns = {
            idx for idx in (
                self._col_indices.get("test_detail"),
                self._col_indices.get("test_detail_powered"),
            )
            if idx is not None
        }
        if col in method_columns:
            allowed = {"", "Macro", "Complex", "Quick test (QT)", "Python code"}
            if value not in allowed:
                return
        elif col in detail_columns:
            if value and value not in self._function_options:
                return
            # Ensure Test Method is 'Macro' when pasting a Macro kind
            tm_col = (
                self._col_indices.get("test_method")
                if col == self._col_indices.get("test_detail")
                else self._col_indices.get("test_method_powered")
            )
            if tm_col is not None:
                for idx in idxs:
                    part_id = model.data(idx, PartIdRole)
                    if part_id is None:
                        continue
                    cur_method = str(model.data(model.index(idx.row(), tm_col)) or "")
                    if cur_method != "Macro":
                        model.setData(model.index(idx.row(), tm_col), "Macro")
                        self._on_method_changed(part_id, "Macro", tm_col)
        elif col == self._col_indices.get("ap"):
            allowed = {"", "active", "passive"}
            if value not in allowed:
                return
        for idx in idxs:
            part_id = model.data(idx, PartIdRole)
            model.setData(idx, value)
            if col in method_columns and part_id is not None:
                self._on_method_changed(part_id, value, col)
            elif col in detail_columns and part_id is not None:
                self._on_detail_changed(part_id, value or None, col)
            elif col == self._col_indices.get("ap") and part_id is not None:
                self._on_value_changed(part_id, (value or None) if value in ("active", "passive") else None)

    def _on_model_changed(self, index: QModelIndex, old, new) -> None:
        if self._updating:
            return
        # Push an undo command capturing this change. The command itself will
        # reapply the change when pushed, but the ``_updating`` flag prevents
        # recursive signal handling.
        cmd = SetCellCommand(self, index, old, new)
        self.undo_stack.push(cmd)
        # Track edits to the Link column and persist/apply as requested
        try:
            link_col = self._col_indices.get("link")
            if link_col is not None and index.column() == link_col:
                # Resolve part id for this row
                part_id = None
                for c in range(self.model.columnCount()):
                    pid = self.model.data(self.model.index(index.row(), c), PartIdRole)
                    if pid is not None:
                        part_id = pid
                        break
                if part_id is not None:
                    url = (new or "").strip() if isinstance(new, str) else (str(new).strip() if new is not None else "")
                    self._part_product_links[part_id] = url or None
                    if self.apply_act.isChecked():
                        try:
                            with app_state.get_session() as session:
                                services.update_part_product_url(session, part_id, url or None)
                        except Exception:
                            pass
                    else:
                        self._dirty_links[part_id] = url or None
                        self.save_act.setEnabled(True)
                    QTimer.singleShot(0, self._install_datasheet_widgets)
        except Exception:
            pass

    def _open_selected_link(self) -> None:
        sel = self.table.selectionModel()
        if not sel or not sel.selectedIndexes():
            return
        row = sel.selectedIndexes()[0].row()
        # Resolve part id
        part_id = None
        for c in range(self.model.columnCount()):
            pid = self.proxy.data(self.proxy.index(row, c), PartIdRole)
            if pid is not None:
                part_id = pid
                break
        if part_id is None:
            return
        link = self._part_product_links.get(part_id)
        if link:
            QDesktopServices.openUrl(QUrl.fromUserInput(link))

    def _copy_selected_link(self) -> None:
        sel = self.table.selectionModel()
        if not sel or not sel.selectedIndexes():
            return
        row = sel.selectedIndexes()[0].row()
        part_id = None
        for c in range(self.model.columnCount()):
            pid = self.proxy.data(self.proxy.index(row, c), PartIdRole)
            if pid is not None:
                part_id = pid
                break
        if part_id is None:
            return
        link = self._part_product_links.get(part_id)
        if link:
            from PyQt6.QtGui import QGuiApplication
            cb = QGuiApplication.clipboard()
            try:
                cb.setText(link)
            except Exception:
                pass

    def _delete_selected_rows(self) -> None:  # pragma: no cover - UI glue
        proxy = self.table.model()
        sel = self.table.selectionModel()
        if proxy is None or sel is None or not sel.selectedIndexes():
            return
        rows = sorted({i.row() for i in sel.selectedIndexes()})
        bom_ids: set[int] = set()
        part_ids: set[int] = set()
        for r in rows:
            # Try to find a BOM item id on any column in this row
            bid = None
            pid = None
            for c in range(proxy.columnCount()):
                idx = proxy.index(r, c)
                if bid is None:
                    b = proxy.data(idx, BOMItemIdRole)
                    if isinstance(b, int):
                        bid = b
                if pid is None:
                    p = proxy.data(idx, PartIdRole)
                    if isinstance(p, int):
                        pid = p
                if bid is not None and pid is not None:
                    break
            if bid is not None:
                bom_ids.add(int(bid))
            elif pid is not None:
                part_ids.add(int(pid))
        if not bom_ids and not part_ids:
            return
        # Confirm
        total = len(bom_ids) + len(part_ids)
        msg = (
            f"Delete {total} row(s)?\n\n"
            + ("Grouped rows (by PN) will remove all references for that part in this assembly." if part_ids else "")
        )
        res = QMessageBox.question(self, "Delete", msg)
        if res != QMessageBox.StandardButton.Yes:
            return
        # Perform deletion
        try:
            with app_state.get_session() as session:
                if bom_ids:
                    services.delete_bom_items(session, list(bom_ids))
                for pid in part_ids:
                    services.delete_bom_items_for_part(session, self._assembly_id, pid)
        except Exception as exc:
            QMessageBox.warning(self, "Delete failed", str(exc))
            return
        # Reload data and model
        self._load_data()
        self._rebuild_model()

    def _visible_columns(self) -> list[int]:
        model = self.table.model()
        if model is None:
            return []
        cols = []
        for c in range(model.columnCount()):
            try:
                if not self.table.isColumnHidden(c):
                    cols.append(c)
            except Exception:
                cols.append(c)
        return cols

    def _is_full_row_selection(self) -> bool:
        proxy = self.table.model()
        sel = self.table.selectionModel()
        if proxy is None or sel is None:
            return False
        idxs = sel.selectedIndexes()
        if not idxs:
            return False
        vis_cols = set(self._visible_columns())
        # Map row -> set of selected visible columns
        rows = sorted({i.row() for i in idxs})
        for r in rows:
            cols_sel = {i.column() for i in idxs if i.row() == r and i.column() in vis_cols}
            if cols_sel != vis_cols:
                return False
        return True

    def _clear_selected_cells(self) -> None:
        proxy = self.table.model()
        sel = self.table.selectionModel()
        if proxy is None or sel is None:
            return
        idxs = sel.selectedIndexes()
        if not idxs:
            cur = self.table.currentIndex()
            if cur.isValid():
                idxs = [cur]
        if not idxs:
            return
        method_cols = {
            idx for idx in (
                self._col_indices.get("test_method"),
                self._col_indices.get("test_method_powered"),
            )
            if idx is not None
        }
        detail_cols = {
            idx for idx in (
                self._col_indices.get("test_detail"),
                self._col_indices.get("test_detail_powered"),
            )
            if idx is not None
        }
        ap_col = self._col_indices.get("ap")
        for idx in idxs:
            try:
                # Only clear editable cells
                if not (proxy.flags(idx) & Qt.ItemFlag.ItemIsEditable):
                    continue
            except Exception:
                pass
            col = idx.column()
            part_id = proxy.data(idx, PartIdRole)
            # Apply column-specific hooks for consistency with paste handler
            proxy.setData(idx, "")
            if part_id is None:
                continue
            if col in method_cols:
                self._on_method_changed(part_id, "", col)
            elif col in detail_cols:
                self._on_detail_changed(part_id, None, col)
            elif ap_col is not None and col == ap_col:
                self._on_value_changed(part_id, None)

    def _delete_or_clear_selection(self) -> None:
        # If the user has selected entire rows (all visible columns), confirm row delete;
        # otherwise treat Delete as clearing the selected cell values only.
        if self._is_full_row_selection():
            self._delete_selected_rows()
        else:
            self._clear_selected_cells()

    # ------------------------------------------------------------------
    def _load_function_options(self) -> list[str]:
        """Load function options from data/function_list.txt.

        Returns an empty list if not found. Lines starting with '#' are ignored.
        """
        candidates = []
        try:
            here = Path(__file__).resolve()
            candidates.append(here.parents[2] / "data" / "function_list.txt")
            candidates.append(here.parents[1] / "data" / "function_list.txt")
            candidates.append(Path.cwd() / "data" / "function_list.txt")
        except Exception:
            candidates.append(Path.cwd() / "data" / "function_list.txt")
        candidates.append(DATA_ROOT / "function_list.txt")
        for p in candidates:
            try:
                if p.exists():
                    lines = p.read_text(encoding="utf-8").splitlines()
                    items = [ln.strip() for ln in lines if ln.strip() and not ln.strip().startswith("#")]
                    if items:
                        return items
            except Exception:
                continue
        # Fallback: derive options from prefix_macros (unique macro names)
        try:
            from ..logic.prefix_macros import load_prefix_macros as _lpm
            macros = sorted({macro for _pref, macro in (_lpm() or [])})
            return macros
        except Exception:
            return []

    # ------------------------------------------------------------------
    @staticmethod
    def _configure_icon_button(button: QToolButton, color: str, enabled: bool, disabled_color: str | None = None) -> None:
        """Apply consistent styling to inline icon buttons."""
        button.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        button.setEnabled(enabled)
        disabled_color = disabled_color or color
        button.setStyleSheet(
            "QToolButton { color: %s; padding: 0px; border: none; } "
            "QToolButton:disabled { color: %s; opacity: 0.6; }" % (color, disabled_color)
        )

    @staticmethod
    def _configure_icon_button(button: QToolButton, color: str, enabled: bool) -> None:
        button.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        button.setCursor(Qt.CursorShape.PointingHandCursor if enabled else Qt.CursorShape.ArrowCursor)
        button.setEnabled(enabled)
        button.setStyleSheet(
            "QToolButton { color: %s; padding: 0px; border: none; }"
            " QToolButton:disabled { color: %s; opacity: 0.5; }" % (color, color)
        )

    def _load_icon(self, name: str, fallback: QStyle.StandardPixmap) -> QIcon:
        try:
            p = Path(__file__).resolve().parent / "icons" / name
            if p.exists():
                icon = QIcon(str(p))
                if not icon.isNull():
                    return icon
        except Exception:
            pass
        return self.style().standardIcon(fallback)

    def _icon_pdf_open(self) -> QIcon:
        return self._load_icon("pdf_green.svg", QStyle.StandardPixmap.SP_FileIcon)

    def _icon_pdf_add(self) -> QIcon:
        return self._load_icon("pdf_plus.svg", QStyle.StandardPixmap.SP_DialogOpenButton)

    def _icon_pdf_error(self) -> QIcon:
        return self._load_icon("pdf_x.svg", QStyle.StandardPixmap.SP_MessageBoxCritical)

    def _icon_for_loading(self) -> QIcon:
        return self._load_icon("pdf_loading.svg", QStyle.StandardPixmap.SP_BrowserReload)

    def _icon_link_open(self) -> QIcon:
        return self._load_icon("link_green.svg", QStyle.StandardPixmap.SP_DialogOpenButton)

    def _icon_link_add(self) -> QIcon:
        return self._load_icon("link_plus.svg", QStyle.StandardPixmap.SP_DialogOpenButton)

    def _install_datasheet_widgets(self) -> None:
        # Rebuild buttons for Datasheet and Link columns without showing raw text
        ds_col = self._col_indices.get("ds")
        link_col = self._col_indices.get("link")
        if ds_col is None and link_col is None:
            return
        self._updating = True
        try:
            for r in range(self.model.rowCount()):
                ds_idx = self.model.index(r, ds_col) if ds_col is not None else None
                link_idx = self.model.index(r, link_col) if link_col is not None else None
                if ds_idx is not None:
                    self.model.setData(ds_idx, "", Qt.ItemDataRole.DisplayRole)
                if link_idx is not None:
                    self.model.setData(link_idx, "", Qt.ItemDataRole.DisplayRole)

                part_id = None
                for c in range(self.model.columnCount()):
                    pid = self.model.data(self.model.index(r, c), PartIdRole)
                    if pid is not None:
                        part_id = pid
                        break
                if part_id is None:
                    continue

                path = self._part_datasheets.get(part_id)
                path_value = path or ""
                path_exists = bool(path_value) and Path(path_value).exists()
                btn = QToolButton(self.table)
                try:
                    from PyQt6.QtCore import QSize as _QSize
                    btn.setAutoRaise(True)
                    btn.setIconSize(_QSize(24, 24))
                    btn.setFixedSize(30, 30)
                except Exception:
                    pass

                staged_flag = part_id in self._dirty_datasheets
                staged_value = self._dirty_datasheets.get(part_id)
                if part_id in self._datasheet_loading:
                    btn.setIcon(self._icon_for_loading())
                    self._configure_icon_button(btn, "#1F6FEB", False)
                    btn.setToolTip("Searching datasheet...")
                elif staged_flag and staged_value is None:
                    btn.setIcon(self._icon_pdf_add())
                    self._configure_icon_button(btn, "#1F6FEB", True)
                    btn.setToolTip("Datasheet will be removed on Save. Click to attach a replacement.")
                    btn.clicked.connect(lambda _=False, pid=part_id: self._open_attach_dialog(pid))
                else:
                    effective_path = path if path_exists else (staged_value if staged_flag else path)
                    if effective_path and Path(effective_path).exists():
                        btn.setIcon(self._icon_pdf_open())
                        self._configure_icon_button(btn, "#28A745", True)
                        tip = "Open datasheet"
                        if staged_flag and staged_value:
                            tip += " (pending save)"
                        btn.setToolTip(tip)
                        btn.clicked.connect(lambda _=False, p=effective_path: self._open_pdf_path(p))
                    else:
                        if path_value.strip() and not path_exists:
                            btn.setIcon(self._icon_pdf_error())
                            self._configure_icon_button(btn, "#DC2626", True)
                            btn.setToolTip("Stored path not found. Click to attach a new datasheet.")
                        elif part_id in self._datasheet_failed:
                            btn.setIcon(self._icon_pdf_error())
                            self._configure_icon_button(btn, "#DC2626", True)
                            btn.setToolTip("Search attempted; no datasheet found. Click to attach manually.")
                        else:
                            btn.setIcon(self._icon_pdf_add())
                            self._configure_icon_button(btn, "#1F6FEB", True)
                            btn.setToolTip("Attach datasheet")
                        btn.clicked.connect(lambda _=False, pid=part_id: self._open_attach_dialog(pid))

                if ds_idx is not None:
                    self.model.setData(ds_idx, self._part_datasheets.get(part_id), DatasheetRole)
                    proxy_idx = self.proxy.mapFromSource(ds_idx)
                    self.table.setIndexWidget(proxy_idx, btn)

                if link_idx is not None:
                    link_btn = QToolButton(self.table)
                    try:
                        from PyQt6.QtCore import QSize as _QSize
                        link_btn.setAutoRaise(True)
                        link_btn.setIconSize(_QSize(24, 24))
                        link_btn.setFixedSize(30, 30)
                    except Exception:
                        pass
                    link = self._dirty_links.get(part_id, self._part_product_links.get(part_id))
                    link_effective = link or ""
                    self.model.setData(link_idx, link_effective, LinkUrlRole)
                    if link:
                        link_btn.setIcon(self._icon_link_open())
                        self._configure_icon_button(link_btn, "#28A745", True)
                        link_btn.setToolTip("Open product link")
                        link_btn.clicked.connect(lambda _=False, u=link_effective: QDesktopServices.openUrl(QUrl.fromUserInput(u)))
                    else:
                        link_btn.setIcon(self._icon_link_add())
                        self._configure_icon_button(link_btn, "#1F6FEB", True)
                        link_btn.setToolTip("Set product link")
                        link_btn.clicked.connect(lambda _=False, pid=part_id: self._prompt_link_entry(pid))
                    proxy_link_idx = self.proxy.mapFromSource(link_idx)
                    self.table.setIndexWidget(proxy_link_idx, link_btn)
        finally:
            self._updating = False

    def _open_attach_dialog(self, part_id: int) -> None:
        from .datasheet_attach_dialog import DatasheetAttachDialog
        dlg = DatasheetAttachDialog(part_id, self)
        dlg.attached.connect(lambda canonical, pid=part_id: self._on_datasheet_attached(pid, canonical))
        dlg.exec()

    def _on_datasheet_attached(self, part_id: int, canonical: str) -> None:
        # Respect Apply toggle: stage or persist
        self._datasheet_failed.discard(part_id)
        if self.apply_act.isChecked():
            try:
                with app_state.get_session() as session:
                    services.update_part_datasheet_url(session, part_id, canonical)
            except Exception as exc:
                QMessageBox.warning(self, "Attach failed", str(exc))
                return
            self._part_datasheets[part_id] = canonical
            self._dirty_datasheets.pop(part_id, None)
        else:
            self._part_datasheets[part_id] = canonical
            self._dirty_datasheets[part_id] = canonical
            self.save_act.setEnabled(True)
        # Update UI labels and link column from DB snapshot (best-effort)
        try:
            from ..models import Part
            with app_state.get_session() as session:
                p = session.get(Part, part_id)
                if p:
                    self._fanout_part_field(part_id, "desc", getattr(p, "description", None))
                    self._part_product_links[part_id] = getattr(p, "product_url", None)
        except Exception:
            pass
        self._install_datasheet_widgets()

    def _on_datasheet_failed(self, part_id: int) -> None:
        self._datasheet_failed.add(part_id)
        self._install_datasheet_widgets()

    def _on_manual_link(self, part_id: int, url: str) -> None:
        self._part_manual_links[part_id] = url
        # Persist as product link on the Part for reuse across projects
        try:
            with app_state.get_session() as session:
                services.update_part_product_url(session, part_id, url)
        except Exception:
            pass
        self._part_product_links[part_id] = url
        self._datasheet_failed.discard(part_id)
        # Update visible cell
        link_col = self._col_indices.get("link")
        if link_col is not None:
            for r in range(self.model.rowCount()):
                for c in range(self.model.columnCount()):
                    if self.model.data(self.model.index(r, c), PartIdRole) == part_id:
                        self.model.setData(self.model.index(r, link_col), url)
                        break
        self._install_datasheet_widgets()

    def _open_pdf_path(self, path: str) -> None:
        """Open a datasheet path using the configured viewer."""
        t0 = time.perf_counter()
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        try:
            size = Path(path).stat().st_size if Path(path).exists() else -1
        except Exception:
            size = -1
        if PDF_OPEN_DEBUG:
            logging.info(
                "Open-PDF[%s]: start part_path=%s size=%s bytes viewer=%s",
                ts,
                path,
                size,
                PDF_VIEWER,
            )
        t1 = time.perf_counter()
        try:
            local_path = str(get_local_open_path(Path(path)))
        except Exception:
            local_path = path
        t2 = time.perf_counter()
        if PDF_OPEN_DEBUG:
            logging.info(
                "Open-PDF: resolved local path in %.1f ms -> %s",
                (t2 - t1) * 1000.0,
                local_path,
            )

        opened = False
        open_err: str | None = None
        t3 = time.perf_counter()
        try:
            if PDF_VIEWER == "chrome":
                exe = (
                    PDF_VIEWER_PATH
                    or shutil.which("chrome")
                    or shutil.which("chrome.exe")
                    or r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
                )
                if exe and os.path.exists(exe):
                    url = QUrl.fromLocalFile(local_path).toString()
                    subprocess.Popen([exe, url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    opened = True
                else:
                    open_err = f"Chrome not found (PDF_VIEWER_PATH={PDF_VIEWER_PATH})"
            elif PDF_VIEWER == "edge":
                exe = (
                    PDF_VIEWER_PATH
                    or shutil.which("msedge")
                    or shutil.which("msedge.exe")
                    or r"C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe"
                )
                if exe and os.path.exists(exe):
                    url = QUrl.fromLocalFile(local_path).toString()
                    subprocess.Popen([exe, url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    opened = True
                else:
                    open_err = f"Edge not found (PDF_VIEWER_PATH={PDF_VIEWER_PATH})"
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(local_path))
                opened = True
        except Exception as e:  # pragma: no cover - environment dependent
            open_err = str(e)
        t4 = time.perf_counter()
        if PDF_OPEN_DEBUG:
            logging.info(
                "Open-PDF: open call returned in %.1f ms (ok=%s)%s",
                (t4 - t3) * 1000.0,
                opened,
                f" err={open_err}" if open_err else "",
            )
            logging.info("Open-PDF: total time %.1f ms", (t4 - t0) * 1000.0)

    def _prompt_link_entry(self, part_id: int) -> None:
        current = self._part_product_links.get(part_id) or ""
        text, ok = QInputDialog.getText(self, "Set Product Link", "Enter product URL:", text=current)
        if not ok:
            return
        text = (text or "").strip()
        link_col = self._col_indices.get("link")
        if link_col is None:
            return
        for row in range(self.model.rowCount()):
            idx = self.model.index(row, link_col)
            if self.model.data(idx, PartIdRole) == part_id:
                self.model.setData(idx, text)
                break
        QTimer.singleShot(0, self._install_datasheet_widgets)

    def _remove_selected_datasheets(self) -> None:
        proxy = self.table.model()
        sel = self.table.selectionModel()
        if proxy is None or sel is None or not sel.selectedIndexes():
            return
        rows = sorted({i.row() for i in sel.selectedIndexes()})
        # Collect part IDs that currently have a datasheet path
        target_pids: list[int] = []
        ds_col = self._col_indices.get("ds")
        for r in rows:
            # find part_id on this row
            part_id = None
            for c in range(proxy.columnCount()):
                pid = proxy.data(proxy.index(r, c), PartIdRole)
                if pid is not None:
                    part_id = pid
                    break
            if part_id is None:
                continue
            path = self._part_datasheets.get(part_id)
            if path:
                target_pids.append(part_id)
        if not target_pids:
            return
        # Confirm
        msg = (
            "Remove datasheet for the selected part?" if len(target_pids) == 1
            else f"Remove datasheets for {len(target_pids)} parts?"
        )
        res = QMessageBox.question(self, "Remove Datasheet", msg)
        if res != QMessageBox.StandardButton.Yes:
            return
        if self.apply_act.isChecked():
            removed_any = False
            with app_state.get_session() as session:
                for pid in target_pids:
                    try:
                        services.remove_part_datasheet(session, pid, delete_file=True)
                        self._part_datasheets[pid] = None
                        self._dirty_datasheets.pop(pid, None)
                        removed_any = True
                    except Exception as exc:
                        QMessageBox.warning(self, "Remove failed", str(exc))
            if removed_any:
                self._install_datasheet_widgets()
        else:
            # Stage removal
            for pid in target_pids:
                self._dirty_datasheets[pid] = None
                self._part_datasheets[pid] = None
            self.save_act.setEnabled(True)
            self._install_datasheet_widgets()


class WrapTextDelegate(QStyledItemDelegate):
    """Delegate that wraps long text (e.g., References) within column width.

    The visible number of lines can be adjusted to change row height.
    """

    def __init__(self, parent=None, lines: int = 1) -> None:
        super().__init__(parent)
        self._lines = lines

    def set_line_count(self, lines: int) -> None:
        self._lines = max(1, lines)

    def paint(self, painter: QPainter, option: 'QStyleOptionViewItem', index):  # pragma: no cover - UI glue
        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        # Draw base item without text
        text = opt.text
        opt.text = ""
        style = opt.widget.style() if opt.widget else QApplication.style()
        style.drawControl(QStyle.ControlElement.CE_ItemViewItem, opt, painter, opt.widget)

        # Draw wrapped text
        doc = QTextDocument()
        doc.setDefaultFont(opt.font)
        topt = QTextOption()
        topt.setWrapMode(QTextOption.WrapMode.WrapAtWordBoundaryOrAnywhere)
        doc.setDefaultTextOption(topt)
        doc.setTextWidth(opt.rect.width())
        doc.setPlainText(text)
        painter.save()
        painter.translate(opt.rect.topLeft())
        doc.drawContents(painter, QRectF(0, 0, opt.rect.width(), opt.rect.height()))
        painter.restore()

    def sizeHint(self, option: 'QStyleOptionViewItem', index):  # pragma: no cover - UI glue
        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        # Provide a sensible width (fallback if opt.rect is empty)
        width = max(100, opt.rect.width())
        line_height = opt.fontMetrics.lineSpacing()
        # Add padding
        return QSize(int(width), int(line_height * self._lines) + 6)




