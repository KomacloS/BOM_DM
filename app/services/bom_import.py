"""BOM import service supporting CSV and XLSX sources."""

from __future__ import annotations

import csv
import io
import re
import shutil
import logging
from decimal import Decimal
from pathlib import Path
from typing import List

from pydantic import BaseModel, Field, validator
from sqlmodel import Session, select
from sqlalchemy.exc import IntegrityError

from ..models import Assembly, BOMItem, Part, PartType
from ..config import get_complex_editor_settings
from ..domain.complex_linker import auto_link_by_pn
from ..integration.ce_bridge_client import CENetworkError
from ..integration.ce_bridge_manager import CEBridgeError, ensure_ce_bridge_ready


logger = logging.getLogger(__name__)


class ImportReport(BaseModel):
    total: int
    matched: int
    unmatched: int
    created_task_ids: List[int] = Field(default_factory=list)
    errors: List[str] = Field(default_factory=list)


# ---------------------------------------------------------------------------
# Header handling


def _norm(s: str) -> str:
    """Normalize a header name by stripping, lowering and removing symbols."""
    return re.sub(r"[^a-z0-9]+", "", s.strip().lower())


HEADER_MAP = {
    "part_number": {"pn", "mpn", "part", "partnumber", "manufacturerpartnumber"},
    "reference": {"ref", "reference", "designator", "refdes", "refdesg", "designators"},
    "qty": {"qty", "quantity", "q'ty", "qnty"},
    "manufacturer": {"mfr", "manufacturer", "vendor", "maker"},
    "active_passive": {"active/passive", "activepassive", "a/p", "ap"},
    "function": {"function", "func"},
    "tol_p": {"tolerancep", "tolerance+", "tol+", "tolp", "tolerancepos"},
    "tol_n": {"tolerancen", "tolerance-", "tol-", "toln", "toleranceneg"},
    "unit_cost": {"price", "unitprice", "cost", "unit_cost"},
    "currency": {"currency", "curr"},
    "datasheet_url": {"datasheet", "datasheeturl", "datasheet_link", "datasheetlink"},
    "notes": {"notes", "comment", "comments"},
}


def validate_headers(headers: List[str]) -> dict[str, int]:
    """Validate headers and return a mapping of canonical name to index."""

    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        hn = _norm(h)
        for canon, variants in HEADER_MAP.items():
            norm_set = {_norm(canon)} | {_norm(v) for v in variants}
            if hn in norm_set and canon not in col_map:
                col_map[canon] = idx
                break
    missing = [c for c in ("part_number", "reference") if c not in col_map]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")
    return col_map


# ---------------------------------------------------------------------------
# Row schema


class BOMRow(BaseModel):
    part_number: str
    reference: str
    qty: int = 1
    manufacturer: str | None = None
    active_passive: str | None = None
    function: str | None = None
    tol_p: str | None = None
    tol_n: str | None = None
    unit_cost: Decimal | None = None
    currency: str | None = None
    datasheet_url: str | None = None
    notes: str | None = None

    @validator("part_number", "reference")
    def _req_trim(cls, v):
        v = (v or "").strip()
        if not v:
            raise ValueError("required")
        return v

    @validator("qty", pre=True)
    def _qty_int(cls, v):
        if v in (None, "", " "):
            return 1
        try:
            return int(str(v).strip())
        except Exception:  # pragma: no cover - defensive
            raise ValueError("qty must be integer")

    @validator("currency")
    def _cur_up(cls, v):
        return (v or "").upper() or None

    @validator("unit_cost", pre=True)
    def _unit_cost_optional(cls, v):
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        return Decimal(s)


_RANGE = re.compile(r"^([A-Za-z]+)(\d+)\s*-\s*([A-Za-z]+)?(\d+)$")


def _expand_references(ref_str: str) -> list[str]:
    # Split on commas first
    tokens = [t.strip() for t in ref_str.split(",") if t.strip()]
    out: list[str] = []
    for tok in tokens:
        m = _RANGE.match(tok)
        if m:
            p1, n1, p2, n2 = m.group(1), m.group(2), m.group(3), m.group(4)
            if p2 and p2 != p1:
                out.append(tok)
                continue
            start, end = int(n1), int(n2)
            if start > end:
                start, end = end, start
            out.extend([f"{p1}{i}" for i in range(start, end + 1)])
        else:
            out.append(tok)
    return out


# ---------------------------------------------------------------------------
# File parsing helpers


def _is_xlsx(data: bytes) -> bool:
    return data[:2] == b"PK"


def _iter_rows(data: bytes) -> tuple[list[str], list[list[str]]]:
    if _is_xlsx(data):
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = [[(c if c is not None else "") for c in r] for r in ws.iter_rows(values_only=True)]
        headers = [str(x or "") for x in rows[0]] if rows else []
        return headers, [list(map(lambda x: "" if x is None else str(x), r)) for r in rows[1:]]
    else:
        text = data.decode("utf-8-sig", errors="ignore")
        rdr = csv.reader(io.StringIO(text))
        rows = list(rdr)
        headers = rows[0] if rows else []
        return headers, rows[1:]


# ---------------------------------------------------------------------------
# Datasheet caching


DATASHEETS_DIR = Path("datasheets")
DATASHEETS_DIR.mkdir(exist_ok=True)


def _safe_pn(pn: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", pn.strip())[:80]


def _cache_datasheet_for_pn(pn: str, url_or_path: str) -> str | None:
    if not url_or_path:
        return None
    try:
        p = Path(url_or_path)
        if p.exists() and p.is_file():
            ext = p.suffix or ".pdf"
            dst = DATASHEETS_DIR / f"{_safe_pn(pn)}{ext}"
            shutil.copyfile(p, dst)
            return str(dst)
        if url_or_path.lower().startswith(("http://", "https://")):
            import requests

            r = requests.get(url_or_path, timeout=15)
            r.raise_for_status()
            ext = ".pdf"
            dst = DATASHEETS_DIR / f"{_safe_pn(pn)}{ext}"
            with open(dst, "wb") as f:
                f.write(r.content)
            return str(dst)
    except Exception:  # pragma: no cover - cache failures are non-fatal
        return url_or_path
    return None


# ---------------------------------------------------------------------------
# Importer


def import_bom(assembly_id: int, data: bytes, session: Session) -> ImportReport:
    errors: List[str] = []
    ce_settings = get_complex_editor_settings()
    bridge_cfg = ce_settings.get('bridge', {}) if isinstance(ce_settings, dict) else {}
    auto_link_enabled = bool(bridge_cfg.get('enabled')) if isinstance(bridge_cfg, dict) else False
    if auto_link_enabled:
        try:
            ensure_ce_bridge_ready()
        except CEBridgeError as exc:
            logger.debug('Complex Editor bridge unavailable before import auto-link: %s', exc)
            auto_link_enabled = False
    attempted_links: set[int] = set()
    try:
        headers, raw_rows = _iter_rows(data)
        col_map = validate_headers(headers)
    except Exception as exc:
        errors.append(str(exc))
        return ImportReport(total=0, matched=0, unmatched=0, errors=errors)

    assembly = session.get(Assembly, assembly_id)
    if not assembly:
        errors.append("assembly not found")
        return ImportReport(total=0, matched=0, unmatched=0, errors=errors)

    total = matched = unmatched = 0

    for i, row in enumerate(raw_rows, start=2):
        data_map = {key: row[idx] if idx < len(row) else "" for key, idx in col_map.items()}
        raw_qty = data_map.get("qty")
        if not data_map.get("part_number") or not data_map.get("reference"):
            errors.append(f"Row {i}: missing part_number or reference")
            continue
        try:
            bom_row = BOMRow(**data_map)
        except Exception as e:
            errors.append(f"Row {i}: {e}")
            continue

        total += 1
        pn = bom_row.part_number
        part = session.exec(select(Part).where(Part.part_number == pn)).first()
        is_new = part is None
        if part:
            matched += 1
        else:
            unmatched += 1
            part = Part(part_number=pn)

        if bom_row.active_passive:
            try:
                ap = PartType(bom_row.active_passive.lower())
            except Exception:
                ap = PartType.passive
            if is_new or not getattr(part, "active_passive", None):
                part.active_passive = ap
        if bom_row.function and (is_new or not part.function):
            part.function = bom_row.function
        if bom_row.tol_p and (is_new or not part.tol_p):
            part.tol_p = bom_row.tol_p
        if bom_row.tol_n and (is_new or not part.tol_n):
            part.tol_n = bom_row.tol_n
        if bom_row.datasheet_url and (is_new or not part.datasheet_url):
            cached = _cache_datasheet_for_pn(pn, bom_row.datasheet_url)
            if cached:
                part.datasheet_url = cached

        session.add(part)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            part = session.exec(select(Part).where(Part.part_number == pn)).first()
            if not part:
                errors.append(f"Row {i}: duplicate part_number {pn}")
                continue
        session.refresh(part)

        if auto_link_enabled and part.id is not None and part.id not in attempted_links:
            try:
                auto_link_by_pn(part.id, part.part_number)
            except CENetworkError:
                logger.debug("Complex Editor bridge unavailable during import auto-link for %s", part.part_number)
            finally:
                attempted_links.add(part.id)

        refs = _expand_references(bom_row.reference)
        if len(refs) > 1:
            for ref in refs:
                item = session.exec(
                    select(BOMItem).where(
                        BOMItem.assembly_id == assembly_id, BOMItem.reference == ref
                    )
                ).first()
                if not item:
                    item = BOMItem(assembly_id=assembly_id, reference=ref)
                item.part_id = part.id
                item.qty = 1
                if bom_row.manufacturer:
                    item.manufacturer = bom_row.manufacturer
                if bom_row.unit_cost is not None:
                    item.unit_cost = bom_row.unit_cost
                if bom_row.currency:
                    item.currency = bom_row.currency
                if bom_row.datasheet_url:
                    item.datasheet_url = bom_row.datasheet_url
                if bom_row.notes:
                    item.notes = bom_row.notes
                session.add(item)
            session.commit()
            if raw_qty not in (None, "", " ") and bom_row.qty != len(refs):
                errors.append(
                    f"Row {i}: qty={bom_row.qty} but {len(refs)} references expanded"
                )
        else:
            ref = refs[0] if refs else bom_row.reference
            item = session.exec(
                select(BOMItem).where(
                    BOMItem.assembly_id == assembly_id, BOMItem.reference == ref
                )
            ).first()
            if not item:
                item = BOMItem(assembly_id=assembly_id, reference=ref)

            item.part_id = part.id
            item.qty = bom_row.qty
            if bom_row.manufacturer:
                item.manufacturer = bom_row.manufacturer
            if bom_row.unit_cost is not None:
                item.unit_cost = bom_row.unit_cost
            if bom_row.currency:
                item.currency = bom_row.currency
            if bom_row.datasheet_url:
                item.datasheet_url = bom_row.datasheet_url
            if bom_row.notes:
                item.notes = bom_row.notes

            session.add(item)
            session.commit()

    return ImportReport(total=total, matched=matched, unmatched=unmatched, errors=errors)


__all__ = ["ImportReport", "validate_headers", "import_bom"]

