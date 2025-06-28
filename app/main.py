# root: app/main.py
from __future__ import annotations
from fastapi import (
    FastAPI,
    status,
    UploadFile,
    File,
    HTTPException,
    Depends,
    APIRouter,
    Request,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse, RedirectResponse, HTMLResponse, Response
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from typing import Optional
import jwt
from sqlmodel import SQLModel, Field, Session, select
from sqlalchemy import (
    text,
    UniqueConstraint,
    Column,
    ForeignKey,
    Integer,
)
import sqlalchemy
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timedelta
from apscheduler.schedulers.background import BackgroundScheduler
import csv
import io
import os
import re
import hashlib
from pathlib import Path
from openpyxl import Workbook, load_workbook

from .security import verify_password, get_password_hash, create_access_token
from .config import (
    SECRET_KEY,
    ALGORITHM,
    ACCESS_TOKEN_EXPIRE_MINUTES,
    MAX_DATASHEET_MB,
    MAX_GLB_MB,
    MAX_EDA_MB,
    MAX_PY_MB,
    BOM_HOURLY_USD,
    BOM_DEFAULT_CURRENCY,
    DATABASE_URL,
    load_settings,
    save_database_url,
    get_engine,
    reload_settings,
)

from .pdf_utils import extract_bom_text, parse_bom_lines
from .quote_utils import calculate_quote
from .trace_utils import component_trace, board_trace

if SQLModel.metadata.tables:
    SQLModel.metadata.clear()

from .vendor import octopart, fixer
from . import fx
from .fx import FXRate

engine = get_engine()
if engine.dialect.name == "sqlite":
    @sqlalchemy.event.listens_for(engine, "connect")
    def _fk_pragma(dbapi_con, rec):
        dbapi_con.execute("PRAGMA foreign_keys=ON")
scheduler = BackgroundScheduler()
templates = Jinja2Templates(directory="app/frontend/templates")


def reload_db(url: str) -> None:
    """Rebuild SQL engine and re-init DB."""
    global engine
    save_database_url(url)
    reload_settings()
    engine = get_engine()
    init_db()


class StatusCheck(SQLModel, table=True):
    """Trivial table used to exercise the database connection."""

    id: int | None = Field(default=None, primary_key=True)
    msg: str


class BOMItemBase(SQLModel):
    """Shared properties for BOM items."""

    part_number: str = Field(min_length=1)
    description: str = Field(min_length=1)
    quantity: int = Field(default=1, ge=1)
    reference: str | None = None
    datasheet_url: str | None = None
    manufacturer: str | None = None
    mpn: str | None = None
    footprint: str | None = None
    unit_cost: float | None = Field(default=None, sa_column=Column(sqlalchemy.Numeric(10,4)))
    dnp: bool | None = Field(default=False)
    currency: str = Field(default=BOM_DEFAULT_CURRENCY, max_length=3)

    assembly_id: int | None = Field(
        default=None,
        sa_column=Column(Integer, ForeignKey("assembly.id", ondelete="CASCADE")),
    )
    part_id: int | None = Field(
        default=None,
        sa_column=Column(Integer, ForeignKey("part.id", ondelete="SET NULL")),
    )


class BOMItem(BOMItemBase, table=True):
    """Database model for a BOM item."""

    id: int | None = Field(default=None, primary_key=True)



    __table_args__ = (UniqueConstraint("part_number", "reference", name="uix_part_ref"),)


class BOMItemCreate(BOMItemBase):
    """Schema for creating items via the API."""

    pass


class BOMItemRead(BOMItemBase):
    """Schema returned from the API."""

    id: int


class BOMItemUpdate(SQLModel):
    """Schema for updating items (all fields optional)."""

    part_number: str | None = Field(default=None, min_length=1)
    description: str | None = Field(default=None, min_length=1)
    quantity: int | None = Field(default=None, ge=1)
    reference: str | None = None
    datasheet_url: str | None = None
    assembly_id: int | None = None
    part_id: int | None = None
    manufacturer: str | None = None
    mpn: str | None = None
    footprint: str | None = None
    unit_cost: float | None = None
    dnp: bool | None = None
    currency: str | None = None


class Customer(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    name: str = Field(sa_column_kwargs={"unique": True})
    contact: str | None = None
    notes: str | None = None
    active: bool = True



class CustomerCreate(SQLModel):
    name: str
    contact: str | None = None
    notes: str | None = None
    active: bool = True


class CustomerRead(CustomerCreate):
    id: int


class CustomerUpdate(SQLModel):
    name: str | None = None
    contact: str | None = None
    notes: str | None = None
    active: bool | None = None


class Project(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    customer_id: int = Field(
        sa_column=Column(Integer, ForeignKey("customer.id", ondelete="CASCADE")),
    )
    name: str
    code: str | None = None
    notes: str | None = None
    description: str | None = None



    __table_args__ = (UniqueConstraint("customer_id", "name", name="uix_customer_name"),)


class ProjectCreate(SQLModel):
    customer_id: int
    name: str
    code: str | None = None
    notes: str | None = None
    description: str | None = None


class ProjectRead(ProjectCreate):
    id: int


class ProjectUpdate(SQLModel):
    customer_id: int | None = None
    name: str | None = None
    code: str | None = None
    notes: str | None = None
    description: str | None = None


class Assembly(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    project_id: int = Field(
        sa_column=Column(Integer, ForeignKey("project.id", ondelete="CASCADE"))
    )
    rev: str = Field(default="A", max_length=16)
    vault_sha: str | None = Field(
        default=None, sa_column=Column(sqlalchemy.String(64), ForeignKey("blob.sha256"))
    )
    notes: str | None = None
    created_at: datetime = Field(default_factory=datetime.utcnow)


class Part(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    number: str = Field(index=True, unique=True)
    description: str | None = None


class PartCreate(SQLModel):
    number: str
    description: str | None = None


class PartRead(PartCreate):
    id: int
    usage_count: int | None = None


class PartUpdate(SQLModel):
    number: str | None = None
    description: str | None = None


class TestMacro(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    name: str
    glb_path: str | None = None
    notes: str | None = None


class Complex(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    name: str
    eda_path: str | None = None
    notes: str | None = None


class PythonTest(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    name: str
    file_path: str | None = None
    notes: str | None = None


class TestMacroCreate(SQLModel):
    name: str
    glb_path: str | None = None
    notes: str | None = None


class TestMacroRead(TestMacroCreate):
    id: int
    usage_count: int | None = None


class TestMacroUpdate(SQLModel):
    name: str | None = None
    glb_path: str | None = None
    notes: str | None = None


class ComplexCreate(SQLModel):
    name: str
    eda_path: str | None = None
    notes: str | None = None


class ComplexRead(ComplexCreate):
    id: int


class ComplexUpdate(SQLModel):
    name: str | None = None
    eda_path: str | None = None
    notes: str | None = None


class PythonTestCreate(SQLModel):
    name: str
    file_path: str | None = None
    notes: str | None = None


class PythonTestRead(PythonTestCreate):
    id: int


class PythonTestUpdate(SQLModel):
    name: str | None = None
    file_path: str | None = None
    notes: str | None = None


class MacroLink(SQLModel):
    test_macro_id: int


class PartTestMap(SQLModel, table=True):
    part_id: int = Field(foreign_key="part.id", primary_key=True)
    test_macro_id: int = Field(foreign_key="testmacro.id", primary_key=True)


class Blob(SQLModel, table=True):
    sha256: str = Field(sa_column=Column(sqlalchemy.String(64), primary_key=True))
    size: int | None = None


class Quote(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    assembly_id: int = Field(foreign_key="assembly.id")
    total_cost: float | None = None


class AuditEvent(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    table_name: str
    row_id: int
    diff: str
    created_at: datetime = Field(default_factory=datetime.utcnow)


class User(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    username: str = Field(sa_column_kwargs={"unique": True})
    hashed_pw: str
    role: str


class UserCreate(SQLModel):
    username: str
    password: str
    role: str = "user"


class QuoteResponse(SQLModel):
    """Schema returned from the quote endpoint."""

    total_components: int
    estimated_time_s: int
    estimated_cost_usd: float
    labor_cost: float
    parts_cost: float
    total_cost: float
    currency: str


class Inventory(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    mpn: str = Field(sa_column_kwargs={"unique": True})
    on_hand: int = 0
    on_order: int = 0


class InventoryCreate(SQLModel):
    mpn: str
    on_hand: int = 0
    on_order: int = 0


class InventoryRead(InventoryCreate):
    id: int


class InventoryUpdate(SQLModel):
    on_hand: int | None = None
    on_order: int | None = None



class TestResult(SQLModel, table=True):
    """Database model for flying-probe test results."""

    test_id: int | None = Field(default=None, primary_key=True)
    assembly_id: int = 1  # placeholder until multi-BOM support
    serial_number: str | None = None
    date_tested: datetime = Field(default_factory=datetime.utcnow)
    result: bool
    failure_details: str | None = None


class TestResultCreate(SQLModel):
    """Schema for creating test results via the API."""

    assembly_id: int = 1
    serial_number: str | None = None
    result: bool
    failure_details: str | None = None


class TestResultRead(TestResultCreate):
    """Schema returned from the test result API."""

    test_id: int
    date_tested: datetime


class ComponentTraceEntry(SQLModel):
    serial_number: str | None = None
    result: bool
    failure_details: str | None = None


class BoardTraceBOMItem(SQLModel):
    id: int
    part_number: str
    description: str
    quantity: int
    reference: str | None = None
    status: str


class BoardTraceResponse(SQLModel):
    serial_number: str | None = None
    result: bool
    failure_details: str | None = None
    bom: list[BoardTraceBOMItem]


def get_all_bom() -> list[BOMItem]:
    with Session(engine) as session:
        return session.exec(select(BOMItem)).all()


def get_all_testresults() -> list[TestResult]:
    with Session(engine) as session:
        return session.exec(select(TestResult)).all()


def get_default_assembly(project_id: int, session: Session) -> Assembly:
    asm = session.exec(select(Assembly).where(Assembly.project_id == project_id)).first()
    if not asm:
        asm = Assembly(project_id=project_id, rev="A")
        session.add(asm)
        session.commit()
        session.refresh(asm)
    return asm


def get_or_create_part(number: str, description: str | None, session: Session) -> Part:
    """Return existing Part matching number (case-insensitive) or create it."""
    part = session.exec(select(Part).where(Part.number.ilike(number))).first()
    if part:
        return part
    part = Part(number=number, description=description)
    session.add(part)
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        part = session.exec(select(Part).where(Part.number.ilike(number))).first()
        if not part:
            raise
    else:
        session.refresh(part)
    return part


def csv_generator(items: list[BOMItem], delimiter: str = ","):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter)
    writer.writerow(["id", "part_number", "description", "quantity", "reference"])
    yield output.getvalue()
    output.seek(0)
    output.truncate(0)
    for item in items:
        writer.writerow([
            item.id,
            item.part_number,
            item.description,
            item.quantity,
            item.reference or "",
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)


def excel_bytes(results: list[TestResult]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(
        [
            "test_id",
            "assembly_id",
            "serial_number",
            "date_tested",
            "result",
            "failure_details",
        ]
    )
    for r in results:
        ws.append(
            [
                r.test_id,
                r.assembly_id,
                r.serial_number,
                r.date_tested.isoformat(),
                r.result,
                r.failure_details,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def nightly_backup(dest: str = "backups") -> None:
    os.makedirs(dest, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d")
    items = get_all_bom()
    with open(os.path.join(dest, f"bom_{stamp}.csv"), "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "part_number", "description", "quantity", "reference"])
        for i in items:
            writer.writerow([i.id, i.part_number, i.description, i.quantity, i.reference or ""])
    results = get_all_testresults()
    excel_data = excel_bytes(results)
    with open(os.path.join(dest, f"testresults_{stamp}.xlsx"), "wb") as f:
        f.write(excel_data)


def save_blob(data: bytes, ext: str) -> str:
    """Store bytes on disk and record/update Blob row."""
    sha = hashlib.sha256(data).hexdigest()
    prefix = sha[:2]
    os.makedirs(os.path.join("assets", prefix), exist_ok=True)
    filename = f"{sha}.{ext.lstrip('.')}"
    path = os.path.join("assets", prefix, filename)
    if not os.path.exists(path):
        with open(path, "wb") as f:
            f.write(data)
    with Session(engine) as session:
        blob = session.get(Blob, sha)
        if not blob:
            blob = Blob(sha256=sha, size=len(data))
            session.add(blob)
        else:
            blob.size = len(data)
            session.add(blob)
        session.commit()
    return f"assets/{prefix}/{filename}"


def _count_refs(sha: str) -> int:
    with Session(engine) as session:
        t = session.exec(
            select(sqlalchemy.func.count()).select_from(TestMacro).where(TestMacro.glb_path.like(f"%{sha}%"))
        ).one()
        c = session.exec(
            select(sqlalchemy.func.count()).select_from(Complex).where(Complex.eda_path.like(f"%{sha}%"))
        ).one()
        p = session.exec(
            select(sqlalchemy.func.count()).select_from(PythonTest).where(PythonTest.file_path.like(f"%{sha}%"))
        ).one()
    return t + c + p


def cleanup_blob(path: str) -> None:
    if not path:
        return
    m = re.search(r"([0-9a-f]{64})", path)
    if not m:
        return
    sha = m.group(1)
    if _count_refs(sha) == 0:
        p = Path(path)
        if p.exists():
            try:
                p.unlink()
            except Exception:
                pass


def migrate_db() -> None:
    """Apply simple in-place migrations for older database schemas."""
    SQLModel.metadata.create_all(engine)
    inspector = sqlalchemy.inspect(engine)
    if "bomitem" in inspector.get_table_names():
        columns = {c["name"] for c in inspector.get_columns("bomitem")}
        if "assembly_id" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN assembly_id INTEGER"))
        if "part_id" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN part_id INTEGER"))
        if "datasheet_url" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN datasheet_url TEXT"))
        if "manufacturer" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN manufacturer TEXT"))
        if "mpn" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN mpn TEXT"))
        if "footprint" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN footprint TEXT"))
        if "unit_cost" not in columns:
            with engine.begin() as conn:
                if engine.dialect.name == "sqlite":
                    conn.execute(text("ALTER TABLE bomitem ADD COLUMN unit_cost NUMERIC DEFAULT 0"))
                else:
                    conn.execute(text("ALTER TABLE bomitem ADD COLUMN IF NOT EXISTS unit_cost NUMERIC(10,4) DEFAULT 0"))
        if "dnp" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN dnp BOOLEAN DEFAULT 0"))
        if "currency" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bomitem ADD COLUMN currency VARCHAR(3) DEFAULT 'USD'"))
        with engine.begin() as conn:
            pass


def init_db() -> None:
    """Create database tables if they do not exist."""

    if engine.dialect.name == "sqlite":
        sqlalchemy.event.listen(
            engine, "connect", lambda conn, rec: conn.execute("PRAGMA foreign_keys=ON")
        )
        with engine.connect() as conn:
            conn.exec_driver_sql("PRAGMA foreign_keys=ON")
    SQLModel.metadata.create_all(engine)
    migrate_db()
    with Session(engine) as session:
        user_exists = session.exec(select(User)).first()
        if not user_exists:
            admin = User(
                username="admin",
                hashed_pw=get_password_hash("change_me"),
                role="admin",
            )
            session.add(admin)
            session.commit()

os.makedirs("datasheets", exist_ok=True)
os.makedirs("assets", exist_ok=True)

app = FastAPI()
app.mount("/datasheets", StaticFiles(directory="datasheets"), name="datasheets")
ui_router = APIRouter(prefix="/ui")
workflow_router = APIRouter(prefix="/ui/workflow")

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/token", auto_error=False)


def get_current_user(token: str = Depends(oauth2_scheme)) -> User:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    with Session(engine) as session:
        user = session.exec(select(User).where(User.username == username)).first()
        if not user:
            raise credentials_exception
        return user


def admin_required(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin only")
    return current_user


def edit_allowed(current_user: User = Depends(get_current_user)) -> User:
    """Deny write operations for operator role."""
    if current_user.role == "operator":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Read only")
    return current_user


def optional_user(token: str | None = Depends(oauth2_scheme)) -> User | None:
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
    except Exception:
        return None
    with Session(engine) as session:
        return session.exec(select(User).where(User.username == username)).first()

origins = ["http://localhost:3000"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup() -> None:
    """Initialise the database on application start."""

    init_db()
    if not scheduler.get_jobs():
        scheduler.add_job(nightly_backup, "cron", hour=2, minute=0)
    if not scheduler.running:
        scheduler.start()


@app.on_event("shutdown")
def on_shutdown() -> None:
    if scheduler.running:
        scheduler.shutdown()


@app.get("/health")
def health() -> dict[str, str]:
    """Return API and DB health status."""

    try:
        with Session(engine) as session:
            session.exec(text("SELECT 1"))
        db_status = "ok"
    except Exception:  # pragma: no cover - network/db errors
        db_status = "error"

    return {"api": "ok", "db": db_status}


@app.post("/auth/token")
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()) -> dict:
    with Session(engine) as session:
        user = session.exec(select(User).where(User.username == form_data.username)).first()
        if not user or not verify_password(form_data.password, user.hashed_pw):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Incorrect username or password")
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    token = create_access_token({"sub": user.username, "role": user.role}, access_token_expires)
    return {"access_token": token, "token_type": "bearer"}


@app.get("/auth/me")
def auth_me(current_user: User = Depends(get_current_user)) -> dict:
    """Return currently authenticated user."""
    return {"username": current_user.username, "role": current_user.role}


@app.post("/auth/register", dependencies=[Depends(admin_required)], status_code=status.HTTP_201_CREATED)
def register_user(user_in: UserCreate) -> dict:
    with Session(engine) as session:
        user = User(username=user_in.username, hashed_pw=get_password_hash(user_in.password), role=user_in.role)
        session.add(user)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Username already exists")
        session.refresh(user)
    return {"id": user.id, "username": user.username, "role": user.role}


@app.get("/customers", response_model=list[CustomerRead])
def list_customers(
    search: str | None = None,
    skip: int = 0,
    limit: int = 50,
) -> list[CustomerRead]:
    if limit > 200:
        limit = 200
    with Session(engine) as session:
        stmt = select(Customer)
        if search:
            stmt = stmt.where(Customer.name.ilike(f"%{search}%"))
        stmt = stmt.offset(skip).limit(limit)
        return session.exec(stmt).all()


@app.post("/customers", response_model=CustomerRead, status_code=status.HTTP_201_CREATED)
def create_customer(customer: CustomerCreate) -> CustomerRead:
    with Session(engine) as session:
        db = Customer.from_orm(customer)
        session.add(db)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            raise HTTPException(status_code=409, detail="Customer exists")
        session.refresh(db)
        return db


@app.patch("/customers/{customer_id}", response_model=CustomerRead)
def update_customer(customer_id: int, customer: CustomerUpdate) -> CustomerRead:
    with Session(engine) as session:
        db_cust = session.get(Customer, customer_id)
        if not db_cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        for f, v in customer.model_dump(exclude_unset=True).items():
            setattr(db_cust, f, v)
        session.add(db_cust)
        session.commit()
        session.refresh(db_cust)
        return db_cust


@app.delete("/customers/{customer_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def delete_customer(customer_id: int) -> Response:
    with Session(engine) as session:
        cust = session.get(Customer, customer_id)
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        session.delete(cust)
        session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/customers/{customer_id}/projects", response_model=list[ProjectRead])
def customer_projects(customer_id: int) -> list[ProjectRead]:
    with Session(engine) as session:
        return session.exec(select(Project).where(Project.customer_id == customer_id)).all()


@app.get("/parts", response_model=list[PartRead])
def list_parts(
    search: str | None = None,
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
) -> list[PartRead]:
    if limit > 200:
        limit = 200
    with Session(engine) as session:
        stmt = select(Part)
        if search:
            stmt = stmt.where(Part.number.ilike(f"%{search}%"))
        stmt = stmt.offset(skip).limit(limit)
        parts = session.exec(stmt).all()
        result: list[PartRead] = []
        for p in parts:
            count = session.exec(
                select(sqlalchemy.func.count()).select_from(BOMItem).where(BOMItem.part_id == p.id)
            ).one()
            result.append(PartRead(id=p.id, number=p.number, description=p.description, usage_count=count))
        return result


@app.post("/parts", response_model=PartRead, status_code=status.HTTP_201_CREATED)
def create_part(part_in: PartCreate, current_user: User = Depends(edit_allowed)) -> PartRead:
    with Session(engine) as session:
        existing = session.exec(select(Part).where(Part.number.ilike(part_in.number))).first()
        if existing:
            raise HTTPException(status_code=409, detail="Part exists")
        db_part = Part.from_orm(part_in)
        session.add(db_part)
        session.commit()
        session.refresh(db_part)
        if re.match(r"(?i)^U\d+", db_part.number):
            default = session.exec(select(TestMacro).where(TestMacro.name == "Continuity")).first()
            if not default:
                default = TestMacro(name="Continuity")
                session.add(default)
                session.commit()
                session.refresh(default)
            link = PartTestMap(part_id=db_part.id, test_macro_id=default.id)
            session.add(link)
            try:
                session.commit()
            except IntegrityError:
                session.rollback()
        return PartRead(id=db_part.id, number=db_part.number, description=db_part.description, usage_count=0)


@app.patch("/parts/{part_id}", response_model=PartRead)
def update_part(part_id: int, part_in: PartUpdate, current_user: User = Depends(edit_allowed)) -> PartRead:
    with Session(engine) as session:
        db_part = session.get(Part, part_id)
        if not db_part:
            raise HTTPException(status_code=404, detail="Part not found")
        if part_in.number is not None:
            existing = session.exec(
                select(Part).where(Part.number.ilike(part_in.number), Part.id != part_id)
            ).first()
            if existing:
                raise HTTPException(status_code=409, detail="Part exists")
        for f, v in part_in.dict(exclude_unset=True).items():
            setattr(db_part, f, v)
        session.add(db_part)
        session.commit()
        session.refresh(db_part)
        count = session.exec(
            select(sqlalchemy.func.count()).select_from(BOMItem).where(BOMItem.part_id == db_part.id)
        ).one()
        return PartRead(id=db_part.id, number=db_part.number, description=db_part.description, usage_count=count)


@app.delete(
    "/parts/{part_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
    dependencies=[Depends(admin_required)],
)
def delete_part(part_id: int) -> Response:
    with Session(engine) as session:
        db_part = session.get(Part, part_id)
        if not db_part:
            raise HTTPException(status_code=404, detail="Part not found")
        session.delete(db_part)
        session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/testmacros", response_model=list[TestMacroRead])
def list_testmacros(
    search: str | None = None,
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
) -> list[TestMacroRead]:
    if limit > 200:
        limit = 200
    with Session(engine) as session:
        stmt = select(TestMacro)
        if search:
            stmt = stmt.where(TestMacro.name.ilike(f"%{search}%"))
        stmt = stmt.offset(skip).limit(limit)
        macros = session.exec(stmt).all()
        result: list[TestMacroRead] = []
        for m in macros:
            count = session.exec(
                select(sqlalchemy.func.count()).select_from(PartTestMap).where(PartTestMap.test_macro_id == m.id)
            ).one()
            result.append(
                TestMacroRead(id=m.id, name=m.name, glb_path=m.glb_path, notes=m.notes, usage_count=count)
            )
        return result


@app.post("/testmacros", response_model=TestMacroRead, status_code=status.HTTP_201_CREATED)
def create_testmacro(macro: TestMacroCreate, current_user: User = Depends(edit_allowed)) -> TestMacroRead:
    with Session(engine) as session:
        db_macro = TestMacro.from_orm(macro)
        session.add(db_macro)
        session.commit()
        session.refresh(db_macro)
        return TestMacroRead(id=db_macro.id, name=db_macro.name, glb_path=db_macro.glb_path, notes=db_macro.notes, usage_count=0)


@app.patch("/testmacros/{macro_id}", response_model=TestMacroRead)
def update_testmacro(macro_id: int, macro: TestMacroUpdate, current_user: User = Depends(edit_allowed)) -> TestMacroRead:
    with Session(engine) as session:
        db_macro = session.get(TestMacro, macro_id)
        if not db_macro:
            raise HTTPException(status_code=404, detail="Macro not found")
        for f, v in macro.model_dump(exclude_unset=True).items():
            setattr(db_macro, f, v)
        session.add(db_macro)
        session.commit()
        session.refresh(db_macro)
        count = session.exec(
            select(sqlalchemy.func.count()).select_from(PartTestMap).where(PartTestMap.test_macro_id == db_macro.id)
        ).one()
        return TestMacroRead(id=db_macro.id, name=db_macro.name, glb_path=db_macro.glb_path, notes=db_macro.notes, usage_count=count)


@app.post("/testmacros/{macro_id}/upload_glb", response_model=TestMacroRead)
async def upload_glb(
    request: Request,
    macro_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(edit_allowed),
) -> Response:
    contents = await file.read()
    if not file.filename.lower().endswith(".glb"):
        raise HTTPException(status_code=400, detail="Only .glb files allowed")
    if len(contents) > MAX_GLB_MB * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"File exceeds {MAX_GLB_MB} MB")
    path = save_blob(contents, "glb")
    with Session(engine) as session:
        db_macro = session.get(TestMacro, macro_id)
        if not db_macro:
            raise HTTPException(status_code=404, detail="Macro not found")
        db_macro.glb_path = path
        session.add(db_macro)
        session.commit()
        session.refresh(db_macro)
        count = session.exec(
            select(sqlalchemy.func.count()).select_from(PartTestMap).where(PartTestMap.test_macro_id == db_macro.id)
        ).one()
        macro_read = TestMacroRead(
            id=db_macro.id,
            name=db_macro.name,
            glb_path=db_macro.glb_path,
            notes=db_macro.notes,
            usage_count=count,
        )
        if request.headers.get("HX-Request"):
            parts = []
            stmt = (
                select(Part)
                .join(PartTestMap, Part.id == PartTestMap.part_id)
                .where(PartTestMap.test_macro_id == db_macro.id)
            )
            items = session.exec(stmt).all()
            for p in items:
                pc = session.exec(
                    select(sqlalchemy.func.count()).select_from(BOMItem).where(BOMItem.part_id == p.id)
                ).one()
                parts.append(
                    PartRead(id=p.id, number=p.number, description=p.description, usage_count=pc)
                )
            return templates.TemplateResponse(
                "testasset_detail.html",
                {"request": request, "macro": macro_read, "parts": parts},
            )
        return macro_read


@app.delete(
    "/testmacros/{macro_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
    dependencies=[Depends(admin_required)],
)
def delete_testmacro(macro_id: int) -> Response:
    with Session(engine) as session:
        db_macro = session.get(TestMacro, macro_id)
        if not db_macro:
            raise HTTPException(status_code=404, detail="Macro not found")
        path = db_macro.glb_path
        session.delete(db_macro)
        session.commit()
    cleanup_blob(path or "")
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/parts/{part_id}/testmacros", response_model=list[TestMacroRead])
def part_macros(part_id: int, current_user: User = Depends(get_current_user)) -> list[TestMacroRead]:
    with Session(engine) as session:
        part = session.get(Part, part_id)
        if not part:
            raise HTTPException(status_code=404, detail="Part not found")
        stmt = (
            select(TestMacro)
            .join(PartTestMap, TestMacro.id == PartTestMap.test_macro_id)
            .where(PartTestMap.part_id == part_id)
        )
        macros = session.exec(stmt).all()
        result: list[TestMacroRead] = []
        for m in macros:
            count = session.exec(
                select(sqlalchemy.func.count()).select_from(PartTestMap).where(PartTestMap.test_macro_id == m.id)
            ).one()
            result.append(
                TestMacroRead(id=m.id, name=m.name, glb_path=m.glb_path, notes=m.notes, usage_count=count)
            )
        return result


@app.post("/parts/{part_id}/testmacros", response_model=TestMacroRead, status_code=status.HTTP_201_CREATED)
def add_part_macro(part_id: int, link: MacroLink, current_user: User = Depends(edit_allowed)) -> TestMacroRead:
    with Session(engine) as session:
        part = session.get(Part, part_id)
        if not part:
            raise HTTPException(status_code=404, detail="Part not found")
        macro = session.get(TestMacro, link.test_macro_id)
        if not macro:
            raise HTTPException(status_code=404, detail="Macro not found")
        mapping = PartTestMap(part_id=part_id, test_macro_id=macro.id)
        session.add(mapping)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
        count = session.exec(
            select(sqlalchemy.func.count()).select_from(PartTestMap).where(PartTestMap.test_macro_id == macro.id)
        ).one()
        return TestMacroRead(id=macro.id, name=macro.name, glb_path=macro.glb_path, notes=macro.notes, usage_count=count)


@app.delete("/parts/{part_id}/testmacros/{macro_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def remove_part_macro(part_id: int, macro_id: int, current_user: User = Depends(edit_allowed)) -> Response:
    with Session(engine) as session:
        mapping = session.get(PartTestMap, {"part_id": part_id, "test_macro_id": macro_id})
        if mapping:
            session.delete(mapping)
            session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/testmacros/{macro_id}/parts", response_model=list[PartRead])
def macro_parts(macro_id: int, current_user: User = Depends(get_current_user)) -> list[PartRead]:
    """Return Parts linked to a TestMacro."""
    with Session(engine) as session:
        macro = session.get(TestMacro, macro_id)
        if not macro:
            raise HTTPException(status_code=404, detail="Macro not found")
        stmt = (
            select(Part)
            .join(PartTestMap, Part.id == PartTestMap.part_id)
            .where(PartTestMap.test_macro_id == macro_id)
        )
        parts = session.exec(stmt).all()
        result: list[PartRead] = []
        for p in parts:
            count = session.exec(
                select(sqlalchemy.func.count()).select_from(BOMItem).where(BOMItem.part_id == p.id)
            ).one()
            result.append(PartRead(id=p.id, number=p.number, description=p.description, usage_count=count))
        return result


@app.get("/complexes", response_model=list[ComplexRead])
def list_complexes(skip: int = 0, limit: int = 50, current_user: User = Depends(get_current_user)) -> list[ComplexRead]:
    if limit > 200:
        limit = 200
    with Session(engine) as session:
        stmt = select(Complex).offset(skip).limit(limit)
        return session.exec(stmt).all()


@app.post("/complexes", response_model=ComplexRead, status_code=status.HTTP_201_CREATED)
def create_complex(obj: ComplexCreate, current_user: User = Depends(edit_allowed)) -> ComplexRead:
    with Session(engine) as session:
        db = Complex.from_orm(obj)
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.patch("/complexes/{cid}", response_model=ComplexRead)
def update_complex(cid: int, obj: ComplexUpdate, current_user: User = Depends(edit_allowed)) -> ComplexRead:
    with Session(engine) as session:
        db = session.get(Complex, cid)
        if not db:
            raise HTTPException(status_code=404, detail="Complex not found")
        for f, v in obj.model_dump(exclude_unset=True).items():
            setattr(db, f, v)
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.post("/complexes/{cid}/upload_eda", response_model=ComplexRead)
async def upload_eda(
    cid: int,
    file: UploadFile = File(...),
    current_user: User = Depends(edit_allowed),
) -> ComplexRead:
    contents = await file.read()
    if not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Only .zip files allowed")
    if len(contents) > MAX_EDA_MB * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"File exceeds {MAX_EDA_MB} MB")
    path = save_blob(contents, "zip")
    with Session(engine) as session:
        db = session.get(Complex, cid)
        if not db:
            raise HTTPException(status_code=404, detail="Complex not found")
        db.eda_path = path
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.delete("/complexes/{cid}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def delete_complex(cid: int, current_user: User = Depends(admin_required)) -> Response:
    with Session(engine) as session:
        db = session.get(Complex, cid)
        if not db:
            raise HTTPException(status_code=404, detail="Complex not found")
        path = db.eda_path
        session.delete(db)
        session.commit()
    cleanup_blob(path or "")
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/pythontests", response_model=list[PythonTestRead])
def list_pythontests(skip: int = 0, limit: int = 50, current_user: User = Depends(get_current_user)) -> list[PythonTestRead]:
    if limit > 200:
        limit = 200
    with Session(engine) as session:
        stmt = select(PythonTest).offset(skip).limit(limit)
        return session.exec(stmt).all()


@app.post("/pythontests", response_model=PythonTestRead, status_code=status.HTTP_201_CREATED)
def create_pythontest(obj: PythonTestCreate, current_user: User = Depends(edit_allowed)) -> PythonTestRead:
    with Session(engine) as session:
        db = PythonTest.from_orm(obj)
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.patch("/pythontests/{pid}", response_model=PythonTestRead)
def update_pythontest(pid: int, obj: PythonTestUpdate, current_user: User = Depends(edit_allowed)) -> PythonTestRead:
    with Session(engine) as session:
        db = session.get(PythonTest, pid)
        if not db:
            raise HTTPException(status_code=404, detail="PythonTest not found")
        for f, v in obj.model_dump(exclude_unset=True).items():
            setattr(db, f, v)
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.post("/pythontests/{pid}/upload_file", response_model=PythonTestRead)
async def upload_pythontest_file(
    pid: int,
    file: UploadFile = File(...),
    current_user: User = Depends(edit_allowed),
) -> PythonTestRead:
    contents = await file.read()
    if not file.filename.lower().endswith(".py"):
        raise HTTPException(status_code=400, detail="Only .py files allowed")
    if len(contents) > MAX_PY_MB * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"File exceeds {MAX_PY_MB} MB")
    path = save_blob(contents, "py")
    with Session(engine) as session:
        db = session.get(PythonTest, pid)
        if not db:
            raise HTTPException(status_code=404, detail="PythonTest not found")
        db.file_path = path
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.delete("/pythontests/{pid}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def delete_pythontest(pid: int, current_user: User = Depends(admin_required)) -> Response:
    with Session(engine) as session:
        db = session.get(PythonTest, pid)
        if not db:
            raise HTTPException(status_code=404, detail="PythonTest not found")
        path = db.file_path
        session.delete(db)
        session.commit()
    cleanup_blob(path or "")
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/projects/{project_id}/assemblies", response_model=list[Assembly])
def list_project_assemblies(project_id: int) -> list[Assembly]:
    with Session(engine) as session:
        return session.exec(select(Assembly).where(Assembly.project_id == project_id)).all()


@app.get("/projects", response_model=list[ProjectRead])
def list_projects(
    customer_id: int | None = None,
    skip: int = 0,
    limit: int = 50,
) -> list[ProjectRead]:
    if limit > 200:
        limit = 200
    with Session(engine) as session:
        stmt = select(Project)
        if customer_id is not None:
            stmt = stmt.where(Project.customer_id == customer_id)
        stmt = stmt.offset(skip).limit(limit)
        return session.exec(stmt).all()


@app.post("/projects", response_model=ProjectRead, status_code=status.HTTP_201_CREATED)
def create_project(project: ProjectCreate) -> ProjectRead:
    with Session(engine) as session:
        db_proj = Project.from_orm(project)
        session.add(db_proj)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            raise HTTPException(status_code=409, detail="Project exists")
        session.refresh(db_proj)
        asm = Assembly(project_id=db_proj.id, rev="A")
        session.add(asm)
        session.commit()
        session.refresh(db_proj)
        return db_proj


@app.patch("/projects/{project_id}", response_model=ProjectRead)
def update_project(project_id: int, project: ProjectUpdate) -> ProjectRead:
    with Session(engine) as session:
        db_proj = session.get(Project, project_id)
        if not db_proj:
            raise HTTPException(status_code=404, detail="Project not found")
        for f, v in project.model_dump(exclude_unset=True).items():
            setattr(db_proj, f, v)
        session.add(db_proj)
        session.commit()
        session.refresh(db_proj)
        return db_proj


@app.delete("/projects/{project_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def delete_project(project_id: int) -> Response:
    with Session(engine) as session:
        db_proj = session.get(Project, project_id)
        if not db_proj:
            raise HTTPException(status_code=404, detail="Project not found")
        session.delete(db_proj)
        session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/projects/{project_id}/bom", response_model=list[BOMItemRead])
def project_bom(
    project_id: int,
    search: str | None = None,
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
) -> list[BOMItemRead]:
    """List BOM items for a specific project."""

    if limit > 200:
        limit = 200
    with Session(engine) as session:
        asm = get_default_assembly(project_id, session)
        stmt = select(BOMItem).where(BOMItem.assembly_id == asm.id)
        if search:
            pattern = f"%{search}%"
            stmt = stmt.where(
                (BOMItem.part_number.ilike(pattern))
                | (BOMItem.description.ilike(pattern))
            )
        stmt = stmt.offset(skip).limit(limit)
        items = session.exec(stmt).all()
        return items


@app.get("/projects/{project_id}/export.csv")
def project_export_csv(project_id: int, comma: bool = True):
    """Export a project's BOM as CSV."""
    with Session(engine) as session:
        asm = get_default_assembly(project_id, session)
        items = session.exec(
            select(BOMItem).where(BOMItem.assembly_id == asm.id)
        ).all()
    delim = "," if comma else ";"
    headers = {"Content-Disposition": "attachment; filename=export.csv"}
    return StreamingResponse(
        csv_generator(items, delimiter=delim),
        media_type="text/csv",
        headers=headers,
    )


@app.get("/projects/{project_id}/cost")
def project_cost(
    project_id: int,
    currency: str | None = None,
    current_user: User = Depends(get_current_user),
) -> dict:
    """Return estimated total cost for a project's BOM."""
    cur = (currency or BOM_DEFAULT_CURRENCY).upper()
    with Session(engine) as session:
        asm = get_default_assembly(project_id, session)
        items = session.exec(
            select(BOMItem).where(
                BOMItem.assembly_id == asm.id, BOMItem.dnp == False
            )
        ).all()
    rate_to = fx.get(cur)
    parts = 0.0
    for item in items:
        if item.unit_cost and float(item.unit_cost) > 0:
            rate_from = fx.get(item.currency or BOM_DEFAULT_CURRENCY)
            parts += float(item.unit_cost) * item.quantity * rate_to / rate_from
    labor = calculate_quote(items)["estimated_cost_usd"] * rate_to
    total = parts + labor
    return {
        "parts_cost": round(parts, 2),
        "labor_cost": round(labor, 2),
        "total_cost": round(total, 2),
        "currency": cur,
    }


@app.get("/projects/{project_id}/quote", response_model=QuoteResponse)
def project_quote(
    project_id: int,
    currency: str | None = None,
    current_user: User = Depends(get_current_user),
) -> QuoteResponse:
    """Return cost/time estimate for a single project's BOM."""
    cur = (currency or BOM_DEFAULT_CURRENCY).upper()
    with Session(engine) as session:
        asm = get_default_assembly(project_id, session)
        items = session.exec(
            select(BOMItem).where(
                BOMItem.assembly_id == asm.id, BOMItem.dnp == False
            )
        ).all()
    data = calculate_quote(items)
    rate_to = fx.get(cur)
    parts = 0.0
    for i in items:
        if i.unit_cost and float(i.unit_cost) > 0:
            rate_from = fx.get(i.currency or BOM_DEFAULT_CURRENCY)
            parts += float(i.unit_cost) * i.quantity * rate_to / rate_from
    labor = data["estimated_cost_usd"] * rate_to
    data["parts_cost"] = round(parts, 2)
    data["labor_cost"] = round(labor, 2)
    data["total_cost"] = round(parts + labor, 2)
    data["currency"] = cur
    return QuoteResponse(**data)


@app.post("/projects/{project_id}/po.pdf")
def project_po_pdf(project_id: int, current_user: User = Depends(edit_allowed)):
    """Generate a simple purchase order PDF and update inventory."""
    with Session(engine) as session:
        proj = session.get(Project, project_id)
        if not proj:
            raise HTTPException(status_code=404, detail="Project not found")
        cust = session.get(Customer, proj.customer_id)
        asm = get_default_assembly(project_id, session)
        items = session.exec(
            select(BOMItem).where(
                BOMItem.assembly_id == asm.id, BOMItem.dnp == False
            )
        ).all()
        groups: dict[tuple[str, str, str], dict[str, float]] = {}
        for it in items:
            key = (
                it.manufacturer or "",
                it.mpn or it.part_number,
                (it.currency or BOM_DEFAULT_CURRENCY).upper(),
            )
            g = groups.setdefault(key, {"qty": 0, "cost": 0.0})
            g["qty"] += it.quantity
            g["cost"] += float(it.unit_cost or 0) * it.quantity
        parts_sub = 0.0
        rate_to = fx.get(BOM_DEFAULT_CURRENCY)
        for (mfr, mpn, cur), g in groups.items():
            parts_sub += g["cost"] * rate_to / fx.get(cur)
            inv = session.exec(select(Inventory).where(Inventory.mpn == mpn)).first()
            if not inv:
                inv = Inventory(mpn=mpn, on_hand=0, on_order=0)
            inv.on_hand -= g["qty"]
            inv.on_order += g["qty"]
            session.add(inv)
        labor = calculate_quote(items)["estimated_cost_usd"] * rate_to
        total = parts_sub + labor
        session.commit()
        from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Image, Spacer
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter)
        styles = getSampleStyleSheet()
        elems = []
        logo = Path("docs/logo.png")
        if logo.exists():
            elems.append(Image(str(logo), width=100))
        elems.append(Paragraph(f"{cust.name} - {proj.name} - {datetime.utcnow().date()}", styles["Heading2"]))
        rows = [["Manufacturer", "MPN", "Qty", "Currency", "Ext. Cost"]]
        for (mfr, mpn, cur), g in groups.items():
            rows.append([mfr, mpn, str(g["qty"]), cur, f"{g['cost']:.2f}"])
        elems.append(Table(rows))
        elems.append(Spacer(1, 12))
        elems.append(Paragraph(f"Parts subtotal: {parts_sub:.2f} {BOM_DEFAULT_CURRENCY}", styles["Normal"]))
        elems.append(Paragraph(f"Labor subtotal: {labor:.2f} {BOM_DEFAULT_CURRENCY}", styles["Normal"]))
        elems.append(Paragraph(f"Grand total: {total:.2f} {BOM_DEFAULT_CURRENCY}", styles["Normal"]))
        doc.build(elems)
        pdf = buf.getvalue()
    headers = {"Content-Disposition": "attachment; filename=po.pdf"}
    return Response(content=pdf, media_type="application/pdf", headers=headers)


@app.get("/bom/items", response_model=list[BOMItemRead])
def list_items(
    search: str | None = None,
    min_qty: int | None = None,
    max_qty: int | None = None,
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
) -> list[BOMItemRead]:
    """Return BOM items with optional filtering and pagination."""

    if limit > 200:
        limit = 200

    stmt = select(BOMItem)
    if search:
        pattern = f"%{search}%"
        stmt = stmt.where(
            (BOMItem.part_number.ilike(pattern))
            | (BOMItem.description.ilike(pattern))
        )
    if min_qty is not None:
        stmt = stmt.where(BOMItem.quantity >= min_qty)
    if max_qty is not None:
        stmt = stmt.where(BOMItem.quantity <= max_qty)
    stmt = stmt.offset(skip).limit(limit)

    with Session(engine) as session:
        items = session.exec(stmt).all()
    return items


@app.post("/bom/items", response_model=BOMItemRead, status_code=status.HTTP_201_CREATED)
def create_item(item: BOMItemCreate, current_user: User = Depends(edit_allowed)) -> BOMItemRead:
    """Create a new BOM item."""

    db_item = BOMItem.from_orm(item)
    with Session(engine) as session:
        if db_item.assembly_id is None and hasattr(item, "project_id") and item.project_id is not None:
            asm = get_default_assembly(item.project_id, session)
            db_item.assembly_id = asm.id
        part = get_or_create_part(db_item.part_number, db_item.description, session)
        db_item.part_id = part.id
        session.add(db_item)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Item with this part_number and reference already exists",
            )
        session.refresh(db_item)
    return db_item


@app.get("/bom/items/{item_id}", response_model=BOMItemRead)
def get_item(item_id: int) -> BOMItemRead:
    """Retrieve a single BOM item by ID."""

    with Session(engine) as session:
        item = session.get(BOMItem, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        return item


@app.put("/bom/items/{item_id}", response_model=BOMItemRead)
def replace_item(
    item_id: int,
    item_in: BOMItemCreate,
    current_user: User = Depends(edit_allowed),
) -> BOMItemRead:
    """Fully replace an existing BOM item."""

    with Session(engine) as session:
        db_item = session.get(BOMItem, item_id)
        if not db_item:
            raise HTTPException(status_code=404, detail="Item not found")
        for field, value in item_in.dict().items():
            setattr(db_item, field, value)
        part = get_or_create_part(db_item.part_number, db_item.description, session)
        db_item.part_id = part.id
        try:
            session.add(db_item)
            session.commit()
        except IntegrityError:
            session.rollback()
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Item with this part_number and reference already exists",
            )
        session.refresh(db_item)
        return db_item


@app.get("/assets/{sha}/{fname}")
def download_asset(sha: str, fname: str):
    path = Path("assets") / sha[:2] / fname
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    media_type = "application/octet-stream"
    headers = {"Content-Disposition": f"attachment; filename={fname}"}
    return StreamingResponse(open(path, "rb"), headers=headers, media_type=media_type)


@app.patch("/bom/items/{item_id}", response_model=BOMItemRead)
def update_item(
    item_id: int,
    item_in: BOMItemUpdate,
    current_user: User = Depends(edit_allowed),
) -> BOMItemRead:
    """Partially update an existing BOM item."""

    with Session(engine) as session:
        db_item = session.get(BOMItem, item_id)
        if not db_item:
            raise HTTPException(status_code=404, detail="Item not found")
        update_data = item_in.dict(exclude_unset=True)
        if "project_id" in update_data:
            asm = get_default_assembly(update_data.pop("project_id"), session)
            update_data["assembly_id"] = asm.id
        new_part_number = update_data.get("part_number")
        if new_part_number is not None:
            part = get_or_create_part(new_part_number, update_data.get("description", db_item.description), session)
            db_item.part_id = part.id
        for field, value in update_data.items():
            setattr(db_item, field, value)
        try:
            session.add(db_item)
            session.commit()
        except IntegrityError:
            session.rollback()
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Item with this part_number and reference already exists",
            )
        session.refresh(db_item)
        return db_item


@app.delete("/bom/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def delete_item(item_id: int, current_user: User = Depends(admin_required)) -> Response:
    """Delete a BOM item."""

    with Session(engine) as session:
        db_item = session.get(BOMItem, item_id)
        if not db_item:
            raise HTTPException(status_code=404, detail="Item not found")
        session.delete(db_item)
        session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.post("/bom/items/{item_id}/datasheet", response_model=BOMItemRead)
async def upload_datasheet(
    item_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(edit_allowed),
) -> BOMItemRead:
    """Attach a datasheet file to an item and return the updated item."""

    contents = await file.read()
    if len(contents) > MAX_DATASHEET_MB * 1024 * 1024:
        raise HTTPException(
            status_code=413,
            detail=f"File exceeds {MAX_DATASHEET_MB} MB",
        )
    dest = Path("datasheets") / str(item_id)
    dest.mkdir(parents=True, exist_ok=True)
    filename = os.path.basename(file.filename)
    path = dest / filename
    with open(path, "wb") as f:
        f.write(contents)
    url = f"/datasheets/{item_id}/{filename}"
    with Session(engine) as session:
        db_item = session.get(BOMItem, item_id)
        if not db_item:
            raise HTTPException(status_code=404, detail="Item not found")
        db_item.datasheet_url = url
        session.add(db_item)
        session.commit()
        session.refresh(db_item)
        return db_item


@app.post("/bom/items/{item_id}/fetch_price", response_model=BOMItemRead)
def fetch_price(
    item_id: int,
    payload: dict,
    current_user: User = Depends(get_current_user),
) -> BOMItemRead:
    """Fetch vendor price and update unit_cost."""
    source = payload.get("source")
    if source != "octopart":
        raise HTTPException(status_code=400, detail="Unsupported source")
    with Session(engine) as session:
        item = session.get(BOMItem, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        mpn = item.mpn or item.part_number
        try:
            data = octopart.lookup(mpn)
        except KeyError:
            raise HTTPException(status_code=413, detail="MPN not found")
        item.unit_cost = data["price"]
        session.add(item)
        session.commit()
        session.refresh(item)
        return item


@app.post("/bom/import", response_model=list[BOMItemRead])
async def import_bom(
    file: UploadFile = File(...),
    assembly_id: int | None = None,
    current_user: User = Depends(edit_allowed),
) -> list[BOMItemRead]:
    """Import BOM items from an uploaded PDF or Excel file."""

    contents = await file.read()
    ext = file.filename.lower().split(".")[-1]
    records: list[dict]
    if ext == "pdf":
        text = extract_bom_text(contents)
        records = parse_bom_lines(text)
    elif ext in {"xlsx", "xls"}:
        if ext == "xls":
            import xlrd

            wb = xlrd.open_workbook(file_contents=contents)
            sheet = wb.sheet_by_index(0)
            headers = [str(sheet.cell_value(0, i)).lower() for i in range(sheet.ncols)]
            records = []
            for row_idx in range(1, sheet.nrows):
                row = sheet.row_values(row_idx)
                data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                records.append(
                    {
                        "part_number": data.get("part_number") or data.get("part number") or "",
                        "description": data.get("description") or data.get("desc") or "",
                        "quantity": int(data.get("quantity") or data.get("qty") or 1),
                        "reference": data.get("reference") or data.get("ref") or None,
                        "manufacturer": data.get("manufacturer"),
                        "mpn": data.get("mpn"),
                        "footprint": data.get("footprint"),
                        "unit_cost": float(data.get("unit_cost")) if data.get("unit_cost") else None,
                    }
                )
        else:
            wb = load_workbook(io.BytesIO(contents), read_only=True)
            ws = wb.active
            headers = [str(c.value).lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                records.append(
                    {
                        "part_number": data.get("part_number") or data.get("part number") or "",
                        "description": data.get("description") or data.get("desc") or "",
                        "quantity": int(data.get("quantity") or data.get("qty") or 1),
                        "reference": data.get("reference") or data.get("ref") or None,
                        "manufacturer": data.get("manufacturer"),
                        "mpn": data.get("mpn"),
                        "footprint": data.get("footprint"),
                        "unit_cost": float(data.get("unit_cost")) if data.get("unit_cost") else None,
                    }
                )
    elif ext == "csv":
        text = contents.decode("utf-8", errors="ignore")
        lines = text.splitlines()
        delimiter = ";" if lines and lines[0].count(";") > lines[0].count(",") else ","
        reader = csv.DictReader(lines, delimiter=delimiter)
        records = []
        for row in reader:
            records.append(
                {
                    "part_number": row.get("part_number") or row.get("part number") or "",
                    "description": row.get("description") or row.get("desc") or "",
                    "quantity": int(row.get("quantity") or row.get("qty") or 1),
                    "reference": row.get("reference") or row.get("ref") or None,
                    "manufacturer": row.get("manufacturer"),
                    "mpn": row.get("mpn"),
                    "footprint": row.get("footprint"),
                    "unit_cost": float(row.get("unit_cost")) if row.get("unit_cost") else None,
                    "dnp": False,
                    "currency": BOM_DEFAULT_CURRENCY,
                }
            )
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    inserted: list[BOMItem] = []
    with Session(engine) as session:
        asm = None
        if assembly_id is not None:
            asm = session.get(Assembly, assembly_id)
        for rec in records:
            if not rec.get("part_number") or not rec.get("description"):
                continue
            item = BOMItem(**rec)
            if asm:
                item.assembly_id = asm.id
            part = get_or_create_part(item.part_number, item.description, session)
            item.part_id = part.id
            session.add(item)
            try:
                session.commit()
            except IntegrityError:
                session.rollback()
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Item with this part_number and reference already exists",
                )
            session.refresh(item)
            inserted.append(item)
    return inserted


@app.get("/bom/quote", response_model=QuoteResponse)
def get_quote(currency: str | None = None) -> QuoteResponse:
    """Return quick cost/time estimates for all BOM items."""

    cur = (currency or BOM_DEFAULT_CURRENCY).upper()
    with Session(engine) as session:
        items = session.exec(select(BOMItem).where(BOMItem.dnp == False)).all()
    data = calculate_quote(items)
    rate_to = fx.get(cur)
    parts = 0.0
    for i in items:
        if i.unit_cost and float(i.unit_cost) > 0:
            rate_from = fx.get(i.currency or BOM_DEFAULT_CURRENCY)
            parts += float(i.unit_cost) * i.quantity * rate_to / rate_from
    labor = data["estimated_cost_usd"] * rate_to
    data["parts_cost"] = round(parts, 2)
    data["labor_cost"] = round(labor, 2)
    data["total_cost"] = round(parts + labor, 2)
    data["currency"] = cur
    return QuoteResponse(**data)


@app.post("/testresults", response_model=TestResultRead, status_code=status.HTTP_201_CREATED)
def create_test_result(
    result_in: TestResultCreate,
    current_user: User = Depends(edit_allowed),
) -> TestResultRead:
    """Log a new flying-probe test result."""

    db_result = TestResult(**result_in.dict())
    with Session(engine) as session:
        session.add(db_result)
        session.commit()
        session.refresh(db_result)
    return db_result


@app.get("/testresults", response_model=list[TestResultRead])
def list_test_results(skip: int = 0, limit: int = 50) -> list[TestResultRead]:
    """Return test result logs with pagination."""

    if limit > 200:
        limit = 200

    with Session(engine) as session:
        stmt = select(TestResult).offset(skip).limit(limit)
        results = session.exec(stmt).all()
    return results


@app.get("/testresults/{test_id}", response_model=TestResultRead)
def get_test_result(test_id: int) -> TestResultRead:
    """Retrieve a single test result by ID."""

    with Session(engine) as session:
        result = session.get(TestResult, test_id)
        if not result:
            raise HTTPException(status_code=404, detail="Test result not found")
        return result


@app.get("/export/bom.csv", dependencies=[Depends(admin_required)])
def export_bom_csv():
    items = get_all_bom()
    headers = {"Content-Disposition": "attachment; filename=bom.csv"}
    return StreamingResponse(
        csv_generator(items), media_type="text/csv", headers=headers
    )


@app.get(
    "/export/testresults.xlsx",
    dependencies=[Depends(admin_required)],
)
def export_testresults_xlsx():
    results = get_all_testresults()
    data = excel_bytes(results)
    headers = {
        "Content-Disposition": "attachment; filename=testresults.xlsx"
    }
    return StreamingResponse(
        (chunk for chunk in [data]),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )


@app.get("/inventory", response_model=list[InventoryRead])
def list_inventory() -> list[InventoryRead]:
    with Session(engine) as session:
        return session.exec(select(Inventory)).all()


@app.post("/inventory", response_model=InventoryRead, status_code=status.HTTP_201_CREATED)
def create_inventory(inv: InventoryCreate, current_user: User = Depends(edit_allowed)) -> InventoryRead:
    with Session(engine) as session:
        db = Inventory.from_orm(inv)
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.patch("/inventory/{inv_id}", response_model=InventoryRead)
def update_inventory(inv_id: int, inv: InventoryUpdate, current_user: User = Depends(edit_allowed)) -> InventoryRead:
    with Session(engine) as session:
        db = session.get(Inventory, inv_id)
        if not db:
            raise HTTPException(status_code=404, detail="Item not found")
        for f, v in inv.model_dump(exclude_unset=True).items():
            setattr(db, f, v)
        session.add(db)
        session.commit()
        session.refresh(db)
        return db


@app.delete("/inventory/{inv_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def delete_inventory(inv_id: int, current_user: User = Depends(edit_allowed)):
    with Session(engine) as session:
        db = session.get(Inventory, inv_id)
        if not db:
            raise HTTPException(status_code=404, detail="Item not found")
        session.delete(db)
        session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/traceability/component/{part_number}", response_model=list[ComponentTraceEntry])
def trace_component(part_number: str):
    with Session(engine) as session:
        data = component_trace(part_number, session)
    if not data:
        raise HTTPException(status_code=404, detail="No failures found for part")
    return data


@app.get("/traceability/board/{serial_number}", response_model=BoardTraceResponse)
def trace_board(serial_number: str):
    with Session(engine) as session:
        data = board_trace(serial_number, session)
    if not data:
        raise HTTPException(status_code=404, detail="Board not found")
    return data


# ---------------------- Web UI -------------------------

@ui_router.get("/", response_class=HTMLResponse)
def ui_dashboard(request: Request):
    return templates.TemplateResponse(
        "dashboard.html",
        {"request": request, "title": "Dashboard", "db": DATABASE_URL},
    )


@ui_router.get("/bom/", response_class=HTMLResponse)
def ui_bom(request: Request):
    return templates.TemplateResponse("bom.html", {"request": request, "title": "BOM"})


@ui_router.get("/bom/table", response_class=HTMLResponse)
def ui_bom_table(request: Request):
    items = get_all_bom()
    return templates.TemplateResponse(
        "bom_table.html", {"request": request, "items": items}
    )


@ui_router.post("/bom/create", response_class=HTMLResponse)
async def ui_bom_create(request: Request, current_user: User = Depends(edit_allowed)):
    form = await request.form()
    item = BOMItemCreate(
        part_number=form.get("part_number"),
        description=form.get("description"),
        quantity=int(form.get("quantity", 1)),
    )
    with Session(engine) as session:
        db_item = BOMItem.from_orm(item)
        session.add(db_item)
        session.commit()
        session.refresh(db_item)
    return templates.TemplateResponse(
        "bom_table.html",
        {"request": request, "items": [db_item]},
    )


@ui_router.get("/import/", response_class=HTMLResponse)
def ui_import(request: Request):
    return templates.TemplateResponse("import.html", {"request": request, "title": "Import"})


@ui_router.get("/quote/", response_class=HTMLResponse)
def ui_quote(request: Request):
    return templates.TemplateResponse("quote.html", {"request": request, "title": "Quote"})


@ui_router.get("/test/", response_class=HTMLResponse)
def ui_test(request: Request):
    return templates.TemplateResponse("test.html", {"request": request, "title": "Test"})


@ui_router.get("/trace/", response_class=HTMLResponse)
def ui_trace(request: Request):
    return templates.TemplateResponse("trace.html", {"request": request, "title": "Trace"})


@ui_router.get("/export/", response_class=HTMLResponse)
def ui_export(request: Request):
    return templates.TemplateResponse("export.html", {"request": request, "title": "Export"})


@ui_router.get("/inventory/", response_class=HTMLResponse)
def ui_inventory(request: Request):
    with Session(engine) as session:
        items = session.exec(select(Inventory)).all()
    return templates.TemplateResponse(
        "inventory.html", {"request": request, "title": "Inventory", "items": items}
    )


@ui_router.get("/testassets/", response_class=HTMLResponse)
def ui_testassets(request: Request):
    return templates.TemplateResponse("testassets.html", {"request": request, "title": "Test Assets"})


@ui_router.get("/testassets/table", response_class=HTMLResponse)
def ui_testassets_table(request: Request):
    with Session(engine) as session:
        macros = session.exec(select(TestMacro)).all()
        rows: list[TestMacroRead] = []
        for m in macros:
            count = session.exec(
                select(sqlalchemy.func.count()).select_from(PartTestMap).where(PartTestMap.test_macro_id == m.id)
            ).one()
            rows.append(
                TestMacroRead(id=m.id, name=m.name, glb_path=m.glb_path, notes=m.notes, usage_count=count)
            )
    return templates.TemplateResponse("testassets_table.html", {"request": request, "macros": rows})


@ui_router.get("/testassets/detail/{macro_id}", response_class=HTMLResponse)
def ui_testassets_detail(macro_id: int, request: Request):
    with Session(engine) as session:
        macro = session.get(TestMacro, macro_id)
        if not macro:
            raise HTTPException(status_code=404, detail="Macro not found")
        count = session.exec(
            select(sqlalchemy.func.count()).select_from(PartTestMap).where(PartTestMap.test_macro_id == macro.id)
        ).one()
        macro_read = TestMacroRead(
            id=macro.id,
            name=macro.name,
            glb_path=macro.glb_path,
            notes=macro.notes,
            usage_count=count,
        )
        parts = []
        stmt = (
            select(Part)
            .join(PartTestMap, Part.id == PartTestMap.part_id)
            .where(PartTestMap.test_macro_id == macro_id)
        )
        for p in session.exec(stmt).all():
            pc = session.exec(
                select(sqlalchemy.func.count()).select_from(BOMItem).where(BOMItem.part_id == p.id)
            ).one()
            parts.append(PartRead(id=p.id, number=p.number, description=p.description, usage_count=pc))
    return templates.TemplateResponse(
        "testasset_detail.html",
        {"request": request, "macro": macro_read, "parts": parts},
    )


@ui_router.get("/users/", response_class=HTMLResponse)
def ui_users(request: Request):
    return templates.TemplateResponse("users.html", {"request": request, "title": "Users"})


@ui_router.get("/settings/", response_class=HTMLResponse)
def ui_settings(request: Request):
    sqlite_mode = "sqlite" in DATABASE_URL
    return templates.TemplateResponse(
        "settings.html",
        {
            "request": request,
            "title": "Settings",
            "sqlite": sqlite_mode,
            "sqlite_path": DATABASE_URL.removeprefix("sqlite:///") if sqlite_mode else "",
            "pg_url": DATABASE_URL if not sqlite_mode else "",
        },
    )


@ui_router.post("/settings/", response_class=RedirectResponse)
async def ui_save_settings(request: Request):
    form = await request.form()
    mode = form.get("mode")
    if mode == "sqlite":
        path = form.get("sqlite_path") or (BASE_DIR / "bom_dev.db")
        url = f"sqlite:///{path}"
    else:
        url = form.get("pg_url") or DATABASE_URL
    reload_db(url)
    return RedirectResponse("/ui/settings/", status_code=303)


# -------- Workflow Endpoints ---------

@workflow_router.get("/", response_class=HTMLResponse)
def ui_workflow(request: Request, user: User | None = Depends(optional_user)):
    return templates.TemplateResponse(
        "workflow.html",
        {
            "request": request,
            "title": "Workflow",
            "max_ds_mb": MAX_DATASHEET_MB,
            "hourly": BOM_HOURLY_USD,
            "default_currency": BOM_DEFAULT_CURRENCY,
            "hide_po": user.role == "operator" if user else False,
        },
    )


@workflow_router.get("/customers", response_model=list[CustomerRead])
def wf_customers():
    with Session(engine) as session:
        return session.exec(select(Customer).where(Customer.active == True)).all()


@workflow_router.post("/customers", response_model=CustomerRead, status_code=status.HTTP_201_CREATED)
def wf_create_customer(customer: CustomerCreate):
    with Session(engine) as session:
        db_cust = Customer.from_orm(customer)
        session.add(db_cust)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            raise HTTPException(status_code=409, detail="Customer exists")
        session.refresh(db_cust)
        return db_cust


@workflow_router.patch("/customers/{customer_id}", response_model=CustomerRead)
def wf_update_customer(customer_id: int, customer: CustomerUpdate):
    with Session(engine) as session:
        db_cust = session.get(Customer, customer_id)
        if not db_cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        for field, value in customer.model_dump(exclude_unset=True).items():
            setattr(db_cust, field, value)
        session.add(db_cust)
        session.commit()
        session.refresh(db_cust)
        return db_cust


@workflow_router.delete("/customers/{customer_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def wf_delete_customer(customer_id: int):
    with Session(engine) as session:
        db_cust = session.get(Customer, customer_id)
        if not db_cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        session.delete(db_cust)
        session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@workflow_router.get("/projects", response_model=list[ProjectRead])
def wf_projects(customer_id: int):
    with Session(engine) as session:
        return session.exec(select(Project).where(Project.customer_id == customer_id)).all()


@workflow_router.post("/projects", response_model=ProjectRead, status_code=status.HTTP_201_CREATED)
def wf_create_project(project: ProjectCreate):
    with Session(engine) as session:
        db_proj = Project.from_orm(project)
        session.add(db_proj)
        session.commit()
        session.refresh(db_proj)
        asm = Assembly(project_id=db_proj.id, rev="A")
        session.add(asm)
        session.commit()
        session.refresh(db_proj)
        return db_proj


@workflow_router.patch("/projects/{project_id}", response_model=ProjectRead)
def wf_update_project(project_id: int, project: ProjectUpdate):
    with Session(engine) as session:
        db_proj = session.get(Project, project_id)
        if not db_proj:
            raise HTTPException(status_code=404, detail="Project not found")
        for field, value in project.model_dump(exclude_unset=True).items():
            setattr(db_proj, field, value)
        session.add(db_proj)
        session.commit()
        session.refresh(db_proj)
        return db_proj


@workflow_router.delete("/projects/{project_id}", status_code=status.HTTP_204_NO_CONTENT, response_class=Response)
def wf_delete_project(project_id: int):
    with Session(engine) as session:
        db_proj = session.get(Project, project_id)
        if not db_proj:
            raise HTTPException(status_code=404, detail="Project not found")
        session.delete(db_proj)
        session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


class BOMSave(SQLModel):
    assembly_id: int
    items: list[BOMItemCreate]


@workflow_router.post("/upload")
async def wf_upload_bom(file: UploadFile = File(...)):
    contents = await file.read()
    ext = file.filename.lower().split(".")[-1]
    items: list[dict] = []
    if ext == "csv":
        text = contents.decode("utf-8", errors="ignore")
        lines = text.splitlines()
        delim = ";" if lines and lines[0].count(";") > lines[0].count(",") else ","
        reader = csv.DictReader(lines, delimiter=delim)
        for row in reader:
            items.append(
                {
                    "part_number": row.get("part_number") or row.get("part number") or "",
                    "description": row.get("description") or row.get("desc") or "",
                    "quantity": int(row.get("quantity") or row.get("qty") or 1),
                    "reference": row.get("reference") or row.get("ref") or None,
                    "manufacturer": row.get("manufacturer"),
                    "mpn": row.get("mpn"),
                    "footprint": row.get("footprint"),
                    "unit_cost": float(row.get("unit_cost")) if row.get("unit_cost") else None,
                }
            )
    elif ext in {"xlsx", "xls"}:
        wb = load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
        headers = [str(c.value).lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            items.append(
                {
                    "part_number": data.get("part_number") or data.get("part number") or "",
                    "description": data.get("description") or data.get("desc") or "",
                    "quantity": int(data.get("quantity") or data.get("qty") or 1),
                    "reference": data.get("reference") or data.get("ref") or None,
                    "manufacturer": data.get("manufacturer"),
                    "mpn": data.get("mpn"),
                    "footprint": data.get("footprint"),
                    "unit_cost": float(data.get("unit_cost")) if data.get("unit_cost") else None,
                    "dnp": False,
                    "currency": BOM_DEFAULT_CURRENCY,
                }
            )
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    return items


@workflow_router.post("/save", response_model=list[BOMItemRead])
def wf_save_bom(payload: BOMSave):
    inserted: list[BOMItem] = []
    with Session(engine) as session:
        for itm in payload.items:
            db_item = BOMItem.from_orm(itm)
            db_item.assembly_id = payload.assembly_id
            part = get_or_create_part(db_item.part_number, db_item.description, session)
            db_item.part_id = part.id
            session.add(db_item)
            try:
                session.commit()
            except IntegrityError:
                session.rollback()
                continue
            session.refresh(db_item)
            inserted.append(db_item)
    return inserted


app.include_router(ui_router)
app.include_router(workflow_router)

